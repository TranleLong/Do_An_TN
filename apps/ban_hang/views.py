"""Views cho app Bán Hàng"""
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Q, Sum
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from apps.kho.models import TonKhoViTri
from apps.danh_muc.models import (HangHoa, KhachHang, Kho, NhomHang,
                                  TaiKhoanKeToan)
from apps.kho.models import TonKho
from apps.so_cai.periods import (AccountingPeriodError,
                                 ensure_accounting_period_open,
                                 ensure_accounting_period_open_for_dates,
                                 get_current_accounting_period,
                                 guard_accounting_period_error)
from apps.so_cai.services import LedgerPostingError, post_to_ledger

from .models import (CongNoCanhBaoConfig, DonBan, DonBan_CT, HoaDonBan,
                     HoaDonBan_CT, PhieuGiaBan, PhieuGiaBan_CT,
                     PhieuGiaBanChietKhau, PhieuGiaoHang, PhieuGiaoHang_CT,
                     PhieuThu, PhieuTraHang, PhieuTraHang_CT)

EXCEL_COMPANY_NAME = 'CÔNG TY PHẦN MỀM QUẢN LÝ DOANH NGHIỆP (ERP TIEN HUONG)'
EXCEL_COMPANY_ADDRESS = 'Tầng 3, Tòa nhà CT1B - Khu VOV, Mễ Trì, Nam Từ Liêm, Hà Nội'
AUTO_CONG_NO_NOTE_PREFIX = '[AUTO_CONG_NO_HOA_DON]'

HOA_DON_EXCEL_HEADERS = [
    ('Số hóa đơn', 'so_hoa_don'),
    ('Ngày lập', 'ngay_lap'),
    ('Ngày hạch toán', 'ngay_hach_toan'),
    ('Mã giao dịch', 'ma_giao_dich'),
    ('Mã khách', 'ma_kh'),
    ('Tên khách', 'ten_kh'),
    ('Địa chỉ', 'dia_chi'),
    ('SĐT', 'so_dien_thoai'),
    ('MST', 'mst'),
    ('Người mua hàng', 'nguoi_mua_hang'),
    ('Mã hàng', 'ma_hang'),
    ('Tên hàng', 'ten_hang'),
    ('Mã kho', 'ma_kho'),
    ('Tên kho', 'ten_kho'),
    ('Số lượng', 'so_luong'),
    ('Giá bán', 'gia_ban'),
    ('CK %', 'ty_le_chiet_khau'),
    ('Thuế %', 'thue_suat'),
    ('TK nợ', 'tk_no'),
    ('TK có đối ứng', 'tk_co'),
    ('TK vật tư', 'tk_vat_tu'),
    ('TK giá vốn', 'tk_gia_von'),
    ('TK doanh thu', 'tk_doanh_thu'),
    ('Diễn giải', 'dien_giai'),
    ('Trạng thái', 'trang_thai'),
]


def _parse_decimal(value, default=Decimal('0')):
    if value in (None, ''):
        return default
    if isinstance(value, Decimal):
        return value
    try:
        raw = str(value).strip()
        if not raw:
            return default

        # New robust logic for VN/EN formats
        # 1. If both , and . exist, identify which is the decimal separator
        if ',' in raw and '.' in raw:
            if raw.rfind(',') > raw.rfind('.'):
                # VN: 1.234.567,89 -> remove . then replace , with .
                raw = raw.replace('.', '').replace(',', '.')
            else:
                # EN: 1,234,567.89 -> remove ,
                raw = raw.replace(',', '')
        elif ',' in raw:
            # Check if it looks like a decimal or thousand separator
            # If 1,234 -> thousand. If 1,23 -> decimal. 
            # In many cases 1,234 is also VN decimal for 1 point 234.
            # We assume single comma followed by 1 or 2 digits is decimal.
            # Else it's a thousand separator.
            parts = raw.split(',')
            if len(parts) == 2 and len(parts[1]) <= 2:
                raw = raw.replace(',', '.')
            else:
                raw = raw.replace(',', '')
        elif '.' in raw:
            # Similar for dot: 1.234 vs 1.23
            parts = raw.split('.')
            if len(parts) == 2 and len(parts[1]) <= 2:
                # 1.23 -> Decimal dot. Keep as-is.
                pass
            else:
                # 1.234 or 1.234.567 -> Thousand dot. Remove.
                raw = raw.replace('.', '')

        return Decimal(raw)
    except (InvalidOperation, AttributeError, ValueError):
        return default


def _normalize_decimal(value, decimal_places, default=Decimal('0')):
    parsed_value = _parse_decimal(value, default)
    if not isinstance(parsed_value, Decimal):
        parsed_value = Decimal(str(parsed_value))
    quantum = Decimal('1').scaleb(-decimal_places)
    try:
        return parsed_value.quantize(quantum)
    except InvalidOperation:
        return default


def _parse_date(value, default=None):
    if value in (None, ''):
        return default or date.today()
    if hasattr(value, 'date') and callable(value.date):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return date.fromisoformat(value) if fmt == '%Y-%m-%d' else timezone.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return default or date.today()


def _gen_so_hoa_don():
    prefix = 'HD'
    max_index = 0
    for code in HoaDonBan.objects.filter(so_hoa_don__istartswith=prefix).values_list('so_hoa_don', flat=True):
        text = str(code or '').strip().upper()
        if not text.startswith(prefix):
            continue
        suffix = text[len(prefix):]
        if suffix.isdigit():
            max_index = max(max_index, int(suffix))
    return f'{prefix}{max_index + 1:05d}'


def _next_code_from_queryset_values(values, prefix, width=5):
    max_index = 0
    for code in values:
        text = str(code or '').strip().upper()
        if not text.startswith(prefix):
            continue
        suffix = text[len(prefix):]
        if suffix.isdigit():
            max_index = max(max_index, int(suffix))
    return f'{prefix}{max_index + 1:0{width}d}'


def _style_export_sheet(ws, title, headers):
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = EXCEL_COMPANY_NAME
    ws['A1'].font = Font(name='Arial', size=12, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = EXCEL_COMPANY_ADDRESS
    ws['A2'].font = Font(name='Arial', size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A4:{last_col}4')
    ws['A4'] = title
    ws['A4'].font = Font(name='Arial', size=16, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    for idx, (header, _) in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A7'
    for idx, (header, _) in enumerate(headers, start=1):
        width = max(len(header) + 2, 12)
        if header in {'Tên khách', 'Địa chỉ', 'Diễn giải', 'Tên hàng', 'Tên kho'}:
            width = 24
        ws.column_dimensions[get_column_letter(idx)].width = width


def _find_header_row(ws, expected_headers):
    expected = [str(item).strip().lower() for item in expected_headers]
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [str(ws.cell(row=row, column=col).value or '').strip().lower() for col in range(1, len(expected) + 1)]
        if values == expected:
            return row
    return None


def _hoa_don_ban_filtered_queryset(request):
    items = HoaDonBan.objects.select_related('khach_hang', 'don_ban', 'nguoi_tao')
    q = request.GET.get('q', '').strip()
    ma_giao_dich = request.GET.get('ma_giao_dich', '').strip()
    trang_thai = request.GET.get('trang_thai', '').strip()
    tu_ngay = request.GET.get('tu_ngay', '').strip()
    den_ngay = request.GET.get('den_ngay', '').strip()
    ma_kh = request.GET.get('ma_kh', '').strip()

    if q:
        items = items.filter(
            Q(so_hoa_don__icontains=q)
            | Q(ten_kh__icontains=q)
            | Q(dia_chi__icontains=q)
            | Q(so_dien_thoai__icontains=q)
            | Q(mst__icontains=q)
            | Q(khach_hang__ma_kh__icontains=q)
            | Q(khach_hang__ten_kh__icontains=q)
        )
    if ma_giao_dich:
        items = items.filter(ma_giao_dich=ma_giao_dich)
    if trang_thai:
        items = items.filter(trang_thai=trang_thai)
    if tu_ngay:
        items = items.filter(ngay_lap__gte=tu_ngay)
    if den_ngay:
        items = items.filter(ngay_lap__lte=den_ngay)
    if ma_kh:
        items = items.filter(khach_hang__ma_kh__icontains=ma_kh)
    return items.order_by('-ngay_lap', '-id')


def _build_hoa_don_context(request, hoa_don=None):
    if hoa_don and hoa_don.khach_hang:
        kh_values = {
            'ten_kh': hoa_don.khach_hang.ten_kh,
            'so_dien_thoai': hoa_don.khach_hang.so_dien_thoai,
            'dia_chi': hoa_don.khach_hang.dia_chi,
            'mst': hoa_don.khach_hang.ma_so_thue,
        }
    else:
        kh_values = {'ten_kh': '', 'so_dien_thoai': '', 'dia_chi': '', 'mst': ''}

    current_period = get_current_accounting_period()

    nv_qs = KhachHang.objects.filter(trang_thai=True, la_nhan_vien=True)
    if hoa_don and hoa_don.ma_nv_ban_hang:
        nv_qs = KhachHang.objects.filter(
            Q(trang_thai=True, la_nhan_vien=True) | Q(ma_kh=hoa_don.ma_nv_ban_hang)
        )

    hang_qs = HangHoa.objects.filter(
        pk__in=TonKho.objects.filter(so_luong__gte=1).values_list('hang_hoa_id', flat=True).distinct()
    ).select_related('don_vi_tinh')
    kho_qs = Kho.objects.filter(trang_thai=True)
    chi_tiet = []
    if hoa_don and hoa_don.pk:
        chi_tiet_qs = hoa_don.chi_tiet.all()
        chi_tiet = list(chi_tiet_qs)
        hang_ids = list(chi_tiet_qs.values_list('hang_hoa_id', flat=True))
        kho_ids = list(chi_tiet_qs.values_list('kho_id', flat=True))
        if hang_ids:
            hang_qs = (hang_qs | HangHoa.objects.filter(pk__in=hang_ids).select_related('don_vi_tinh')).distinct()
        if kho_ids:
            kho_qs = (kho_qs | Kho.objects.filter(pk__in=kho_ids)).distinct()

    return {
        'hoa_don': hoa_don,
        'chi_tiet': chi_tiet,
        'editing': bool(hoa_don),
        'kh_list': KhachHang.objects.filter(Q(la_khach_hang=True) | Q(la_nhan_vien=True), trang_thai=True).order_by('ma_kh'),
        'nv_list': nv_qs,
        'kho_list': kho_qs.order_by('ma_kho'),
        'hang_list': hang_qs.order_by('ma_hang'),
        'so_hoa_don_default': hoa_don.so_hoa_don if hoa_don else _gen_so_hoa_don(),
        'today': date.today(),
        'current_period': current_period,
        **kh_values,
    }


def _save_hoa_don_from_request(request, hoa_don=None):
    data = request.POST
    old_trang_thai = str(getattr(hoa_don, 'trang_thai', '') or '').strip() if hoa_don and hoa_don.pk else ''
    kh_id = data.get('khach_hang') or None
    kh = KhachHang.objects.filter(pk=kh_id).first() if kh_id else None

    hoa_don = hoa_don or HoaDonBan()
    hoa_don.ma_giao_dich = data.get('ma_giao_dich', '1')
    hoa_don.so_hoa_don = data.get('so_hoa_don') or hoa_don.so_hoa_don or _gen_so_hoa_don()
    ngay_lap = _parse_date(data.get('ngay_lap') or data.get('ngay_chung_tu') or data.get('ngay_hach_toan'))
    ngay_chung_tu = _parse_date(data.get('ngay_chung_tu') or data.get('ngay_hach_toan') or data.get('ngay_lap'))
    hoa_don.ngay_lap = ngay_lap
    hoa_don.ngay_hach_toan = ngay_chung_tu
    ensure_accounting_period_open_for_dates([hoa_don.ngay_lap, hoa_don.ngay_hach_toan], 'hóa đơn bán hàng')

    hoa_don.ma_ngoai_te = data.get('ma_ngoai_te', 'VND')
    hoa_don.ty_gia = _normalize_decimal(data.get('ty_gia'), 4, Decimal('1'))
    hoa_don.khach_hang = kh
    hoa_don.ten_kh = data.get('ten_kh') or (kh.ten_kh if kh else '')
    hoa_don.dia_chi = data.get('dia_chi') or (kh.dia_chi if kh else '')
    hoa_don.so_dien_thoai = data.get('so_dien_thoai') or (kh.so_dien_thoai if kh else '')
    hoa_don.mst = data.get('mst') or (kh.ma_so_thue if kh else '')
    hoa_don.nguoi_mua_hang = data.get('nguoi_mua_hang', '')
    ma_nv_ban_hang = (data.get('ma_nv_ban_hang') or '').strip()
    if not ma_nv_ban_hang:
        ma_nv_ban_hang, _ = _parse_nv_display_value(data.get('nv_display'))
    if not ma_nv_ban_hang and data.get('don_ban'):
        don_lien_ket = DonBan.objects.filter(pk=data.get('don_ban')).only('ma_nv_ban_hang').first()
        if don_lien_ket and don_lien_ket.ma_nv_ban_hang:
            ma_nv_ban_hang = don_lien_ket.ma_nv_ban_hang
    if not ma_nv_ban_hang and hoa_don.pk:
        ma_nv_ban_hang = (hoa_don.ma_nv_ban_hang or '').strip()
    if not ma_nv_ban_hang:
        raise ValueError('Mã NV bán hàng là bắt buộc')
    hoa_don.ma_nv_ban_hang = ma_nv_ban_hang
    tk_no = (data.get('tk_no') or '131').strip()
    tk_co = (data.get('tk_co') or '').strip()
    if not tk_co:
        tk_co = '511'
    hoa_don.tk_no = tk_no
    hoa_don.tk_co = tk_co
    hoa_don.dien_giai = data.get('dien_giai', '')
    trang_thai_input = str(data.get('trang_thai', '1') or '1').strip()
    if trang_thai_input == '3':
        trang_thai_input = '2'
    hoa_don.trang_thai = trang_thai_input if trang_thai_input in ('1', '2') else '1'
    if data.get('don_ban'):
        hoa_don.don_ban_id = data.get('don_ban')
    elif not hoa_don.pk:
        hoa_don.don_ban_id = None
    hoa_don.save()

    # Nếu hóa đơn trước đó đã ghi sổ, hoàn tồn cũ trước khi cập nhật chi tiết mới.
    if old_trang_thai in ('2', '3'):
        _restore_ton_kho_from_hoa_don(hoa_don)

    if hoa_don.pk:
        hoa_don.chi_tiet.all().delete()

    hang_ids = data.getlist('hang_id[]')
    kho_ids = data.getlist('kho_id[]')
    so_luongs = data.getlist('so_luong[]')
    gia_bans = data.getlist('gia_ban[]')
    ty_le_cks = data.getlist('ty_le_ck[]')
    thue_suats = data.getlist('thue_suat[]')
    tk_doanh_thus = data.getlist('tk_doanh_thu[]')
    tk_gia_vons = data.getlist('tk_gia_von[]')
    tk_khos = data.getlist('tk_kho[]')

    valid_lines = 0
    for index in range(len(hang_ids)):
        hang_id = hang_ids[index] if index < len(hang_ids) else ''
        kho_id = kho_ids[index] if index < len(kho_ids) else ''
        if not hang_id or not kho_id:
            continue
        hang = HangHoa.objects.select_related('nhom_hang', 'don_vi_tinh').filter(pk=hang_id).first()
        if not hang:
            continue
        so_luong = _normalize_decimal(so_luongs[index] if index < len(so_luongs) else 0, 2)
        gia_ban = _normalize_decimal(gia_bans[index] if index < len(gia_bans) else 0, 0)
        ty_le_ck = _normalize_decimal(ty_le_cks[index] if index < len(ty_le_cks) else 0, 2)
        # 2 cách lấy giá: nhập tay hoặc để trống/0 để tự lấy từ phiếu giá bán.
        if gia_ban <= 0:
            gia_ban = _resolve_gia_ban_hang_hoa(hang)
        if ty_le_ck <= 0:
            ty_le_ck = _resolve_chiet_khau_hang_hoa(hang, so_luong)
        if so_luong <= 0:
            continue

        HoaDonBan_CT.objects.create(
            hoa_don=hoa_don,
            hang_hoa_id=hang_id,
            kho_id=kho_id,
            so_luong=so_luong,
            gia_ban=gia_ban,
            ty_le_chiet_khau=ty_le_ck,
            thue_suat=_normalize_decimal(thue_suats[index] if index < len(thue_suats) else 10, 2),
            tk_doanh_thu=(tk_doanh_thus[index] if index < len(tk_doanh_thus) else '511'),
            tk_gia_von=(tk_gia_vons[index] if index < len(tk_gia_vons) else '632'),
            tk_kho=(tk_khos[index] if index < len(tk_khos) else '156'),
        )
        valid_lines += 1


    if valid_lines == 0:
        hoa_don.delete()
        raise ValueError('Hóa đơn phải có ít nhất 1 dòng hàng hóa hợp lệ')

    hoa_don.tinh_tong()
    _apply_ton_kho_for_hoa_don(hoa_don)
    _sync_cong_no_from_hoa_don(hoa_don)
    return hoa_don


def _apply_ton_kho_for_hoa_don(hoa_don):
    if str(hoa_don.trang_thai or '').strip() not in ('2', '3'):
        return

    from apps.kho.models import PhieuXuat, PhieuXuat_CT
    from apps.kho.views import _gen_so_phieu

    rows = list(hoa_don.chi_tiet.select_related('hang_hoa', 'kho'))
    for ct in rows:
        ton = TonKho.objects.filter(hang_hoa=ct.hang_hoa, kho=ct.kho).first()
        if not ton or Decimal(ton.so_luong or 0) < Decimal(ct.so_luong or 0):
            raise ValueError(f'Không đủ tồn kho cho hàng {ct.hang_hoa.ma_hang} tại kho {ct.kho.ma_kho}')

    for ct in rows:
        ton = TonKho.objects.get(hang_hoa=ct.hang_hoa, kho=ct.kho)
        ton.so_luong = int(Decimal(ton.so_luong or 0) - Decimal(ct.so_luong or 0))
        ton.save(update_fields=['so_luong', 'ngay_cap_nhat'])
        
        from apps.kho.models import TonKhoViTri
        con_lai = int(Decimal(ct.so_luong or 0))
        vt_rows = list(
            TonKhoViTri.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=ct.kho, so_luong__gt=0)
            .exclude(vi_tri__ma_vi_tri__startswith='HL-')
            .order_by('-so_luong', 'id')
        )
        for vt in vt_rows:
            if con_lai <= 0: break
            tru = min(int(vt.so_luong or 0), con_lai)
            vt.so_luong = int(vt.so_luong or 0) - tru
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])
            con_lai -= tru

    # Tao phieu xuat tu dong neu ma_giao_dich = 1 (Hoa don kiem phieu xuat)
    if hoa_don.ma_giao_dich == '1':
        # Nhom theo kho_id de tao phieu (moi phieu 1 kho)
        kho_dict = {}
        for ct in rows:
            kho_dict.setdefault(ct.kho_id, []).append(ct)
        for kho_id, items in kho_dict.items():
            px = PhieuXuat.objects.create(
                so_phieu=_gen_so_phieu('PX'),
                ngay_lap=hoa_don.ngay_lap,
                ngay_hach_toan=hoa_don.ngay_hach_toan,
                ngay_chung_tu=hoa_don.ngay_lap,
                ngay_xuat=hoa_don.ngay_lap,
                loai_xuat='ban_hang',
                kho_id=kho_id,
                tong_gia_von=0,
                trang_thai='3', # Da vao so cai
                nguoi_tao=hoa_don.nguoi_tao,
                ghi_chu=f'Xuất kho tự động từ hóa đơn {hoa_don.so_hoa_don}',
            )
            tong_gia_von = 0
            for ct in items:
                ton = TonKho.objects.get(hang_hoa=ct.hang_hoa, kho_id=kho_id)
                # Lay gia von tb tu he thong
                gv_don_vi = int(ton.gia_von_tb or 0)
                gv_tong = gv_don_vi * int(ct.so_luong or 0)
                tong_gia_von += gv_tong
                PhieuXuat_CT.objects.create(
                    phieu_xuat=px,
                    hang_hoa=ct.hang_hoa,
                    so_luong=int(ct.so_luong or 0),
                    gia_von=gv_don_vi,
                    tong_gia_von=gv_tong,
                    tk_no=str(ct.tk_gia_von or '632').strip() or '632',
                    tk_co=str(getattr(ct, 'tk_kho', None) or '156').strip() or '156'
                )
            px.tong_gia_von = tong_gia_von
            px.save(update_fields=['tong_gia_von'])


def _restore_ton_kho_from_hoa_don(hoa_don):
    from apps.kho.models import PhieuXuat

    rows = list(hoa_don.chi_tiet.select_related('hang_hoa', 'kho'))
    for ct in rows:
        ton, _ = TonKho.objects.get_or_create(
            hang_hoa=ct.hang_hoa,
            kho=ct.kho,
            defaults={'so_luong': 0, 'gia_von_tb': 0},
        )
        ton.so_luong = int(Decimal(ton.so_luong or 0) + Decimal(ct.so_luong or 0))
        ton.save(update_fields=['so_luong', 'ngay_cap_nhat'])
        
        from apps.kho.models import TonKhoViTri, ViTriKho
        vt = TonKhoViTri.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=ct.kho).exclude(vi_tri__ma_vi_tri__startswith='HL-').first()
        if not vt:
            vi_tri_kho = ViTriKho.objects.filter(kho=ct.kho, trang_thai='hoat_dong').first()
            if vi_tri_kho:
                vt, _ = TonKhoViTri.objects.get_or_create(
                    hang_hoa=ct.hang_hoa, kho=ct.kho, vi_tri=vi_tri_kho, defaults={'so_luong': 0}
                )
        if vt:
            vt.so_luong = int(vt.so_luong or 0) + int(Decimal(ct.so_luong or 0))
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])
        
    # Xoa phieu xuat auto neu co
    if hoa_don.ma_giao_dich == '1':
        PhieuXuat.objects.filter(ghi_chu__startswith=f'Xuất kho tự động từ hóa đơn {hoa_don.so_hoa_don}').delete()


def _export_hoa_don_workbook(title, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'HoaDonBan'
    _style_export_sheet(ws, title, HOA_DON_EXCEL_HEADERS)

    for row_index, row in enumerate(rows, start=7):
        for col_index, (_, field_name) in enumerate(HOA_DON_EXCEL_HEADERS, start=1):
            value = row.get(field_name, '')
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if field_name in {'ngay_lap', 'ngay_hach_toan'} and hasattr(value, 'strftime'):
                cell.value = value.strftime('%d/%m/%Y')
            if field_name in {'so_luong', 'gia_ban', 'ty_le_chiet_khau', 'thue_suat', 'tong_cong'}:
                cell.alignment = Alignment(horizontal='right')
    return wb


def _gen_so_don(prefix):
    if prefix == 'BH':
        return _next_code_from_queryset_values(
            DonBan.objects.filter(so_don__istartswith='ĐH').values_list('so_don', flat=True),
            'ĐH',
        )

    if prefix == 'PT':
        return _next_code_from_queryset_values(
            PhieuThu.objects.filter(so_phieu__istartswith='PT').values_list('so_phieu', flat=True),
            'PT',
        )

    if prefix == 'PC':
        return _next_code_from_queryset_values(
            PhieuThu.objects.filter(so_phieu__istartswith='PC').values_list('so_phieu', flat=True),
            'PC',
        )

    now = timezone.now()
    return f"{prefix}-{now.strftime('%Y%m%d-%H%M%S')}"


def _gen_so_phieu_thu():
    return _next_code_from_queryset_values(
        PhieuThu.objects.filter(so_phieu__istartswith='PT').values_list('so_phieu', flat=True),
        'PT',
    )


def _gen_so_phieu_chi():
    return _next_code_from_queryset_values(
        PhieuThu.objects.filter(so_phieu__istartswith='PC').values_list('so_phieu', flat=True),
        'PC',
    )


def _ensure_unique_so_phieu_thu(candidate=None):
    value = (candidate or '').strip()
    if value and not PhieuThu.objects.filter(so_phieu=value).exists():
        return value

    for _ in range(5):
        value = _gen_so_phieu_thu()
        if not PhieuThu.objects.filter(so_phieu=value).exists():
            return value

    raise ValueError('Không tạo được số phiếu thu duy nhất')


def _ensure_unique_so_phieu_chi(candidate=None):
    value = (candidate or '').strip()
    if value and not PhieuThu.objects.filter(so_phieu=value).exists():
        return value

    for _ in range(5):
        value = _gen_so_phieu_chi()
        if not PhieuThu.objects.filter(so_phieu=value).exists():
            return value

    raise ValueError('Không tạo được số phiếu chi duy nhất')


def _is_phieu_chi_record(phieu):
    return '[LOAI_PHIEU:CHI]' in str(getattr(phieu, 'ghi_chu', '') or '')


def _resolve_han_thanh_toan(ngay_ban, khach_hang, han_thanh_toan_raw=''):
    if han_thanh_toan_raw:
        return _parse_date(han_thanh_toan_raw)
    so_ngay_no = int(getattr(khach_hang, 'so_ngay_no_max', 0) or 0)
    if so_ngay_no > 0 and ngay_ban:
        return ngay_ban + timedelta(days=so_ngay_no)
    return None


def _parse_nv_display_value(raw_value):
    raw = str(raw_value or '').strip()
    if not raw:
        return '', ''
    if ' - ' in raw:
        ma, ten = raw.split(' - ', 1)
    else:
        ma, ten = raw, raw
    ma = ma.strip().upper()
    ten = ten.strip() or ma
    return ma, ten


def _resolve_or_create_sales_employee(data):
    """Resolve salesperson from selected id or free text; auto-create catalog item if needed."""
    nv_ban_id = data.get('nv_ban_id') or None
    if nv_ban_id:
        nv = KhachHang.objects.filter(pk=nv_ban_id, trang_thai=True, la_nhan_vien=True).first()
        if nv:
            return nv.ma_kh, nv

    ma_nv, ten_nv = _parse_nv_display_value(data.get('nv_display'))
    if not ma_nv:
        return '', None

    nv = KhachHang.objects.filter(ma_kh=ma_nv).first()
    if nv:
        if not nv.la_nhan_vien or not nv.trang_thai:
            nv.la_nhan_vien = True
            nv.trang_thai = True
            nv.save(update_fields=['la_nhan_vien', 'trang_thai'])
        return ma_nv, nv

    # Tạo nhanh danh mục nhân viên bán hàng khi người dùng nhập tay mã mới.
    nv = KhachHang.objects.create(
        ma_kh=ma_nv,
        ten_kh=ten_nv,
        so_dien_thoai='0000000000',
        la_khach_hang=False,
        la_nha_cung_cap=False,
        la_nhan_vien=True,
        loai_kh='1',
        trang_thai=True,
        ghi_chu='[AUTO_CREATE_NV] Tạo tự động từ màn hình đơn hàng bán',
    )
    return ma_nv, nv


def _is_hd_ghi_nhan_cong_no(hd):
    return hd and hd.tk_no == '131' and hd.trang_thai in ('2', '3')


def _is_hd_thu_tien_ngay(hd):
    return hd and hd.tk_no in ('111', '112') and hd.trang_thai in ('2', '3')


def _sync_cong_no_from_hoa_don(hoa_don):
    if not hoa_don or not hoa_don.khach_hang_id:
        return

    # Cho phép tính toán công nợ ngay cả khi ở trạng thái Lập chứng từ (trạng thái 1) 
    # để có thể lập phiếu thu/chi kế thừa từ hóa đơn nháp.
    if str(hoa_don.trang_thai or '').strip() == '4': # Bỏ qua nếu đã hủy
        return

    hoa_don.tinh_tong() # Đảm bảo con_no đã được tính lại từ chi tiết

    # Chỉ ghi nhận công nợ khi hóa đơn hạch toán vào TK phải thu 131.
    if str(hoa_don.tk_no or '').strip() != '131':
        if hoa_don.don_ban:
            _recompute_don_ban_from_linked_hoa_don(hoa_don.don_ban)
        return

    chi_tiet_qs = hoa_don.chi_tiet.select_related('kho')
    kho_default = chi_tiet_qs.first().kho if chi_tiet_qs.exists() else Kho.objects.filter(trang_thai=True).first()
    if not kho_default:
        raise ValueError('Không tìm thấy kho để đồng bộ công nợ từ hóa đơn.')

    don = hoa_don.don_ban
    if not don:
        don = DonBan.objects.create(
            so_don=f'CN-{hoa_don.id}',
            ngay_chung_tu=hoa_don.ngay_lap,
            ngay_ban=hoa_don.ngay_lap,
            loai_ban='ban_buon',
            khach_hang=hoa_don.khach_hang,
            ten_kh=hoa_don.ten_kh or hoa_don.khach_hang.ten_kh,
            sdt_kh=hoa_don.so_dien_thoai or hoa_don.khach_hang.so_dien_thoai,
            dia_chi_kh=hoa_don.dia_chi or hoa_don.khach_hang.dia_chi,
            mst_kh=hoa_don.mst or hoa_don.khach_hang.ma_so_thue,
            nguoi_mua_hang=hoa_don.nguoi_mua_hang,
            ma_nv_ban_hang=hoa_don.ma_nv_ban_hang or '',
            kho=kho_default,
            nhan_vien_ban=hoa_don.nguoi_tao,
            phuong_thuc_tt='no',
            chiet_khau_dh=Decimal('0'),
            ma_ngoai_te=hoa_don.ma_ngoai_te or 'VND',
            ty_gia=hoa_don.ty_gia or Decimal('1'),
            tong_tien_hang=hoa_don.tien_hang or Decimal('0'),
            tong_thue=hoa_don.tong_tien_thue or Decimal('0'),
            tong_thanh_toan=hoa_don.tong_cong or Decimal('0'),
            da_thu=Decimal('0'),
            con_no=hoa_don.tong_cong or Decimal('0'),
            han_thanh_toan=_resolve_han_thanh_toan(hoa_don.ngay_lap, hoa_don.khach_hang),
            ghi_chu=f'{AUTO_CONG_NO_NOTE_PREFIX} {hoa_don.so_hoa_don}',
        )
    else:
        da_thu = Decimal(don.da_thu or 0)
        tong_thanh_toan = Decimal(hoa_don.tong_cong or 0)
        don.ngay_chung_tu = hoa_don.ngay_lap
        don.ngay_ban = hoa_don.ngay_lap
        don.khach_hang = hoa_don.khach_hang
        don.ten_kh = hoa_don.ten_kh or hoa_don.khach_hang.ten_kh
        don.sdt_kh = hoa_don.so_dien_thoai or hoa_don.khach_hang.so_dien_thoai
        don.dia_chi_kh = hoa_don.dia_chi or hoa_don.khach_hang.dia_chi
        don.mst_kh = hoa_don.mst or hoa_don.khach_hang.ma_so_thue
        don.nguoi_mua_hang = hoa_don.nguoi_mua_hang
        don.ma_nv_ban_hang = hoa_don.ma_nv_ban_hang or don.ma_nv_ban_hang
        don.kho = kho_default
        don.phuong_thuc_tt = 'no'
        don.ma_ngoai_te = hoa_don.ma_ngoai_te or 'VND'
        don.ty_gia = hoa_don.ty_gia or Decimal('1')
        don.tong_tien_hang = hoa_don.tien_hang or Decimal('0')
        don.tong_thue = hoa_don.tong_tien_thue or Decimal('0')
        don.tong_thanh_toan = tong_thanh_toan
        don.con_no = tong_thanh_toan - da_thu
        don.han_thanh_toan = _resolve_han_thanh_toan(hoa_don.ngay_lap, hoa_don.khach_hang, don.han_thanh_toan)
        if not (don.ghi_chu or '').strip():
            don.ghi_chu = f'{AUTO_CONG_NO_NOTE_PREFIX} {hoa_don.so_hoa_don}'
        don.save()

    if hoa_don.don_ban_id != don.id:
        hoa_don.don_ban = don
        hoa_don.save(update_fields=['don_ban'])


def _recompute_don_ban_from_linked_hoa_don(don):
    if not don:
        return

    linked = don.hoa_don_lien_ket.filter(tk_no='131', trang_thai__in=('2', '3'))
    if not linked.exists():
        if (don.ghi_chu or '').startswith(AUTO_CONG_NO_NOTE_PREFIX) and Decimal(don.da_thu or 0) == 0:
            don.delete()
        return

    tong = linked.aggregate(total=Sum('tong_cong'))['total'] or Decimal('0')
    da_thu = Decimal(don.da_thu or 0)
    tong_tien_hang = linked.aggregate(total=Sum('tien_hang'))['total'] or Decimal('0')
    tong_thue = linked.aggregate(total=Sum('tong_tien_thue'))['total'] or Decimal('0')
    don.tong_thanh_toan = tong
    don.tong_tien_hang = tong_tien_hang
    don.tong_thue = tong_thue
    don.con_no = tong - da_thu
    don.save(update_fields=['tong_thanh_toan', 'tong_tien_hang', 'tong_thue', 'con_no'])


# ─── ĐƠN BÁN HÀNG ───────────────────────────────────────────
@login_required
def don_ban_list(request):
    q = request.GET.get('q', '')
    trang_thai = request.GET.get('trang_thai', '')
    items = DonBan.objects.select_related('khach_hang', 'kho')
    if q:
        items = items.filter(Q(so_don__icontains=q) | Q(ten_kh__icontains=q) |
                             Q(sdt_kh__icontains=q))
    if trang_thai:
        items = items.filter(trang_thai=trang_thai)
    context = {
        'items': items[:50],
        'q': q,
        'trang_thai_filter': trang_thai,
        'page_title': 'Đơn bán hàng',
        'active_menu': 'don_ban',
    }
    return render(request, 'ban_hang/don_ban_list.html', context)


@login_required
def don_ban_them(request):
    # Xử lý copy_from: nếu có tham số copy_from trên URL, lấy dữ liệu đơn bán gốc để đổ vào form tạo mới
    copy_from = request.GET.get('copy_from')
    don_copy = None
    chi_tiet_copy = []
    if copy_from:
        try:
            don_copy = DonBan.objects.get(pk=copy_from)
            chi_tiet_copy = list(don_copy.chi_tiet.all())
        except DonBan.DoesNotExist:
            don_copy = None
            chi_tiet_copy = []

    if request.method == 'POST':
        data = request.POST
        kh_id = data.get('khach_hang') or None
        kh = KhachHang.objects.filter(pk=kh_id).first() if kh_id else None
        ma_nv_ban_hang, _nv_obj = _resolve_or_create_sales_employee(data)
        loai_ban = data.get('loai_ban', 'ban_le')
        phuong_thuc_tt = data.get('phuong_thuc_tt')
        if not phuong_thuc_tt:
            phuong_thuc_tt = 'tien_mat' if loai_ban == 'ban_le' else 'no'

        if not ma_nv_ban_hang:
            messages.error(request, 'Mã NV bán hàng là bắt buộc')
            return redirect('don_ban_them')

        don = DonBan.objects.create(
            so_don=data.get('so_don') or _gen_so_don('BH'),
            ngay_chung_tu=_parse_date(data.get('ngay_chung_tu')),
            ngay_ban=_parse_date(data.get('ngay_chung_tu')),
            loai_ban=loai_ban,
            khach_hang=kh,
            ten_kh=data.get('ten_kh') or (kh.ten_kh if kh else 'Khách lẻ'),
            sdt_kh=data.get('sdt_kh', ''),
            dia_chi_kh=data.get('dia_chi_kh', ''),
            mst_kh=data.get('mst_kh', ''),
            ma_ngoai_te=data.get('ma_ngoai_te', 'VND'),
            ty_gia=_parse_decimal(data.get('ty_gia', 1), Decimal('1')),
            ma_nv_ban_hang=ma_nv_ban_hang,
            xe_kh=data.get('xe_kh', ''),
            nguoi_mua_hang=data.get('nguoi_mua_hang', ''),
            kho_id=data.get('kho'),
            nhan_vien_ban=request.user,
            phuong_thuc_tt=phuong_thuc_tt,
            chiet_khau_dh=_parse_decimal(data.get('chiet_khau_dh', 0), Decimal('0')),
            ghi_chu=data.get('ghi_chu', ''),
            trang_thai=data.get('trang_thai', '1'),
            han_thanh_toan=_resolve_han_thanh_toan(_parse_date(data.get('ngay_chung_tu')), kh, data.get('han_thanh_toan', '')),
        )

        hang_ids = data.getlist('hang_id[]')
        so_luongs = data.getlist('so_luong[]')
        don_gias = data.getlist('don_gia[]')
        cks = data.getlist('chiet_khau[]')
        vats = data.getlist('vat[]')

        so_dong_hop_le = 0
        for i in range(len(hang_ids)):
            if hang_ids[i] and so_luongs[i]:
                hang = HangHoa.objects.select_related('nhom_hang', 'don_vi_tinh').filter(pk=hang_ids[i]).first()
                if not hang:
                    continue
                so_luong = int(so_luongs[i])
                raw_gia = don_gias[i] if i < len(don_gias) else ''
                raw_ck = cks[i] if i < len(cks) else ''
                raw_vat = vats[i] if i < len(vats) else ''

                don_gia = _parse_decimal(raw_gia, Decimal('0'))
                chiet_khau = _parse_decimal(raw_ck, Decimal('0'))
                thue_vat = _parse_decimal(raw_vat, Decimal('10'))

                if raw_gia == '' or (don_gia <= 0 and not raw_gia.strip() == '0'):
                    don_gia = _resolve_gia_ban_hang_hoa(hang)
                if raw_ck == '' or (chiet_khau <= 0 and not raw_ck.strip() == '0'):
                    chiet_khau = _resolve_chiet_khau_hang_hoa(hang, so_luong)
                if raw_vat == '':
                    thue_vat = Decimal('10')

                if so_luong <= 0 or don_gia < 0:
                    continue

                DonBan_CT.objects.create(
                    don_ban=don,
                    hang_hoa=hang,
                    so_luong=so_luong,
                    don_gia=don_gia,
                    chiet_khau=chiet_khau,
                    thue_vat=thue_vat,
                )
                so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            don.delete()
            messages.error(request, 'Đơn bán phải có ít nhất 1 dòng hàng hóa hợp lệ')
            return redirect('don_ban_them')

        don.tinh_tong()
        messages.success(request, f'Đã tạo đơn {don.so_don}')
        return redirect('don_ban_detail', pk=don.pk)

    context = {
        'kh_list': KhachHang.objects.filter(Q(la_khach_hang=True) | Q(la_nhan_vien=True), trang_thai=True).order_by('ma_kh'),
        'nv_list': KhachHang.objects.filter(trang_thai=True, la_nhan_vien=True).order_by('ma_kh'),
        'kho_list': Kho.objects.filter(trang_thai=True),
        'hang_list': HangHoa.objects.all().order_by('ma_hang'),
        'so_don_default': _gen_so_don('BH'),
        'today': date.today(),
        'page_title': 'Tạo đơn bán hàng',
        'active_menu': 'don_ban',
        'don_copy': don_copy,
        'chi_tiet_copy': chi_tiet_copy,
    }
    return render(request, 'ban_hang/don_ban_form.html', context)


@login_required
def don_ban_detail(request, pk):
    don = get_object_or_404(DonBan, pk=pk)
    chi_tiet = don.chi_tiet.select_related('hang_hoa')
    nv_ban_hang = KhachHang.objects.filter(ma_kh=don.ma_nv_ban_hang, la_nhan_vien=True).first()
    context = {
        'don': don,
        'chi_tiet': chi_tiet,
        'nv_ban_hang_ten': nv_ban_hang.ten_kh if nv_ban_hang else (don.ma_nv_ban_hang or '-'),
        'page_title': f'Đơn bán {don.so_don}',
        'active_menu': 'don_ban',
    }
    return render(request, 'ban_hang/don_ban_detail.html', context)


@login_required
def don_ban_xac_nhan(request, pk):
    don = get_object_or_404(DonBan, pk=pk)
    if request.method == 'POST':
        ok, msg = don.xac_nhan_don_ban()
        if ok:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    return redirect('don_ban_detail', pk=pk)


@login_required
def don_ban_huy(request, pk):
    don = get_object_or_404(DonBan, pk=pk)
    if request.method == 'POST' and don.trang_thai == 'nhap':
        don.trang_thai = 'huy'
        don.save()
        messages.success(request, f'Đã hủy đơn {don.so_don}')
    return redirect('don_ban_list')


@login_required
def don_ban_sua(request, pk):
    don = get_object_or_404(DonBan.objects.select_related('khach_hang', 'kho').prefetch_related('chi_tiet__hang_hoa'), pk=pk)

    if request.method == 'POST':
        data = request.POST
        kh_id = data.get('khach_hang') or None
        kh = KhachHang.objects.filter(pk=kh_id).first() if kh_id else None
        ma_nv_ban_hang, _nv_obj = _resolve_or_create_sales_employee(data)

        if not ma_nv_ban_hang and not (don.ma_nv_ban_hang or '').strip():
            messages.error(request, 'Mã NV bán hàng là bắt buộc')
            return redirect('don_ban_sua', pk=pk)

        loai_ban = data.get('loai_ban', don.loai_ban or 'ban_le')
        phuong_thuc_tt = data.get('phuong_thuc_tt') or ('tien_mat' if loai_ban == 'ban_le' else 'no')

        don.ngay_chung_tu = _parse_date(data.get('ngay_chung_tu'))
        don.ngay_ban = don.ngay_chung_tu
        don.loai_ban = loai_ban
        don.khach_hang = kh
        don.ten_kh = data.get('ten_kh') or (kh.ten_kh if kh else 'Khách lẻ')
        don.sdt_kh = data.get('sdt_kh', '')
        don.dia_chi_kh = data.get('dia_chi_kh', '')
        don.mst_kh = data.get('mst_kh', '')
        don.ma_ngoai_te = data.get('ma_ngoai_te', 'VND')
        don.ty_gia = _parse_decimal(data.get('ty_gia', 1), Decimal('1'))
        don.ma_nv_ban_hang = ma_nv_ban_hang or (don.ma_nv_ban_hang or '')
        don.xe_kh = data.get('xe_kh', '')
        don.nguoi_mua_hang = data.get('nguoi_mua_hang', '')
        don.kho_id = data.get('kho')
        don.phuong_thuc_tt = phuong_thuc_tt
        don.chiet_khau_dh = _parse_decimal(data.get('chiet_khau_dh', 0), Decimal('0'))
        don.ghi_chu = data.get('ghi_chu', '')
        don.trang_thai = data.get('trang_thai', don.trang_thai or '1')
        don.han_thanh_toan = _resolve_han_thanh_toan(don.ngay_ban, kh, data.get('han_thanh_toan', ''))
        don.nhan_vien_ban = request.user
        don.save()

        don.chi_tiet.all().delete()
        hang_ids = data.getlist('hang_id[]')
        so_luongs = data.getlist('so_luong[]')
        don_gias = data.getlist('don_gia[]')
        cks = data.getlist('chiet_khau[]')
        vats = data.getlist('vat[]')

        so_dong_hop_le = 0
        for i in range(len(hang_ids)):
            if hang_ids[i] and so_luongs[i]:
                hang = HangHoa.objects.select_related('nhom_hang', 'don_vi_tinh').filter(pk=hang_ids[i]).first()
                if not hang:
                    continue
                so_luong = int(so_luongs[i])
                raw_gia = don_gias[i] if i < len(don_gias) else ''
                raw_ck = cks[i] if i < len(cks) else ''
                raw_vat = vats[i] if i < len(vats) else ''

                don_gia = _parse_decimal(raw_gia, Decimal('0'))
                chiet_khau = _parse_decimal(raw_ck, Decimal('0'))
                thue_vat = _parse_decimal(raw_vat, Decimal('10'))

                if raw_gia == '' or (don_gia <= 0 and not raw_gia.strip() == '0'):
                    don_gia = _resolve_gia_ban_hang_hoa(hang)
                if raw_ck == '' or (chiet_khau <= 0 and not raw_ck.strip() == '0'):
                    chiet_khau = _resolve_chiet_khau_hang_hoa(hang, so_luong)
                if raw_vat == '':
                    thue_vat = Decimal('10')

                if so_luong <= 0 or don_gia < 0:
                    continue

                DonBan_CT.objects.create(
                    don_ban=don,
                    hang_hoa=hang,
                    so_luong=so_luong,
                    don_gia=don_gia,
                    chiet_khau=chiet_khau,
                    thue_vat=thue_vat,
                )
                so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            messages.error(request, 'Đơn bán phải có ít nhất 1 dòng hàng hóa hợp lệ')
            return redirect('don_ban_sua', pk=pk)

        don.tinh_tong()
        messages.success(request, f'Đã cập nhật đơn {don.so_don}')
        return redirect('don_ban_detail', pk=don.pk)

    nv_list = KhachHang.objects.filter(trang_thai=True, la_nhan_vien=True)
    if (don.ma_nv_ban_hang or '').strip():
        nv_list = (nv_list | KhachHang.objects.filter(ma_kh=don.ma_nv_ban_hang)).distinct()

    hang_ids = list(don.chi_tiet.values_list('hang_hoa_id', flat=True))
    hang_qs = HangHoa.objects.filter(
        pk__in=TonKho.objects.filter(so_luong__gte=1).values_list('hang_hoa_id', flat=True).distinct()
    )
    if hang_ids:
        hang_qs = (hang_qs | HangHoa.objects.filter(pk__in=hang_ids)).distinct()

    kho_qs = Kho.objects.filter(trang_thai=True)
    if don.kho_id:
        kho_qs = (kho_qs | Kho.objects.filter(pk=don.kho_id)).distinct()

    context = {
        'don': don,
        'editing': True,
        'kh_list': KhachHang.objects.filter(Q(la_khach_hang=True) | Q(la_nhan_vien=True), trang_thai=True).order_by('ma_kh'),
        'nv_list': nv_list.order_by('ma_kh'),
        'kho_list': kho_qs.order_by('ma_kho'),
        'hang_list': hang_qs.order_by('ma_hang'),
        'so_don_default': don.so_don,
        'today': don.ngay_chung_tu or date.today(),
        'chi_tiet': don.chi_tiet.select_related('hang_hoa'),
        'page_title': f'Sửa đơn bán {don.so_don}',
        'active_menu': 'don_ban',
    }
    return render(request, 'ban_hang/don_ban_form.html', context)


def don_ban_xoa(request, pk):
    don = get_object_or_404(DonBan, pk=pk)
    if request.method == 'POST':
        so_don = don.so_don
        don.delete()
        messages.success(request, f'Đã xóa đơn {so_don}')
    return redirect('don_ban_list')


@login_required
def don_ban_xoa_nhieu(request):
    if request.method != 'POST':
        return redirect('don_ban_list')

    ids = request.POST.getlist('ids[]') or request.POST.getlist('ids')
    if not ids:
        messages.error(request, 'Vui lòng chọn ít nhất 1 đơn hàng để xóa.')
        return redirect('don_ban_list')

    items = DonBan.objects.filter(pk__in=ids)
    tong = items.count()
    if tong == 0:
        messages.error(request, 'Không tìm thấy đơn hàng cần xóa.')
        return redirect('don_ban_list')

    items.delete()
    messages.success(request, f'Đã xóa {tong} đơn hàng.')
    return redirect('don_ban_list')


def don_ban_chuyen_so_cai(request, pk):
    return _view_not_ready('chuyển sổ cái đơn bán hàng')


@login_required
def don_ban_export_data(request):
    rows = []
    items = DonBan.objects.select_related('khach_hang', 'kho', 'nhan_vien_ban')
    q = request.GET.get('q', '').strip()
    if q:
        items = items.filter(Q(so_don__icontains=q) | Q(ten_kh__icontains=q) | Q(sdt_kh__icontains=q))
    
    for don in items.prefetch_related('chi_tiet__hang_hoa'):
        for ct in don.chi_tiet.all():
            rows.append({
                'so_don': don.so_don,
                'ngay_ban': don.ngay_ban,
                'ma_kh': don.khach_hang.ma_kh if don.khach_hang else '',
                'ten_kh': don.ten_kh,
                'sdt_kh': don.sdt_kh,
                'ma_hang': ct.hang_hoa.ma_hang,
                'ten_hang': ct.hang_hoa.ten_hang,
                'so_luong': ct.so_luong,
                'don_gia': ct.don_gia,
                'chiet_khau': ct.chiet_khau,
                'thue_vat': ct.thue_vat,
                'trang_thai': don.get_trang_thai_display(),
            })
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'DonBan'
    headers = [
        ('Số đơn', 'so_don'),
        ('Ngày bán', 'ngay_ban'),
        ('Mã khách', 'ma_kh'),
        ('Tên khách', 'ten_kh'),
        ('SĐT', 'sdt_kh'),
        ('Mã hàng', 'ma_hang'),
        ('Tên hàng', 'ten_hang'),
        ('Số lượng', 'so_luong'),
        ('Đơn giá', 'don_gia'),
        ('Chiết khấu', 'chiet_khau'),
        ('Thuế VAT', 'thue_vat'),
        ('Trạng thái', 'trang_thai'),
    ]
    _style_export_sheet(ws, 'DANH SÁCH ĐƠN BÁN HÀNG', headers)
    
    for row_index, row in enumerate(rows, start=7):
        for col_index, (_, field_name) in enumerate(headers, start=1):
            value = row.get(field_name, '')
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if field_name == 'ngay_ban' and hasattr(value, 'strftime'):
                cell.value = value.strftime('%d/%m/%Y')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=don_ban.xlsx'
    wb.save(response)
    return response


@login_required
def don_ban_export_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'DonBan'
    headers = [
        ('Số đơn', 'so_don'),
        ('Ngày bán', 'ngay_ban'),
        ('Mã khách', 'ma_kh'),
        ('Tên khách', 'ten_kh'),
        ('SĐT', 'sdt_kh'),
        ('Mã hàng', 'ma_hang'),
        ('Tên hàng', 'ten_hang'),
        ('Số lượng', 'so_luong'),
        ('Đơn giá', 'don_gia'),
        ('Chiết khấu', 'chiet_khau'),
        ('Thuế VAT', 'thue_vat'),
    ]
    _style_export_sheet(ws, 'MẪU NHẬP ĐƠN BÁN HÀNG', headers)
    ws['A7'] = 'Nhập dữ liệu từ dòng 7 trở đi, giữ nguyên các cột đã có.'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=mau_don_ban.xlsx'
    wb.save(response)
    return response


@login_required
def don_ban_import_excel(request):
    if request.method != 'POST':
        return redirect('don_ban_list')
    
    uploaded = request.FILES.get('excel_file')
    if not uploaded:
        messages.error(request, 'Vui lòng chọn tệp Excel.')
        return redirect('don_ban_list')
    
    try:
        wb = load_workbook(uploaded)
    except Exception:
        messages.error(request, 'Không đọc được tệp Excel.')
        return redirect('don_ban_list')
    
    ws = wb.active
    created = updated = skipped = 0
    
    for row_idx in range(8, ws.max_row + 1):
        so_don = str(ws.cell(row=row_idx, column=1).value or '').strip()
        ngay_ban = ws.cell(row=row_idx, column=2).value
        ma_kh = str(ws.cell(row=row_idx, column=3).value or '').strip()
        ten_kh = str(ws.cell(row=row_idx, column=4).value or '').strip()
        sdt_kh = str(ws.cell(row=row_idx, column=5).value or '').strip()
        ma_hang = str(ws.cell(row=row_idx, column=6).value or '').strip()
        so_luong = ws.cell(row=row_idx, column=8).value
        don_gia = ws.cell(row=row_idx, column=9).value
        
        if not so_don or not ma_hang:
            skipped += 1
            continue
        
        kh = KhachHang.objects.filter(ma_kh=ma_kh).first() if ma_kh else None
        hang = HangHoa.objects.filter(ma_hang=ma_hang).first()
        
        if not hang:
            skipped += 1
            continue
        
        don = DonBan.objects.filter(so_don=so_don).first()
        if don:
            updated += 1
            don.chi_tiet.all().delete()
        else:
            don = DonBan()
            created += 1
        
        don.so_don = so_don
        don.ngay_ban = _parse_date(ngay_ban) if ngay_ban else date.today()
        don.khach_hang = kh
        don.ten_kh = ten_kh or (kh.ten_kh if kh else 'Khách lẻ')
        don.sdt_kh = sdt_kh
        don.nhan_vien_ban = request.user
        don.han_thanh_toan = _resolve_han_thanh_toan(don.ngay_ban, kh)
        don.save()
        
        if so_luong and don_gia:
            DonBan_CT.objects.create(
                don_ban=don,
                hang_hoa=hang,
                so_luong=int(_parse_decimal(so_luong, 0)),
                don_gia=_parse_decimal(don_gia),
                chiet_khau=_parse_decimal(ws.cell(row=row_idx, column=10).value or 0),
                thue_vat=_parse_decimal(ws.cell(row=row_idx, column=11).value or 10),
            )
        don.tinh_tong()
    
    messages.success(request, f'Đã nhập đơn bán: tạo mới {created}, cập nhật {updated}, bỏ qua {skipped}.')
    return redirect('don_ban_list')



@login_required
def phieu_thu_list(request, mode='thu'):
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    items = PhieuThu.objects.select_related('khach_hang').order_by('-ngay_thu')
    if is_chi_mode:
        items = items.filter(ghi_chu__icontains='[LOAI_PHIEU:CHI]')
    else:
        items = items.exclude(ghi_chu__icontains='[LOAI_PHIEU:CHI]')

    return render(request, 'ban_hang/phieu_thu_list.html', {
        'items': items[:50],
        'page_title': 'Phiếu chi tiền' if is_chi_mode else 'Phiếu thu tiền',
        'active_menu': 'phieu_chi' if is_chi_mode else 'phieu_thu',
        'loai': 'chi' if is_chi_mode else 'thu',
    })


def _phieu_thu_hoa_don_queryset():
    return (
        HoaDonBan.objects.select_related('khach_hang')
        .exclude(trang_thai='4')  # Không lấy hóa đơn đã hủy
        .filter(con_no__gt=0)     # Chỉ lấy những đơn thực sự còn nợ (đã tính lại qua shell command)
        .distinct()
        .order_by('-ngay_lap', '-id')
    )


def _phieu_thu_khach_hang_cong_no_map():
    hd_qs = HoaDonBan.objects.select_related('khach_hang').filter(
        khach_hang__isnull=False,
        khach_hang__trang_thai=True,
    ).distinct()

    data = {}
    for hd in hd_qs:
        kh_id = hd.khach_hang_id
        if not kh_id:
            continue
        if kh_id not in data:
            data[kh_id] = {
                'tong_hoa_don': Decimal('0'),
                'tong_da_thu': Decimal('0'),
                'con_no': Decimal('0'),
            }
        data[kh_id]['tong_hoa_don'] += Decimal(hd.tong_cong or 0)
        data[kh_id]['tong_da_thu'] += Decimal(hd.da_thu or 0)

    thu_theo_kh_qs = PhieuThu.objects.filter(
        trang_thai='2',
        hoa_don__isnull=True,
        ghi_chu__icontains='[LOAI_PHIEU_THU:2]',
        khach_hang__isnull=False,
    )
    for pt in thu_theo_kh_qs:
        kh_id = pt.khach_hang_id
        if not kh_id:
            continue
        if kh_id not in data:
            data[kh_id] = {
                'tong_hoa_don': Decimal('0'),
                'tong_da_thu': Decimal('0'),
                'con_no': Decimal('0'),
            }
        data[kh_id]['tong_da_thu'] += Decimal(pt.tong_thu or 0)

    for _, row in data.items():
        row['con_no'] = row['tong_hoa_don'] - row['tong_da_thu']
        if row['con_no'] < 0:
            row['con_no'] = Decimal('0')

    return data


def _build_phieu_thu_context(form_values=None, hoa_don_selected=None, editing=False, mode='thu'):
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    form_values = form_values or {}
    hinh_thuc = str(form_values.get('hinh_thuc_thu') or 'tien_mat').strip()
    tk_tien = '112' if hinh_thuc == 'chuyen_khoan' else '111'
    if is_chi_mode:
        form_values.setdefault('tk_no', tk_tien)
        form_values.setdefault('tk_co', '131')
    else:
        form_values.setdefault('tk_no', '131')
        form_values.setdefault('tk_co', tk_tien)
    if str(form_values.get('trang_thai') or '').strip() == '3':
        form_values['trang_thai'] = '2'
    elif not str(form_values.get('trang_thai') or '').strip():
        form_values['trang_thai'] = '1'
    if is_chi_mode:
        # Phiếu chi: Chỉ hiện Nhà cung cấp
        kh_list = list(KhachHang.objects.filter(trang_thai=True, la_nha_cung_cap=True).order_by('ma_kh'))
    else:
        # Phiếu thu: Chỉ hiện Khách hàng hoặc Nhân viên
        kh_list = list(KhachHang.objects.filter(
            Q(la_khach_hang=True) | Q(la_nhan_vien=True),
            trang_thai=True
        ).order_by('ma_kh'))
    kh_cong_no_map = _phieu_thu_khach_hang_cong_no_map()
    for kh in kh_list:
        info = kh_cong_no_map.get(kh.pk) or {}
        kh.tong_hoa_don = info.get('tong_hoa_don', Decimal('0'))
        kh.tong_da_thu = info.get('tong_da_thu', Decimal('0'))
        kh.con_no_hien_tai = info.get('con_no', Decimal('0'))
    hoa_don_list = _phieu_thu_hoa_don_queryset()
    if hoa_don_selected and hoa_don_selected.pk:
        hoa_don_list = (hoa_don_list | HoaDonBan.objects.select_related('khach_hang').filter(pk=hoa_don_selected.pk)).distinct()

    tai_khoan_list = list(TaiKhoanKeToan.objects.filter(trang_thai=True).order_by('ma_tk'))

    return {
        'kh_list': kh_list,
        'kh_cong_no_map': kh_cong_no_map,
        'hoa_don_list': hoa_don_list,
        'hoa_don_selected': hoa_don_selected,
        'tai_khoan_list': tai_khoan_list,
        'so_phieu_default': form_values.get('so_phieu') or _gen_so_don('PT'),
        'today': date.today(),
        'page_title': ('Cập nhật phiếu chi tiền' if editing else 'Lập phiếu chi tiền') if is_chi_mode else ('Cập nhật phiếu thu tiền' if editing else 'Lập phiếu thu tiền'),
        'active_menu': 'phieu_chi' if is_chi_mode else 'phieu_thu',
        'loai_form': 'chi' if is_chi_mode else 'thu',
        'form_values': form_values,
    }


@login_required
def phieu_thu_them(request, mode='thu'):
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    # Xử lý copy_from: nếu có tham số copy_from trên URL, lấy dữ liệu phiếu thu gốc để đổ vào form tạo mới
    copy_from = request.GET.get('copy_from')
    phieu_copy = None
    if copy_from:
        try:
            phieu_copy = PhieuThu.objects.get(pk=copy_from)
        except PhieuThu.DoesNotExist:
            phieu_copy = None
    if request.method == 'POST':
        data = request.POST
        loai_phieu_input = (data.get('loai_phieu_thu') or '1').strip()
        loai_phieu = loai_phieu_input if loai_phieu_input in ('1', '2') else ('2' if is_chi_mode else '1')
        hoa_don_id = (data.get('hoa_don') or None) if loai_phieu == '1' else None
        hoa_don_obj = HoaDonBan.objects.filter(pk=hoa_don_id).first() if (hoa_don_id and loai_phieu == '1') else None
        form_values = {
            'so_phieu': data.get('so_phieu', ''),
            'ngay_thu': data.get('ngay_thu', ''),
            'ngay_hach_toan': data.get('ngay_hach_toan', ''),
            'khach_hang': data.get('khach_hang', ''),
            'nguoi_nop_tien': data.get('nguoi_nop_tien', ''),
            'dia_chi': data.get('dia_chi', ''),
            'ly_do_nop': data.get('ly_do_nop', ''),
            'loai_phieu_thu': loai_phieu,
            'hinh_thuc_thu': data.get('hinh_thuc_thu', 'tien_mat'),
            'tong_thu': data.get('tong_thu', ''),
            'hoa_don': data.get('hoa_don', '') if loai_phieu == '1' else '',
            'so_tham_chieu': data.get('so_tham_chieu', ''),
            'trang_thai': data.get('trang_thai', '1'),
            'ghi_chu': data.get('ghi_chu', ''),
            'tk_no': data.get('tk_no', ''),
            'tk_co': data.get('tk_co', ''),
        }
        trang_thai_input = (data.get('trang_thai') or '1').strip()
        if trang_thai_input == '3':
            trang_thai_input = '2'
        trang_thai_luu = trang_thai_input if trang_thai_input in ('1', '2') else '1'
        form_values['trang_thai'] = trang_thai_luu

        try:
            tong_thu = _parse_decimal(data.get('tong_thu', 0))
        except InvalidOperation:
            messages.error(request, 'Số tiền thu không hợp lệ')
            return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        if tong_thu <= 0:
            messages.error(request, 'Số tiền thu phải lớn hơn 0')
            return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        kh_id = data.get('khach_hang') or (hoa_don_obj.khach_hang_id if hoa_don_obj else None)
        if not kh_id:
            messages.error(request, 'Vui lòng chọn khách hàng cho phiếu thu')
            return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        if loai_phieu == '1' and not is_chi_mode:
            if not hoa_don_obj:
                messages.error(request, 'Loại 1 yêu cầu chọn hóa đơn liên kết.')
                return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))
            
            if _is_hd_ghi_nhan_cong_no(hoa_don_obj):
                if tong_thu > Decimal(hoa_don_obj.con_no or 0):
                    messages.error(request, f'Số tiền thu vượt công nợ còn lại ({Decimal(hoa_don_obj.con_no or 0):,.0f} đ)')
                    return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        elif not is_chi_mode:
            kh_cong_no_map = _phieu_thu_khach_hang_cong_no_map()
            try:
                kh_id_int = int(kh_id)
            except (TypeError, ValueError):
                kh_id_int = 0
            con_no_kh = Decimal((kh_cong_no_map.get(kh_id_int) or {}).get('con_no', Decimal('0')))
            if tong_thu > con_no_kh:
                messages.error(request, f'Số tiền thu vượt công nợ còn lại của khách hàng ({con_no_kh:,.0f} đ)')
                return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        ghi_chu_raw = (data.get('ghi_chu') or '').strip()
        ghi_chu_luu = ghi_chu_raw
        if is_chi_mode:
            ghi_chu_luu = f"[LOAI_PHIEU:CHI] {ghi_chu_raw}"
        else:
            loai_map = {'1': 'Thu tiền chi tiết theo hóa đơn', '2': 'Thu của khách hàng'}
            loai_label = loai_map.get(loai_phieu, 'Thu tiền chi tiết theo hóa đơn')
            ghi_chu_luu = f"[LOAI_PHIEU_THU:{loai_phieu}] {loai_label}\n{ghi_chu_raw}".strip()

        pt = None
        last_error = None
        for _ in range(3):
            so_phieu = _ensure_unique_so_phieu_chi(form_values['so_phieu']) if is_chi_mode else _ensure_unique_so_phieu_thu(form_values['so_phieu'])
            try:
                with transaction.atomic():
                    pt = PhieuThu.objects.create(
                        so_phieu=so_phieu,
                        ngay_thu=data.get('ngay_thu') or date.today(),
                        khach_hang_id=kh_id,
                        hinh_thuc_thu=data.get('hinh_thuc_thu', 'tien_mat'),
                        so_tham_chieu=data.get('so_tham_chieu', ''),
                        tong_thu=tong_thu,
                        hoa_don=hoa_don_obj if loai_phieu == '1' else None,
                        trang_thai=trang_thai_luu,
                        ghi_chu=ghi_chu_luu,
                        nguoi_tao=request.user,
                    )
                break
            except IntegrityError as exc:
                last_error = exc
                form_values['so_phieu'] = ''
                pt = None

        if pt is None:
            messages.error(request, f'Không thể lưu phiếu thu do trùng số phiếu. {last_error}')
            form_values['so_phieu'] = _ensure_unique_so_phieu_chi() if is_chi_mode else _ensure_unique_so_phieu_thu()
            return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        if trang_thai_luu == '2' and pt.hoa_don and not is_chi_mode:
            da_thu_moi = Decimal(pt.hoa_don.da_thu or 0) + Decimal(pt.tong_thu or 0)
            pt.hoa_don.da_thu = da_thu_moi
            pt.hoa_don.con_no = Decimal(pt.hoa_don.tong_cong or 0) - da_thu_moi
            pt.hoa_don.save(update_fields=['da_thu', 'con_no'])

        if trang_thai_luu == '2':
            try:
                post_to_ledger('phieu_thu', pt.id, user=request.user)
            except LedgerPostingError as exc:
                messages.warning(request, f'Lỗi ghi sổ cái: {str(exc)}')

        messages.success(request, f'Đã tạo {"phiếu chi" if is_chi_mode else "phiếu thu"} {pt.so_phieu}')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')

    if phieu_copy:
        hoa_don_obj = getattr(phieu_copy, 'hoa_don', None)
        form_values = {
            'ngay_thu': phieu_copy.ngay_thu,
            'ngay_hach_toan': phieu_copy.ngay_hach_toan,
            'hinh_thuc_thu': phieu_copy.hinh_thuc_thu,
            'loai_phieu_thu': getattr(phieu_copy, 'loai_phieu_thu', '1'),
            'trang_thai': '1',
            'tong_thu': phieu_copy.tong_thu,
            'so_phieu': '',
            'hoa_don': str(phieu_copy.hoa_don.pk) if hoa_don_obj else '',
            'khach_hang': str(phieu_copy.khach_hang_id) if phieu_copy.khach_hang_id else '',
            'tk_no': phieu_copy.tk_no if hasattr(phieu_copy, 'tk_no') else '',
            'tk_co': phieu_copy.tk_co if hasattr(phieu_copy, 'tk_co') else '',
        }
        return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))
    else:
        hd_id = request.GET.get('hoa_don')
        hoa_don_obj = HoaDonBan.objects.filter(pk=hd_id).first() if hd_id else None
        loai_phieu_get = request.GET.get('loai_phieu_thu', '')
        if loai_phieu_get not in ('1', '2'):
            loai_phieu_get = '1' if (is_chi_mode and hd_id) else ('2' if is_chi_mode else '1')
        form_values = {
            'ngay_thu': date.today().isoformat(),
            'ngay_hach_toan': date.today().isoformat(),
            'hinh_thuc_thu': 'tien_mat',
            'loai_phieu_thu': loai_phieu_get,
            'trang_thai': '1',
            'tong_thu': str(int(Decimal(hoa_don_obj.con_no or 0))) if hoa_don_obj else '',
            'so_phieu': _ensure_unique_so_phieu_chi() if is_chi_mode else _ensure_unique_so_phieu_thu(),
            'hoa_don': str(hoa_don_obj.pk) if hoa_don_obj else '',
            'khach_hang': str(hoa_don_obj.khach_hang_id) if hoa_don_obj and hoa_don_obj.khach_hang_id else '',
            'tk_no': '',
            'tk_co': '',
        }
        return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))


@login_required
def phieu_thu_sua(request, pk, mode='thu'):
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    phieu = get_object_or_404(PhieuThu, pk=pk)
    if is_chi_mode != _is_phieu_chi_record(phieu):
        messages.error(request, 'Không tìm thấy chứng từ phù hợp.')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')

    if request.method == 'POST':
        data = request.POST
        loai_phieu_input = (data.get('loai_phieu_thu') or '1').strip()
        loai_phieu = loai_phieu_input if loai_phieu_input in ('1', '2') else ('2' if is_chi_mode else '1')
        hoa_don_obj = phieu.hoa_don
        form_values = {
            'so_phieu': phieu.so_phieu,
            'ngay_thu': data.get('ngay_thu', phieu.ngay_thu.isoformat() if phieu.ngay_thu else ''),
            'ngay_hach_toan': data.get('ngay_hach_toan', ''),
            'khach_hang': str(phieu.khach_hang_id) if phieu.khach_hang else '',
            'nguoi_nop_tien': data.get('nguoi_nop_tien', ''),
            'dia_chi': data.get('dia_chi', ''),
            'ly_do_nop': data.get('ly_do_nop', ''),
            'loai_phieu_thu': loai_phieu,
            'hinh_thuc_thu': data.get('hinh_thuc_thu', phieu.hinh_thuc_thu or 'tien_mat'),
            'tong_thu': data.get('tong_thu', str(phieu.tong_thu) if phieu.tong_thu else ''),
            'hoa_don': str(phieu.hoa_don_id) if phieu.hoa_don else '',
            'so_tham_chieu': data.get('so_tham_chieu', phieu.so_tham_chieu or ''),
            'trang_thai': data.get('trang_thai', phieu.trang_thai),
            'ghi_chu': data.get('ghi_chu', ''),
            'tk_no': data.get('tk_no', ''),
            'tk_co': data.get('tk_co', ''),
        }
        
        try:
            tong_thu = _parse_decimal(data.get('tong_thu') or (phieu.tong_thu or 0))
        except InvalidOperation:
            messages.error(request, 'Số tiền thu không hợp lệ')
            return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))
        
        if tong_thu <= 0:
            messages.error(request, 'Số tiền thu phải lớn hơn 0')
            return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))
        
        if loai_phieu == '1' and hoa_don_obj and not is_chi_mode:
            if _is_hd_ghi_nhan_cong_no(hoa_don_obj):
                if tong_thu > Decimal(hoa_don_obj.con_no or 0):
                    messages.error(request, f'Số tiền thu vượt công nợ còn lại ({Decimal(hoa_don_obj.con_no or 0):,.0f} đ)')
                    return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, hoa_don_obj, mode=mode))

        trang_thai_input = (data.get('trang_thai') or phieu.trang_thai or '1').strip()
        if trang_thai_input == '3':
            trang_thai_input = '2'
        trang_thai_luu = trang_thai_input if trang_thai_input in ('1', '2') else '1'
        
        phieu.ngay_thu = data.get('ngay_thu') or phieu.ngay_thu or date.today()
        phieu.hinh_thuc_thu = data.get('hinh_thuc_thu', 'tien_mat')
        phieu.so_tham_chieu = data.get('so_tham_chieu', '')
        phieu.tong_thu = tong_thu
        phieu.trang_thai = trang_thai_luu
        phieu.hoa_don = hoa_don_obj if loai_phieu == '1' else None
        
        ghi_chu = (data.get('ghi_chu') or '').strip()
        if ghi_chu:
            phieu.ghi_chu = ghi_chu
        
        phieu.save()

        if trang_thai_luu == '2':
            try:
                post_to_ledger('phieu_thu', phieu.id, user=request.user)
            except LedgerPostingError as exc:
                messages.warning(request, f'Lỗi ghi sổ cái: {str(exc)}')

        messages.success(request, f'Đã cập nhật {"phiếu chi" if is_chi_mode else "phiếu thu"} {phieu.so_phieu}')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')
    
    loai_phieu = '1' if phieu.hoa_don else '2'
    form_values = {
        'ngay_thu': phieu.ngay_thu.isoformat() if phieu.ngay_thu else date.today().isoformat(),
        'ngay_hach_toan': date.today().isoformat(),
        'hinh_thuc_thu': phieu.hinh_thuc_thu or 'tien_mat',
        'loai_phieu_thu': loai_phieu,
        'trang_thai': phieu.trang_thai,
        'tong_thu': str(phieu.tong_thu) if phieu.tong_thu else '',
        'so_phieu': phieu.so_phieu,
        'hoa_don': str(phieu.hoa_don_id) if phieu.hoa_don else '',
        'khach_hang': str(phieu.khach_hang_id) if phieu.khach_hang else '',
        'so_tham_chieu': phieu.so_tham_chieu or '',
        'ghi_chu': phieu.ghi_chu or '',
        'tk_no': '',
        'tk_co': '',
    }
    
    return render(request, 'ban_hang/phieu_thu_form.html', _build_phieu_thu_context(form_values, phieu.hoa_don, editing=True, mode=mode))


@login_required
def phieu_thu_xem(request, pk, mode='thu'):
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    phieu = get_object_or_404(PhieuThu, pk=pk)
    if is_chi_mode != _is_phieu_chi_record(phieu):
        messages.error(request, 'Không tìm thấy chứng từ phù hợp.')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')

    raw_note = (phieu.ghi_chu or '').strip()
    ghi_chu_hien_thi = raw_note
    if is_chi_mode and raw_note.startswith('[LOAI_PHIEU:CHI]'):
        ghi_chu_hien_thi = raw_note[len('[LOAI_PHIEU:CHI]'):].strip()

    # Tính TK dựa vào hình thức thanh toán
    hinh_thuc = str(phieu.hinh_thuc_thu or 'tien_mat').strip()
    tk_tien = '112' if hinh_thuc == 'chuyen_khoan' else '111'
    if is_chi_mode:
        phieu_tk_no = tk_tien
        phieu_tk_co = '131'
    else:
        phieu_tk_no = '131'
        phieu_tk_co = tk_tien

    return render(request, 'ban_hang/phieu_thu_detail.html', {
        'phieu': phieu,
        'is_chi_mode': is_chi_mode,
        'ghi_chu_hien_thi': ghi_chu_hien_thi,
        'phieu_tk_no': phieu_tk_no,
        'phieu_tk_co': phieu_tk_co,
    })


@login_required
def phieu_thu_xoa(request, pk, mode='thu'): 
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    phieu = get_object_or_404(PhieuThu, pk=pk)
    if is_chi_mode != _is_phieu_chi_record(phieu):
        messages.error(request, 'Không tìm thấy chứng từ phù hợp.')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')
    if request.method == 'POST':
        so_phieu = phieu.so_phieu
        don = phieu.hoa_don
        tong_thu = phieu.tong_thu
        phieu.delete()
        
        if don:
            if phieu.trang_thai == '2' and not is_chi_mode:
                don.da_thu = Decimal(don.da_thu or 0) - Decimal(tong_thu or 0)
                don.con_no = Decimal(don.tong_thanh_toan or 0) - Decimal(don.da_thu or 0)
                don.save(update_fields=['da_thu', 'con_no'])
        
        messages.success(request, f'Đã xóa {"phiếu chi" if is_chi_mode else "phiếu thu"} {so_phieu}')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')
    
    return render(request, 'ban_hang/phieu_thu_form.html', {
        'phieu': phieu,
        'confirm_delete': True,
        'so_phieu': phieu.so_phieu,
    })


@login_required
def phieu_thu_xoa_nhieu(request, mode='thu'):
    is_chi_mode = str(mode or 'thu').strip().lower() == 'chi'
    if request.method == 'POST':
        ids = request.POST.getlist('ids[]')
        phieus = PhieuThu.objects.filter(pk__in=ids)
        count = 0
        
        for phieu in phieus:
            if is_chi_mode != _is_phieu_chi_record(phieu):
                continue
            try:
                so_phieu = phieu.so_phieu
                don = phieu.hoa_don
                tong_thu = phieu.tong_thu
                
                phieu.delete()
                
                if don and phieu.trang_thai == '2' and not is_chi_mode:
                    don.da_thu = Decimal(don.da_thu or 0) - Decimal(tong_thu or 0)
                    don.con_no = Decimal(don.tong_thanh_toan or 0) - Decimal(don.da_thu or 0)
                    don.save(update_fields=['da_thu', 'con_no'])
                
                count += 1
            except Exception as exc:
                print(f'Lỗi xóa phiếu thu {so_phieu}: {exc}')
        
        messages.success(request, f'Đã xóa {count} {"phiếu chi" if is_chi_mode else "phiếu thu"}')
        return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')
    
    return redirect('bh_phieu_chi_list' if is_chi_mode else 'phieu_thu_list')


def _gen_so_phieu_doi_tra():
    return _next_code_from_queryset_values(
        PhieuTraHang.objects.filter(so_phieu__istartswith='DT').values_list('so_phieu', flat=True),
        'DT',
    )


def _returned_qty_map(hoa_don, exclude_phieu_id=None):
    qs = PhieuTraHang_CT.objects.filter(
        phieu_tra__hoa_don_goc=hoa_don,
        phieu_tra__trang_thai='2',
    )
    if exclude_phieu_id:
        qs = qs.exclude(phieu_tra_id=exclude_phieu_id)

    result = {}
    for row in qs.values('hoa_don_ct_goc_id').annotate(total=Sum('so_luong')):
        key = row.get('hoa_don_ct_goc_id')
        if not key:
            continue
        result[key] = int(Decimal(row.get('total') or 0))
    return result


def _apply_inventory_and_cong_no_for_doi_tra(phieu):
    rows = list(phieu.chi_tiet.select_related('hang_hoa', 'hang_hoa_doi', 'kho'))

    for ct in rows:
        if int(ct.so_luong_doi or 0) <= 0 or not ct.hang_hoa_doi_id:
            continue
        ton = TonKho.objects.filter(hang_hoa_id=ct.hang_hoa_doi_id, kho_id=ct.kho_id).first()
        if not ton or int(ton.so_luong or 0) < int(ct.so_luong_doi or 0):
            ma_hang = ct.hang_hoa_doi.ma_hang if ct.hang_hoa_doi else 'N/A'
            raise ValueError(f'Không đủ tồn kho cho hàng đổi {ma_hang}.')

    for ct in rows:
        ton_tra, _ = TonKho.objects.get_or_create(
            hang_hoa_id=ct.hang_hoa_id,
            kho_id=ct.kho_id,
            defaults={'so_luong': 0, 'gia_von_tb': 0},
        )
        ton_tra.so_luong = int(ton_tra.so_luong or 0) + int(ct.so_luong or 0)
        ton_tra.save(update_fields=['so_luong', 'ngay_cap_nhat'])

        vt = TonKhoViTri.objects.filter(hang_hoa_id=ct.hang_hoa_id, kho_id=ct.kho_id).exclude(vi_tri__ma_vi_tri__startswith='HL-').first()
        if vt:
            vt.so_luong = int(vt.so_luong or 0) + int(ct.so_luong or 0)
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])

        if int(ct.so_luong_doi or 0) > 0 and ct.hang_hoa_doi_id:
            ton_doi = TonKho.objects.get(hang_hoa_id=ct.hang_hoa_doi_id, kho_id=ct.kho_id)
            ton_doi.so_luong = int(ton_doi.so_luong or 0) - int(ct.so_luong_doi or 0)
            ton_doi.save(update_fields=['so_luong', 'ngay_cap_nhat'])

            con_lai = int(ct.so_luong_doi or 0)
            vt_rows = list(
                TonKhoViTri.objects.select_for_update().filter(hang_hoa_id=ct.hang_hoa_doi_id, kho_id=ct.kho_id, so_luong__gt=0)
                .exclude(vi_tri__ma_vi_tri__startswith='HL-')
                .order_by('-so_luong', 'id')
            )
            for vtd in vt_rows:
                if con_lai <= 0: break
                tru = min(int(vtd.so_luong or 0), con_lai)
                vtd.so_luong = int(vtd.so_luong or 0) - tru
                vtd.save(update_fields=['so_luong', 'ngay_cap_nhat'])
                con_lai -= tru

    if phieu.hoa_don_goc_id:
        don = phieu.hoa_don_goc
        don.con_no = Decimal(don.con_no or 0) + Decimal(phieu.chenh_lech_tien or 0)
        if don.con_no < 0:
            don.con_no = Decimal('0')
        don.save(update_fields=['con_no'])

    # Hoan tien mat tu dong: tao phieu chi khi phieu doi/tra da hoan tat va phat sinh tien hoan.
    so_tien_hoan = abs(Decimal(phieu.chenh_lech_tien or 0)) if Decimal(phieu.chenh_lech_tien or 0) < 0 else Decimal('0')
    if (
        so_tien_hoan > 0
        and str(phieu.hinh_thuc_hoan or '').strip() == '1'
        and phieu.khach_hang_id
    ):
        doi_tra_tag = f'[DOI_TRA_PHIEU:{phieu.pk}]'
        da_tao_phieu_chi = PhieuThu.objects.filter(ghi_chu__contains=doi_tra_tag).exists()
        if not da_tao_phieu_chi:
            PhieuThu.objects.create(
                so_phieu=_ensure_unique_so_phieu_chi(),
                ngay_thu=phieu.ngay_hach_toan or phieu.ngay_lap or date.today(),
                khach_hang_id=phieu.khach_hang_id,
                hinh_thuc_thu='tien_mat',
                so_tham_chieu='',
                tong_thu=so_tien_hoan,
                don_ban=None,
                trang_thai='2',
                ghi_chu=f'{doi_tra_tag} Hoan tien cho phieu doi/tra {phieu.so_phieu}\n[LOAI_PHIEU:CHI] Hoan tien doi/tra',
                nguoi_tao=phieu.nguoi_tao,
            )

    try:
        post_to_ledger('hang_ban_tra_lai', phieu.id, phieu.nguoi_tao)
    except LedgerPostingError:
        pass

    phieu.da_cap_nhat_kho_cong_no = True
    phieu.save(update_fields=['da_cap_nhat_kho_cong_no'])


def _rollback_inventory_and_cong_no_for_doi_tra(phieu):
    from apps.so_cai.models import JournalEntry
    from apps.kho.models import TonKhoViTri

    rows = list(phieu.chi_tiet.select_related('hang_hoa', 'hang_hoa_doi', 'kho'))

    for ct in rows:
        ton_tra = TonKho.objects.filter(hang_hoa_id=ct.hang_hoa_id, kho_id=ct.kho_id).first()
        if ton_tra:
            ton_tra.so_luong = int(ton_tra.so_luong or 0) - int(ct.so_luong or 0)
            if ton_tra.so_luong < 0:
                ton_tra.so_luong = 0
            ton_tra.save(update_fields=['so_luong', 'ngay_cap_nhat'])

        vt = TonKhoViTri.objects.filter(hang_hoa_id=ct.hang_hoa_id, kho_id=ct.kho_id).exclude(vi_tri__ma_vi_tri__startswith='HL-').first()
        if vt:
            vt.so_luong = int(vt.so_luong or 0) - int(ct.so_luong or 0)
            if vt.so_luong < 0:
                vt.so_luong = 0
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])

        if int(ct.so_luong_doi or 0) > 0 and ct.hang_hoa_doi_id:
            ton_doi, _ = TonKho.objects.get_or_create(
                hang_hoa_id=ct.hang_hoa_doi_id,
                kho_id=ct.kho_id,
                defaults={'so_luong': 0, 'gia_von_tb': 0},
            )
            ton_doi.so_luong = int(ton_doi.so_luong or 0) + int(ct.so_luong_doi or 0)
            ton_doi.save(update_fields=['so_luong', 'ngay_cap_nhat'])
            
            vt_doi = TonKhoViTri.objects.filter(hang_hoa_id=ct.hang_hoa_doi_id, kho_id=ct.kho_id).exclude(vi_tri__ma_vi_tri__startswith='HL-').first()
            if vt_doi:
                vt_doi.so_luong = int(vt_doi.so_luong or 0) + int(ct.so_luong_doi or 0)
                vt_doi.save(update_fields=['so_luong', 'ngay_cap_nhat'])

    if phieu.hoa_don_goc_id:
        don = phieu.hoa_don_goc
        don.con_no = Decimal(don.con_no or 0) - Decimal(phieu.chenh_lech_tien or 0)
        if don.con_no < 0:
            don.con_no = Decimal('0')
        don.save(update_fields=['con_no'])

    doi_tra_tag = f'[DOI_TRA_PHIEU:{phieu.pk}]'
    PhieuThu.objects.filter(ghi_chu__contains=doi_tra_tag).delete()
    JournalEntry.objects.filter(document_type='hang_ban_tra_lai', document_id=phieu.pk).delete()

    phieu.da_cap_nhat_kho_cong_no = False
    phieu.save(update_fields=['da_cap_nhat_kho_cong_no'])


def _build_doi_tra_context(form_values=None, phieu=None):
    form_values = form_values or {}
    form_values.setdefault('tk_no', '131')
    form_values.setdefault('tk_co', '131')
    hoa_don_qs = HoaDonBan.objects.select_related('khach_hang').filter(trang_thai__in=('2', '3')).order_by('-ngay_lap', '-id')
    if phieu and phieu.hoa_don_goc_id:
        hoa_don_qs = (hoa_don_qs | HoaDonBan.objects.select_related('khach_hang').filter(pk=phieu.hoa_don_goc_id)).distinct()

    return {
        'phieu': phieu,
        'form_values': form_values,
        'hoa_don_list': hoa_don_qs[:300],
        'hang_list': HangHoa.objects.all().order_by('ma_hang'),
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'tai_khoan_list': TaiKhoanKeToan.objects.filter(trang_thai=True).order_by('ma_tk'),
        'page_title': 'Sửa phiếu đổi trả' if phieu else 'Lập phiếu đổi trả',
        'active_menu': 'doi_tra',
        'today': date.today(),
        'is_locked': False,
    }


def _phieu_doi_tra_is_linked(phieu):
    if not phieu or not phieu.pk:
        return False
    doi_tra_tag = f'[DOI_TRA_PHIEU:{phieu.pk}]'
    return PhieuThu.objects.filter(ghi_chu__contains=doi_tra_tag).exists()


def _save_phieu_doi_tra_from_request(request, phieu=None):
    data = request.POST
    is_edit = bool(phieu and phieu.pk)

    if is_edit and phieu.da_cap_nhat_kho_cong_no:
        _rollback_inventory_and_cong_no_for_doi_tra(phieu)

    hoa_don_id = (data.get('hoa_don_goc') or '').strip()
    if not hoa_don_id:
        raise ValueError('Vui lòng chọn hóa đơn/phiếu xuất gốc')

    hoa_don = HoaDonBan.objects.select_related('khach_hang', 'don_ban').filter(pk=hoa_don_id).first()
    if not hoa_don:
        raise ValueError('Đơn hàng không tồn tại')

    hinh_thuc = (data.get('hinh_thuc_xu_ly') or 'tra_hang').strip()
    if hinh_thuc not in ('tra_hang', 'doi_hang'):
        hinh_thuc = 'tra_hang'

    hinh_thuc_hoan_raw = (data.get('hinh_thuc_hoan') or '1').strip()
    map_hinh_thuc_hoan = {
        'tien_mat': '1',
        'bu_tru_no': '2',
        'doi_hang': '3',
        '1': '1',
        '2': '2',
        '3': '3',
    }
    hinh_thuc_hoan = map_hinh_thuc_hoan.get(hinh_thuc_hoan_raw, '1')

    submit_action = (data.get('submit_action') or '').strip().lower()
    trang_thai = '2' if submit_action == 'confirm' else (data.get('trang_thai') or '1').strip()
    if trang_thai not in ('1', '2'):
        trang_thai = '1'

    source_rows = {ct.id: ct for ct in hoa_don.chi_tiet.select_related('hang_hoa', 'kho')}
    returned_map = _returned_qty_map(hoa_don, phieu.pk if phieu else None)

    ct_ids = data.getlist('hoa_don_ct_id[]')
    hang_tra_ids = data.getlist('hang_tra_id[]')
    kho_ids = data.getlist('kho_id[]')
    so_luong_tras = data.getlist('so_luong_tra[]')
    don_gia_tras = data.getlist('don_gia_tra[]')
    loai_hang_tras = data.getlist('loai_hang_tra[]')
    hang_doi_ids = data.getlist('hang_doi_id[]')
    so_luong_dois = data.getlist('so_luong_doi[]')
    gia_ban_dois = data.getlist('gia_ban_doi[]')
    include_rows = data.getlist('include_row[]')

    rows_payload = []
    max_len = max(len(hang_tra_ids), len(ct_ids), len(so_luong_tras))
    for i in range(max_len):
        include_flag = (include_rows[i] if i < len(include_rows) else '').strip()
        if include_flag not in ('1', 'true', 'on'):
            continue

        hang_tra_id = (hang_tra_ids[i] if i < len(hang_tra_ids) else '').strip()
        ct_id_raw = (ct_ids[i] if i < len(ct_ids) else '').strip()
        if not hang_tra_id and not ct_id_raw:
            continue

        try:
            so_luong_tra = int(_parse_decimal(so_luong_tras[i] if i < len(so_luong_tras) else 0))
        except (TypeError, ValueError):
            so_luong_tra = 0
        if so_luong_tra <= 0:
            raise ValueError('Số lượng trả phải lớn hơn 0')

        ct_goc = None
        if ct_id_raw:
            ct_goc = source_rows.get(int(ct_id_raw))
        if not ct_goc:
            for item in source_rows.values():
                if str(item.hang_hoa_id) == hang_tra_id:
                    ct_goc = item
                    break
        if not ct_goc:
            raise ValueError('Chỉ cho phép đổi/trả sản phẩm đã mua')

        da_tra = int(returned_map.get(ct_goc.id, 0))
        sl_mua = int(Decimal(ct_goc.so_luong or 0))
        if so_luong_tra > (sl_mua - da_tra):
            raise ValueError('Số lượng vượt quá số đã mua')

        kho_id = (kho_ids[i] if i < len(kho_ids) else '').strip() or str(ct_goc.kho_id)
        don_gia_tra = _parse_decimal(don_gia_tras[i] if i < len(don_gia_tras) else ct_goc.gia_ban)
        loai_hang_tra = (loai_hang_tras[i] if i < len(loai_hang_tras) else 'binh_thuong').strip()
        if loai_hang_tra not in ('binh_thuong', 'hang_loi'):
            loai_hang_tra = 'binh_thuong'

        hang_doi_id = (hang_doi_ids[i] if i < len(hang_doi_ids) else '').strip()
        so_luong_doi = int(_parse_decimal(so_luong_dois[i] if i < len(so_luong_dois) else 0))
        gia_ban_doi = _parse_decimal(gia_ban_dois[i] if i < len(gia_ban_dois) else 0)

        if hinh_thuc == 'doi_hang':
            if not hang_doi_id or so_luong_doi <= 0:
                raise ValueError('Hình thức đổi hàng yêu cầu chọn sản phẩm mới và số lượng đổi hợp lệ')
        else:
            hang_doi_id = ''
            so_luong_doi = 0
            gia_ban_doi = Decimal('0')

        rows_payload.append({
            'hoa_don_ct_goc_id': ct_goc.id,
            'kho_id': int(kho_id),
            'hang_hoa_id': int(hang_tra_id or ct_goc.hang_hoa_id),
            'so_luong': so_luong_tra,
            'don_gia': don_gia_tra,
            'loai_hang_tra': loai_hang_tra,
            'hang_hoa_doi_id': int(hang_doi_id) if hang_doi_id else None,
            'so_luong_doi': so_luong_doi,
            'gia_ban_doi': gia_ban_doi,
        })

    if not rows_payload:
        raise ValueError('Phiếu đổi trả phải có ít nhất 1 dòng hàng hóa hợp lệ')

    tong_tien_tra = sum(Decimal(r['so_luong']) * Decimal(r['don_gia']) for r in rows_payload)
    tong_tien_doi = sum(Decimal(r['so_luong_doi']) * Decimal(r['gia_ban_doi']) for r in rows_payload)
    chenh_lech = tong_tien_doi - tong_tien_tra
    tong_tien_hoan = abs(chenh_lech)

    phieu = phieu or PhieuTraHang()
    phieu.so_phieu = (data.get('so_phieu') or phieu.so_phieu or _gen_so_phieu_doi_tra()).strip()
    phieu.ngay_lap = _parse_date(data.get('ngay_lap') or date.today())
    phieu.ngay_hach_toan = _parse_date(data.get('ngay_hach_toan') or phieu.ngay_lap)
    phieu.ngay_tra = phieu.ngay_lap
    phieu.hoa_don_goc = hoa_don
    phieu.hoa_don_goc = hoa_don.don_ban
    phieu.khach_hang = hoa_don.khach_hang
    phieu.tk_no = (data.get('tk_no') or '131').strip()
    phieu.tk_co = (data.get('tk_co') or '131').strip()
    phieu.dien_giai = (data.get('dien_giai') or '').strip()
    phieu.ly_do_tra = (data.get('ly_do_tra') or 'Đổi/trả theo hóa đơn gốc').strip()
    phieu.hinh_thuc_xu_ly = hinh_thuc
    phieu.hinh_thuc_hoan = hinh_thuc_hoan
    phieu.tong_tien_tra = tong_tien_tra
    phieu.tong_tien_doi = tong_tien_doi
    phieu.chenh_lech_tien = chenh_lech
    phieu.tong_tien_hoan = tong_tien_hoan
    phieu.trang_thai = trang_thai
    if not phieu.pk:
        phieu.nguoi_tao = request.user if request.user.is_authenticated else None
    phieu.save()

    phieu.chi_tiet.all().delete()
    for row in rows_payload:
        PhieuTraHang_CT.objects.create(phieu_tra=phieu, **row)

    if trang_thai == '2' and not phieu.da_cap_nhat_kho_cong_no:
        _apply_inventory_and_cong_no_for_doi_tra(phieu)

    return phieu


@login_required
def phieu_doi_tra_list(request):
    q = request.GET.get('q', '').strip()
    trang_thai = request.GET.get('trang_thai', '').strip()

    items = PhieuTraHang.objects.select_related('khach_hang', 'hoa_don_goc').order_by('-ngay_lap', '-id')
    if q:
        items = items.filter(
            Q(so_phieu__icontains=q)
            | Q(khach_hang__ma_kh__icontains=q)
            | Q(khach_hang__ten_kh__icontains=q)
            | Q(hoa_don_goc__so_hoa_don__icontains=q)
        )
    if trang_thai:
        items = items.filter(trang_thai=trang_thai)

    return render(request, 'ban_hang/doi_tra_list.html', {
        'items': items[:200],
        'q': q,
        'trang_thai_filter': trang_thai,
        'page_title': 'Quản lý đổi trả',
        'active_menu': 'doi_tra',
    })


@login_required
def phieu_doi_tra_them(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                phieu = _save_phieu_doi_tra_from_request(request)
        except (ValueError, IntegrityError) as exc:
            messages.error(request, str(exc))
            return render(request, 'ban_hang/doi_tra_form.html', _build_doi_tra_context(request.POST.dict()))

        messages.success(request, f'Đã tạo phiếu đổi trả {phieu.so_phieu}')
        return redirect('phieu_doi_tra_list')

    hoa_don_goc = (request.GET.get('hoa_don_goc') or '').strip()
    copy_from_id = (request.GET.get('copy_from') or '').strip()
    phieu_copy = None
    if copy_from_id and copy_from_id.isdigit():
        phieu_copy = PhieuTraHang.objects.select_related('hoa_don_goc', 'khach_hang').filter(pk=copy_from_id).first()

    if phieu_copy:
        form_values = {
            'so_phieu': _gen_so_phieu_doi_tra(),
            'ngay_lap': date.today().isoformat(),
            'ngay_hach_toan': date.today().isoformat(),
            'hoa_don_goc': str(phieu_copy.hoa_don_goc_id or ''),
            'khach_hang': str(phieu_copy.khach_hang_id or ''),
            'tk_no': phieu_copy.tk_no or '131',
            'tk_co': phieu_copy.tk_co or '131',
            'dien_giai': phieu_copy.dien_giai or '',
            'ly_do_tra': phieu_copy.ly_do_tra or '',
            'hinh_thuc_xu_ly': phieu_copy.hinh_thuc_xu_ly or 'tra_hang',
            'hinh_thuc_hoan': {'tien_mat': '1', 'bu_tru_no': '2', 'doi_hang': '3'}.get(phieu_copy.hinh_thuc_hoan, '1'),
            'trang_thai': '1',
        }
        chi_tiet_copy = list(phieu_copy.chi_tiet.select_related('hang_hoa', 'hang_hoa_doi', 'kho', 'hoa_don_ct_goc'))
        return render(request, 'ban_hang/doi_tra_form.html', {
            **_build_doi_tra_context(form_values),
            'chi_tiet_copy': chi_tiet_copy,
            'is_copy_mode': True,
        })

    form_values = {
        'so_phieu': _gen_so_phieu_doi_tra(),
        'ngay_lap': date.today().isoformat(),
        'ngay_hach_toan': date.today().isoformat(),
        'hoa_don_goc': hoa_don_goc,
        'tk_no': '131',
        'tk_co': '131',
        'hinh_thuc_xu_ly': 'tra_hang',
        'hinh_thuc_hoan': '1',
        'trang_thai': '1',
    }
    return render(request, 'ban_hang/doi_tra_form.html', _build_doi_tra_context(form_values))


@login_required
def phieu_doi_tra_sua(request, pk):
    phieu = get_object_or_404(PhieuTraHang.objects.select_related('hoa_don_goc', 'khach_hang'), pk=pk)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                _save_phieu_doi_tra_from_request(request, phieu)
        except (ValueError, IntegrityError) as exc:
            messages.error(request, str(exc))
            return render(request, 'ban_hang/doi_tra_form.html', _build_doi_tra_context(request.POST.dict(), phieu))

        messages.success(request, f'Đã cập nhật phiếu đổi trả {phieu.so_phieu}')
        return redirect('phieu_doi_tra_list')

    form_values = {
        'so_phieu': phieu.so_phieu,
        'ngay_lap': phieu.ngay_lap.isoformat() if phieu.ngay_lap else '',
        'ngay_hach_toan': phieu.ngay_hach_toan.isoformat() if phieu.ngay_hach_toan else '',
        'hoa_don_goc': str(phieu.hoa_don_goc_id or ''),
        'tk_no': phieu.tk_no,
        'tk_co': phieu.tk_co,
        'dien_giai': phieu.dien_giai,
        'ly_do_tra': phieu.ly_do_tra,
        'hinh_thuc_xu_ly': phieu.hinh_thuc_xu_ly,
        'hinh_thuc_hoan': {'tien_mat': '1', 'bu_tru_no': '2', 'doi_hang': '3'}.get(phieu.hinh_thuc_hoan, phieu.hinh_thuc_hoan),
        'trang_thai': phieu.trang_thai,
    }
    return render(request, 'ban_hang/doi_tra_form.html', {
        **_build_doi_tra_context(form_values, phieu),
        'chi_tiet': phieu.chi_tiet.select_related('hang_hoa', 'hang_hoa_doi', 'kho', 'hoa_don_ct_goc'),
    })


@login_required
def phieu_doi_tra_xoa(request, pk):
    phieu = get_object_or_404(PhieuTraHang, pk=pk)
    if request.method == 'POST':
        if _phieu_doi_tra_is_linked(phieu):
            messages.error(request, 'Phiếu đổi trả đã có tham chiếu, không được phép xóa')
            return redirect('phieu_doi_tra_list')

        if phieu.da_cap_nhat_kho_cong_no:
            _rollback_inventory_and_cong_no_for_doi_tra(phieu)

        so_phieu = phieu.so_phieu
        phieu.delete()
        messages.success(request, f'Đã xóa phiếu đổi trả {so_phieu}')
    return redirect('phieu_doi_tra_list')


@login_required
def phieu_doi_tra_xoa_nhieu(request):
    if request.method != 'POST':
        return redirect('phieu_doi_tra_list')

    ids = request.POST.getlist('ids[]') or request.POST.getlist('ids')
    if not ids:
        messages.error(request, 'Vui lòng chọn ít nhất 1 phiếu đổi trả để xóa.')
        return redirect('phieu_doi_tra_list')

    items = list(PhieuTraHang.objects.filter(pk__in=ids))
    deleted = 0
    blocked_codes = []

    for item in items:
        if _phieu_doi_tra_is_linked(item):
            blocked_codes.append(item.so_phieu)
            continue

        if item.da_cap_nhat_kho_cong_no:
            _rollback_inventory_and_cong_no_for_doi_tra(item)

        item.delete()
        deleted += 1

    if deleted:
        messages.success(request, f'Đã xóa {deleted} phiếu đổi trả.')

    if blocked_codes:
        preview = ', '.join(blocked_codes[:5])
        suffix = '...' if len(blocked_codes) > 5 else ''
        messages.error(
            request,
            f'Không thể xóa {len(blocked_codes)} phiếu đã có tham chiếu: {preview}{suffix}',
        )

    if not deleted and not blocked_codes:
        messages.error(request, 'Không có phiếu hợp lệ để xóa.')
    return redirect('phieu_doi_tra_list')


@login_required
def doi_tra_hoa_don_detail_api(request, pk):
    hoa_don = HoaDonBan.objects.select_related('khach_hang', 'don_ban').filter(pk=pk, trang_thai__in=('2', '3')).first()
    if not hoa_don:
        return JsonResponse({'error': 'Đơn hàng không tồn tại'}, status=404)

    returned_map = _returned_qty_map(hoa_don)
    rows = []
    for ct in hoa_don.chi_tiet.select_related('hang_hoa__don_vi_tinh', 'kho').all():
        sl_mua = int(Decimal(ct.so_luong or 0))
        sl_da_tra = int(returned_map.get(ct.id, 0))
        sl_con_lai = max(0, sl_mua - sl_da_tra)
        rows.append({
            'hoa_don_ct_id': ct.id,
            'hang_tra_id': ct.hang_hoa_id,
            'ma_hang': ct.hang_hoa.ma_hang,
            'ten_hang': ct.hang_hoa.ten_hang,
            'don_vi_tinh': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa and ct.hang_hoa.don_vi_tinh else '',
            'kho_id': ct.kho_id,
            'ma_kho': ct.kho.ma_kho,
            'so_luong_mua': sl_mua,
            'so_luong_da_tra': sl_da_tra,
            'so_luong_con_lai': sl_con_lai,
            'don_gia_tra': int(Decimal(ct.gia_ban or 0)),
        })

    return JsonResponse({
        'hoa_don_id': hoa_don.id,
        'so_hoa_don': hoa_don.so_hoa_don,
        'khach_hang_id': hoa_don.khach_hang_id,
        'ma_kh': hoa_don.khach_hang.ma_kh if hoa_don.khach_hang else '',
        'ten_kh': hoa_don.khach_hang.ten_kh if hoa_don.khach_hang else '',
        'don_ban_goc_id': hoa_don.don_ban_id,
        'rows': rows,
    })


# ─── BÁO CÁO BÁN HÀNG ───────────────────────────────────────
@login_required
def bao_cao_doanh_thu(request):
    from .revenue_report_service import (RevenueReportFilters,
                                         RevenueReportService)

    payload = {
        'from_date': request.GET.get('from_date', ''),
        'to_date': request.GET.get('to_date', ''),
        'date_type': request.GET.get('date_type', 'chung_tu'),
        'customer_id': request.GET.get('customer_id', ''),
        'salesperson_code': request.GET.get('salesperson_code', ''),
        'product_id': request.GET.get('product_id', ''),
        'group_id': request.GET.get('group_id', ''),
        'warehouse_id': request.GET.get('warehouse_id', ''),
    }
    group_by = (request.GET.get('group_by') or 'day').strip()

    try:
        filters = RevenueReportFilters.from_payload(payload)
    except ValueError as exc:
        messages.error(request, str(exc))
        filters = RevenueReportFilters.from_payload({})
        group_by = 'day'

    service = RevenueReportService(filters)
    try:
        summary_data = service.get_revenue_summary(group_by=group_by)
    except ValueError as exc:
        messages.error(request, str(exc))
        group_by = 'day'
        summary_data = service.get_revenue_summary(group_by=group_by)

    customer_data = service.get_revenue_by_customer()
    product_data = service.get_revenue_by_product()
    salesperson_data = service.get_revenue_by_salesperson()

    export_params = request.GET.copy()
    export_params.pop('report_type', None)

    salesperson_codes = [
        code for code in (
            HoaDonBan.objects
            .exclude(ma_nv_ban_hang='')
            .values_list('ma_nv_ban_hang', flat=True)
            .distinct()
            .order_by('ma_nv_ban_hang')
        )
        if code
    ]

    context = {
        'filters': filters,
        'group_by': group_by,
        'summary_data': summary_data,
        'customer_data': customer_data,
        'product_data': product_data,
        'salesperson_data': salesperson_data,
        'khach_hang_list': KhachHang.objects.filter(trang_thai=True).order_by('ma_kh'),
        'hang_hoa_list': HangHoa.objects.filter(trang_thai='dang_ban').order_by('ma_hang'),
        'nhom_hang_list': NhomHang.objects.order_by('ten_nhom'),
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'salesperson_codes': salesperson_codes,
        'export_params': export_params.urlencode(),
        'page_title': 'Báo cáo doanh thu',
        'active_menu': 'bao_cao_bh',
    }
    return render(request, 'ban_hang/bao_cao_doanh_thu.html', context)


# ─── CÔNG NỢ KHÁCH HÀNG ─────────────────────────────────────
@login_required
def cong_no_kh(request):
    today = timezone.localdate()
    ky_tu_ngay = _parse_date(request.GET.get('tu_ngay') or '', today.replace(day=1))
    ky_den_ngay = _parse_date(request.GET.get('den_ngay') or '', today)

    if request.method == 'POST':
        bat_canh_bao = request.POST.get('bat_canh_bao_qua_han') == '1'
        config, _ = CongNoCanhBaoConfig.objects.get_or_create(user=request.user)
        config.bat_canh_bao_qua_han = bat_canh_bao
        config.save(update_fields=['bat_canh_bao_qua_han', 'ngay_cap_nhat'])
        messages.success(request, 'Đã lưu thiết lập cảnh báo nợ quá hạn.')
        return redirect(request.POST.get('next') or 'cong_no_kh')

    config, _ = CongNoCanhBaoConfig.objects.get_or_create(user=request.user)
    ma_kh = request.GET.get('ma_kh', '').strip()
    ten_kh = request.GET.get('ten_kh', '').strip()
    so_dien_thoai = request.GET.get('so_dien_thoai', '').strip()
    trang_thai_filter = request.GET.get('trang_thai', '').strip()
    tu_ngay = request.GET.get('tu_ngay', '').strip()
    den_ngay = request.GET.get('den_ngay', '').strip()
    selected_khach_hang_id = request.GET.get('khach_hang', '').strip()

    don_qs = DonBan.objects.select_related('khach_hang').prefetch_related('hoa_don_lien_ket').filter(
        khach_hang__isnull=False,
        khach_hang__trang_thai=True,
        hoa_don_lien_ket__tk_no='131',
        hoa_don_lien_ket__trang_thai__in=('2', '3'),
    ).distinct()
    if ma_kh:
        don_qs = don_qs.filter(khach_hang__ma_kh__icontains=ma_kh)
    if ten_kh:
        don_qs = don_qs.filter(khach_hang__ten_kh__icontains=ten_kh)
    if so_dien_thoai:
        don_qs = don_qs.filter(khach_hang__so_dien_thoai__icontains=so_dien_thoai)
    if tu_ngay:
        don_qs = don_qs.filter(ngay_ban__gte=tu_ngay)
    if den_ngay:
        don_qs = don_qs.filter(ngay_ban__lte=den_ngay)

    thu_theo_kh_map = {
        row['khach_hang_id']: Decimal(row['tong'] or 0)
        for row in (
            PhieuThu.objects.filter(
                trang_thai='2',
                hoa_don__isnull=True,
                ghi_chu__icontains='[LOAI_PHIEU_THU:2]',
                khach_hang__isnull=False,
            )
            .values('khach_hang_id')
            .annotate(tong=Sum('tong_thu'))
        )
    }

    tong_hop_map = {}
    for don in don_qs:
        kh = don.khach_hang
        if not kh:
            continue

        da_thanh_toan_don = Decimal(don.da_thu or 0)
        tong_hoa_don_don = Decimal(don.tong_thanh_toan or 0)
        cong_no_don = tong_hoa_don_don - da_thanh_toan_don
        if cong_no_don < 0:
            cong_no_don = Decimal('0')
        han_thanh_toan = don.han_thanh_toan
        if not han_thanh_toan and kh.so_ngay_no_max and don.ngay_ban:
            han_thanh_toan = don.ngay_ban + timedelta(days=int(kh.so_ngay_no_max or 0))

        row = tong_hop_map.get(kh.pk)
        if not row:
            row = {
                'khach_hang_id': kh.pk,
                'ma_kh': kh.ma_kh,
                'ten_kh': kh.ten_kh,
                'so_dien_thoai': kh.so_dien_thoai,
                'loai_kh': kh.loai_kh,
                'tong_gia_tri_hoa_don': Decimal('0'),
                'tong_da_thanh_toan': Decimal('0'),
                'thu_theo_khach_hang': Decimal('0'),
                'cong_no_hien_tai': Decimal('0'),
                'han_thanh_toan': None,
                'so_ngay_qua_han': 0,
                'trang_thai': 'khong_no',
            }
            tong_hop_map[kh.pk] = row

        row['tong_gia_tri_hoa_don'] += tong_hoa_don_don
        row['tong_da_thanh_toan'] += da_thanh_toan_don

        if cong_no_don > 0 and han_thanh_toan:
            if row['han_thanh_toan'] is None or han_thanh_toan < row['han_thanh_toan']:
                row['han_thanh_toan'] = han_thanh_toan
            if today > han_thanh_toan:
                row['so_ngay_qua_han'] = max(row['so_ngay_qua_han'], (today - han_thanh_toan).days)

    data = []
    for row in tong_hop_map.values():
        row['thu_theo_khach_hang'] = thu_theo_kh_map.get(row['khach_hang_id'], Decimal('0'))
        row['tong_da_thanh_toan'] += row['thu_theo_khach_hang']
        row['cong_no_hien_tai'] = row['tong_gia_tri_hoa_don'] - row['tong_da_thanh_toan']
        if row['cong_no_hien_tai'] < 0:
            row['cong_no_hien_tai'] = Decimal('0')

        if row['cong_no_hien_tai'] <= 0:
            row['trang_thai'] = 'khong_no'
            row['so_ngay_qua_han'] = 0
            row['han_thanh_toan'] = None
        elif row['so_ngay_qua_han'] > 0:
            row['trang_thai'] = 'qua_han'
        else:
            row['trang_thai'] = 'con_no'

        if trang_thai_filter and row['trang_thai'] != trang_thai_filter:
            continue
        data.append(row)

    data.sort(key=lambda x: (0 if x['trang_thai'] == 'qua_han' else 1, -x['so_ngay_qua_han'], -x['cong_no_hien_tai']))

    chi_tiet_cong_no = []
    khach_hang_duoc_chon = None
    tong_chi_tiet = {
        'tong_tien': Decimal('0'),
        'da_thanh_toan': Decimal('0'),
        'con_no': Decimal('0'),
    }
    thu_theo_kh_bo_sung = Decimal('0')
    so_chi_tiet_cong_no = []
    tong_ps_no = Decimal('0')
    tong_ps_co = Decimal('0')
    so_du_dau_ky = Decimal('0')
    so_du_cuoi_ky = Decimal('0')
    if selected_khach_hang_id:
        try:
            selected_khach_hang_id_int = int(selected_khach_hang_id)
        except (TypeError, ValueError):
            selected_khach_hang_id_int = 0
        thu_theo_kh_bo_sung = thu_theo_kh_map.get(selected_khach_hang_id_int, Decimal('0'))

        chi_tiet_qs = don_qs.filter(khach_hang_id=selected_khach_hang_id).order_by('-ngay_ban', '-id')
        if chi_tiet_qs.exists():
            khach_hang_duoc_chon = chi_tiet_qs.first().khach_hang
        for don in chi_tiet_qs:
            han_thanh_toan = don.han_thanh_toan
            if not han_thanh_toan and don.khach_hang and don.khach_hang.so_ngay_no_max and don.ngay_ban:
                han_thanh_toan = don.ngay_ban + timedelta(days=int(don.khach_hang.so_ngay_no_max or 0))

            con_no = Decimal(don.con_no or 0)
            so_ngay_qua_han = 0
            if con_no > 0 and han_thanh_toan and today > han_thanh_toan:
                so_ngay_qua_han = (today - han_thanh_toan).days

            if con_no <= 0:
                trang_thai = 'khong_no'
            elif so_ngay_qua_han > 0:
                trang_thai = 'qua_han'
            else:
                trang_thai = 'con_no'

            hoa_don_lien_ket = don.hoa_don_lien_ket.order_by('-ngay_lap', '-id').first()
            ma_hoa_don = hoa_don_lien_ket.so_hoa_don if hoa_don_lien_ket else don.so_don
            ngay_hoa_don = hoa_don_lien_ket.ngay_lap if hoa_don_lien_ket else don.ngay_ban

            row = {
                'ma_hoa_don': ma_hoa_don,
                'ngay_hoa_don': ngay_hoa_don,
                'tong_tien': Decimal(don.tong_thanh_toan or 0),
                'da_thanh_toan': Decimal(don.da_thu or 0),
                'con_no': con_no,
                'han_thanh_toan': han_thanh_toan,
                'so_ngay_qua_han': so_ngay_qua_han,
                'trang_thai': trang_thai,
            }
            chi_tiet_cong_no.append(row)
            tong_chi_tiet['tong_tien'] += row['tong_tien']
            tong_chi_tiet['da_thanh_toan'] += row['da_thanh_toan']
            tong_chi_tiet['con_no'] += row['con_no']

        # Sổ chi tiết công nợ theo kỳ: mô phỏng màn hình "Sổ chi tiết công nợ của một khách hàng"
        dong_so = 1
        phat_sinh_items = []

        don_lich_su_qs = DonBan.objects.select_related('khach_hang').prefetch_related('hoa_don_lien_ket').filter(
            khach_hang_id=selected_khach_hang_id_int,
            hoa_don_lien_ket__tk_no='131',
            hoa_don_lien_ket__trang_thai__in=('2', '3'),
        ).distinct().order_by('ngay_ban', 'id')

        for don in don_lich_su_qs:
            hoa_don_lien_ket = don.hoa_don_lien_ket.order_by('-ngay_lap', '-id').first()
            ngay_ct = hoa_don_lien_ket.ngay_lap if hoa_don_lien_ket and hoa_don_lien_ket.ngay_lap else (don.ngay_ban or don.ngay_chung_tu)
            so_ct = hoa_don_lien_ket.so_hoa_don if hoa_don_lien_ket and hoa_don_lien_ket.so_hoa_don else don.so_don
            tk_doi_ung = (hoa_don_lien_ket.tk_co if hoa_don_lien_ket and hoa_don_lien_ket.tk_co else '51111')
            so_tien = Decimal(don.tong_thanh_toan or 0)

            if ngay_ct and ngay_ct < ky_tu_ngay:
                so_du_dau_ky += so_tien
                continue
            if ngay_ct and (ngay_ct < ky_tu_ngay or ngay_ct > ky_den_ngay):
                continue

            phat_sinh_items.append({
                'ngay_ghi_so': ngay_ct,
                'ngay_lap': ngay_ct,
                'ma_ct': 'HD',
                'so_ct': so_ct,
                'dien_giai': (don.ghi_chu or 'Ghi nhận doanh thu hóa đơn')[:200],
                'tk_doi_ung': tk_doi_ung,
                'phat_sinh_no': so_tien,
                'phat_sinh_co': Decimal('0'),
            })

        phieu_thu_lich_su_qs = PhieuThu.objects.select_related('khach_hang', 'don_ban').filter(
            khach_hang_id=selected_khach_hang_id_int,
            trang_thai='2',
        ).order_by('ngay_thu', 'id')

        for pt in phieu_thu_lich_su_qs:
            so_tien = Decimal(pt.tong_thu or 0)
            ngay_ct = pt.ngay_thu
            tk_doi_ung = '1121' if pt.hinh_thuc_thu == 'chuyen_khoan' else '1111'
            dien_giai = (pt.ghi_chu or 'Thu tiền khách hàng')[:200]

            if ngay_ct and ngay_ct < ky_tu_ngay:
                so_du_dau_ky -= so_tien
                continue
            if ngay_ct and (ngay_ct < ky_tu_ngay or ngay_ct > ky_den_ngay):
                continue

            phat_sinh_items.append({
                'ngay_ghi_so': ngay_ct,
                'ngay_lap': ngay_ct,
                'ma_ct': 'PT',
                'so_ct': pt.so_phieu,
                'dien_giai': dien_giai,
                'tk_doi_ung': tk_doi_ung,
                'phat_sinh_no': Decimal('0'),
                'phat_sinh_co': so_tien,
            })

        phat_sinh_items.sort(key=lambda x: (x['ngay_lap'] or date.min, x['ma_ct'], x['so_ct']))
        for row in phat_sinh_items:
            tong_ps_no += row['phat_sinh_no']
            tong_ps_co += row['phat_sinh_co']
            row['stt'] = dong_so
            dong_so += 1
            so_chi_tiet_cong_no.append(row)

        so_du_cuoi_ky = so_du_dau_ky + tong_ps_no - tong_ps_co

    query_without_selected = request.GET.copy()
    query_without_selected.pop('khach_hang', None)

    return render(request, 'ban_hang/cong_no_kh.html', {
        'data': data,
        'chi_tiet_cong_no': chi_tiet_cong_no,
        'khach_hang_duoc_chon': khach_hang_duoc_chon,
        'tong_chi_tiet': tong_chi_tiet,
        'thu_theo_kh_bo_sung': thu_theo_kh_bo_sung,
        'so_chi_tiet_cong_no': so_chi_tiet_cong_no,
        'so_du_dau_ky': so_du_dau_ky,
        'tong_ps_no': tong_ps_no,
        'tong_ps_co': tong_ps_co,
        'so_du_cuoi_ky': so_du_cuoi_ky,
        'ky_tu_ngay': ky_tu_ngay,
        'ky_den_ngay': ky_den_ngay,
        'bat_canh_bao_qua_han': config.bat_canh_bao_qua_han,
        'selected_khach_hang_id': selected_khach_hang_id,
        'query_without_selected': query_without_selected.urlencode(),
        'filters': {
            'ma_kh': ma_kh,
            'ten_kh': ten_kh,
            'so_dien_thoai': so_dien_thoai,
            'trang_thai': trang_thai_filter,
            'tu_ngay': tu_ngay,
            'den_ngay': den_ngay,
        },
        'page_title': 'Quản lý công nợ khách hàng',
        'active_menu': 'cong_no_kh',
    })


@login_required
def bao_cao_cong_no_kh_chi_tiet(request):
    today = timezone.localdate()
    tu_ngay = _parse_date(request.GET.get('tu_ngay') or '', today.replace(day=1))
    den_ngay = _parse_date(request.GET.get('den_ngay') or '', today)
    selected_khach_hang_id = (request.GET.get('khach_hang') or '').strip()

    khach_hang_list = KhachHang.objects.filter(trang_thai=True).order_by('ma_kh')
    khach_hang_duoc_chon = None
    if selected_khach_hang_id:
        khach_hang_duoc_chon = khach_hang_list.filter(pk=selected_khach_hang_id).first()

    so_chi_tiet_cong_no = []
    so_du_dau_ky = Decimal('0')
    tong_ps_no = Decimal('0')
    tong_ps_co = Decimal('0')

    if khach_hang_duoc_chon:
        don_lich_su_qs = DonBan.objects.select_related('khach_hang').prefetch_related('hoa_don_lien_ket').filter(
            khach_hang_id=khach_hang_duoc_chon.pk,
            hoa_don_lien_ket__tk_no='131',
            hoa_don_lien_ket__trang_thai__in=('2', '3'),
        ).distinct().order_by('ngay_ban', 'id')

        for don in don_lich_su_qs:
            hoa_don_lien_ket = don.hoa_don_lien_ket.order_by('-ngay_lap', '-id').first()
            ngay_ct = hoa_don_lien_ket.ngay_lap if hoa_don_lien_ket and hoa_don_lien_ket.ngay_lap else (don.ngay_ban or don.ngay_chung_tu)
            so_tien = Decimal(don.tong_thanh_toan or 0)
            if ngay_ct and ngay_ct < tu_ngay:
                so_du_dau_ky += so_tien
                continue
            if ngay_ct and (ngay_ct < tu_ngay or ngay_ct > den_ngay):
                continue
            so_chi_tiet_cong_no.append({
                'ngay_ghi_so': ngay_ct,
                'ngay_lap': ngay_ct,
                'ma_ct': 'HD',
                'so_ct': hoa_don_lien_ket.so_hoa_don if hoa_don_lien_ket and hoa_don_lien_ket.so_hoa_don else don.so_don,
                'dien_giai': (don.ghi_chu or 'Ghi nhận doanh thu hóa đơn')[:200],
                'tk_doi_ung': hoa_don_lien_ket.tk_co if hoa_don_lien_ket and hoa_don_lien_ket.tk_co else '51111',
                'phat_sinh_no': so_tien,
                'phat_sinh_co': Decimal('0'),
            })

        phieu_thu_qs = PhieuThu.objects.select_related('khach_hang', 'don_ban').filter(
            khach_hang_id=khach_hang_duoc_chon.pk,
            trang_thai='2',
        ).order_by('ngay_thu', 'id')

        for pt in phieu_thu_qs:
            ngay_ct = pt.ngay_thu
            so_tien = Decimal(pt.tong_thu or 0)
            if ngay_ct and ngay_ct < tu_ngay:
                so_du_dau_ky -= so_tien
                continue
            if ngay_ct and (ngay_ct < tu_ngay or ngay_ct > den_ngay):
                continue
            so_chi_tiet_cong_no.append({
                'ngay_ghi_so': ngay_ct,
                'ngay_lap': ngay_ct,
                'ma_ct': 'PT',
                'so_ct': pt.so_phieu,
                'dien_giai': (pt.ghi_chu or 'Thu tiền khách hàng')[:200],
                'tk_doi_ung': '1121' if pt.hinh_thuc_thu == 'chuyen_khoan' else '1111',
                'phat_sinh_no': Decimal('0'),
                'phat_sinh_co': so_tien,
            })

        so_chi_tiet_cong_no.sort(key=lambda x: (x['ngay_lap'] or date.min, x['ma_ct'], x['so_ct']))
        for row in so_chi_tiet_cong_no:
            tong_ps_no += row['phat_sinh_no']
            tong_ps_co += row['phat_sinh_co']
        for idx, row in enumerate(so_chi_tiet_cong_no, start=1):
            row['stt'] = idx

    so_du_cuoi_ky = so_du_dau_ky + tong_ps_no - tong_ps_co

    return render(request, 'ban_hang/bao_cao_cong_no_kh_chi_tiet.html', {
        'page_title': 'Báo cáo công nợ chi tiết khách hàng',
        'active_menu': 'bao_cao_cong_no_kh_ct',
        'khach_hang_list': khach_hang_list,
        'khach_hang_duoc_chon': khach_hang_duoc_chon,
        'selected_khach_hang_id': selected_khach_hang_id,
        'tu_ngay': tu_ngay,
        'den_ngay': den_ngay,
        'so_chi_tiet_cong_no': so_chi_tiet_cong_no,
        'so_du_dau_ky': so_du_dau_ky,
        'tong_ps_no': tong_ps_no,
        'tong_ps_co': tong_ps_co,
        'so_du_cuoi_ky': so_du_cuoi_ky,
    })


def _view_not_ready(feature_name):
    return HttpResponse(f'Chức năng {feature_name} đang được khôi phục.', status=501)


def khach_hang_api_lookup(request):
    query = (request.GET.get('q') or request.GET.get('term') or '').strip()
    role = (request.GET.get('role') or '').strip().lower()
    items = KhachHang.objects.filter(trang_thai=True)
    if role == 'khach_hang':
        items = items.filter(la_khach_hang=True)
    elif role == 'nha_cung_cap':
        items = items.filter(la_nha_cung_cap=True)
    elif role == 'nhan_vien':
        items = items.filter(la_nhan_vien=True)
    elif role == 'kh_nv':
        items = items.filter(Q(la_khach_hang=True) | Q(la_nhan_vien=True))
    if query:
        items = items.filter(
            Q(ma_kh__icontains=query)
            | Q(ten_kh__icontains=query)
            | Q(so_dien_thoai__icontains=query)
            | Q(dia_chi__icontains=query)
        )
    data = [
        {
            'id': kh.pk,
            'text': f'{kh.ma_kh} - {kh.ten_kh}',
            'ma_kh': kh.ma_kh,
            'ten_kh': kh.ten_kh,
            'so_dien_thoai': kh.so_dien_thoai,
            'dia_chi': kh.dia_chi,
            'la_nhan_vien': kh.la_nhan_vien,
        }
        for kh in items[:20]
    ]
    return JsonResponse({'results': data})


def khach_hang_lookup(request):
    return khach_hang_api_lookup(request)


def don_ban_api_lookup(request):
    q = request.GET.get('q', '').strip()
    khach_hang_id = request.GET.get('khach_hang_id')
    hoa_don_id = request.GET.get('hoa_don_id')

    # Don ke thua hoa don: cho phep chon cac don con hieu luc, khong chi 1 trang thai co dinh.
    items = DonBan.objects.select_related('khach_hang').exclude(trang_thai='4')
    if q:
        items = items.filter(
            Q(so_don__icontains=q)
            | Q(ten_kh__icontains=q)
            | Q(khach_hang__ma_kh__icontains=q)
            | Q(khach_hang__ten_kh__icontains=q)
        )
    if khach_hang_id:
        items = items.filter(khach_hang_id=khach_hang_id)
    if hoa_don_id:
        items = items.exclude(hoa_don_lien_ket__id=hoa_don_id)

    results = []
    for item in items[:20]:
        results.append({
            'id': item.pk,
            'so_don': item.so_don,
            'ngay_chung_tu': item.ngay_chung_tu.strftime('%d/%m/%Y') if item.ngay_chung_tu else '',
            'ma_kh': item.khach_hang.ma_kh if item.khach_hang else '',
            'ten_kh': item.ten_kh,
            'tong_thanh_toan': float(item.tong_thanh_toan or 0),
        })
    return JsonResponse({'results': results})


def hoa_don_ban_api_lookup(request):
    q = request.GET.get('q', '').strip()
    khach_hang_id = request.GET.get('khach_hang_id')

    items = HoaDonBan.objects.select_related('khach_hang').exclude(trang_thai='4')
    if q:
        items = items.filter(
            Q(so_hoa_don__icontains=q)
            | Q(ten_kh__icontains=q)
            | Q(khach_hang__ma_kh__icontains=q)
            | Q(khach_hang__ten_kh__icontains=q)
        )
    if khach_hang_id:
        items = items.filter(khach_hang_id=khach_hang_id)

    results = []
    for item in items[:20]:
        results.append({
            'id': item.pk,
            'so_hoa_don': item.so_hoa_don,
            'ngay_lap': item.ngay_lap.strftime('%d/%m/%Y') if item.ngay_lap else '',
            'khach_hang_id': item.khach_hang_id,
            'ma_kh': item.khach_hang.ma_kh if item.khach_hang else '',
            'ten_kh': item.ten_kh,
            'tong_cong': float(item.tong_cong or 0),
            'con_no': float(item.con_no or 0),
        })
    return JsonResponse({'results': results})


def don_ban_api_detail(request, pk):
    don = get_object_or_404(DonBan.objects.select_related('khach_hang', 'kho'), pk=pk)
    rows = []
    for ct in don.chi_tiet.select_related('hang_hoa'):
        rows.append({
            'hang_hoa_id': ct.hang_hoa_id,
            'ma_hang': ct.hang_hoa.ma_hang,
            'ten_hang': ct.hang_hoa.ten_hang,
            'dvt': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa.don_vi_tinh else '',
            'kho_id': don.kho_id,
            'so_luong': int(Decimal(ct.so_luong or 0)),
            'gia_ban': float(ct.don_gia or 0),
            'ty_le_ck': float(ct.chiet_khau or 0),
            'thue_suat': float(ct.thue_vat or 10),
            'tk_no': '632',
            'tk_co': '156',
        })
    return JsonResponse({
        'id': don.pk,
        'so_don': don.so_don,
        'khach_hang_id': don.khach_hang_id,
        'ma_kh': don.khach_hang.ma_kh if don.khach_hang else '',
        'ten_kh': don.ten_kh or (don.khach_hang.ten_kh if don.khach_hang else ''),
        'dia_chi': don.dia_chi_kh or (don.khach_hang.dia_chi if don.khach_hang else ''),
        'so_dien_thoai': don.sdt_kh or (don.khach_hang.so_dien_thoai if don.khach_hang else ''),
        'mst': don.mst_kh or (don.khach_hang.ma_so_thue if don.khach_hang else ''),
        'nguoi_mua_hang': don.nguoi_mua_hang or '',
        'ma_nv_ban_hang': don.ma_nv_ban_hang or '',
        'ma_ngoai_te': don.ma_ngoai_te or 'VND',
        'ty_gia': float(don.ty_gia or 1),
        'dien_giai': don.ghi_chu or '',
        'kho_id': don.kho_id,
        'rows': rows,
    })


@login_required
def hoa_don_ban_api_detail(request, pk):
    """API: lấy chi tiết hóa đơn để kế thừa vào phiếu xuất kho manual."""
    hd = get_object_or_404(HoaDonBan.objects.select_related('khach_hang'), pk=pk)
    rows = []
    for ct in hd.chi_tiet.select_related('hang_hoa', 'kho'):
        rows.append({
            'hang_hoa_id': ct.hang_hoa_id,
            'ma_hang': ct.hang_hoa.ma_hang,
            'ten_hang': ct.hang_hoa.ten_hang,
            'dvt': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa.don_vi_tinh else '',
            'kho_id': ct.kho_id,
            'ma_kho': ct.kho.ma_kho,
            'so_luong': float(ct.so_luong or 0),
            'tk_no': ct.tk_gia_von or '632',
            'tk_co': ct.tk_kho or '156',
        })
    return JsonResponse({
        'ok': True,
        'id': hd.pk,
        'so_hoa_don': hd.so_hoa_don,
        'khach_hang_id': hd.khach_hang_id,
        'ten_kh': hd.ten_kh or (hd.khach_hang.ten_kh if hd.khach_hang else ''),
        'ghi_chu': f'Xuất hàng theo HĐ {hd.so_hoa_don} - {hd.ten_kh or (hd.khach_hang.ten_kh if hd.khach_hang else "")}',
        'rows': rows,
    })


def phieu_thu_xac_nhan(request, pk):
    phieu = get_object_or_404(PhieuThu, pk=pk)
    if request.method == 'POST':
        if str(phieu.trang_thai or '').strip() == '2':
            messages.warning(request, f'Phiếu thu {phieu.so_phieu} đã chuyển sổ cái trước đó.')
            return redirect('phieu_thu_list')

        if phieu.hoa_don:
            da_thu_moi = Decimal(phieu.hoa_don.da_thu or 0) + Decimal(phieu.tong_thu or 0)
            phieu.hoa_don.da_thu = da_thu_moi
            phieu.hoa_don.con_no = Decimal(phieu.hoa_don.tong_cong or 0) - da_thu_moi
            phieu.hoa_don.save(update_fields=['da_thu', 'con_no'])

        phieu.trang_thai = '2'
        phieu.save(update_fields=['trang_thai'])
        messages.success(request, f'Đã chuyển phiếu thu {phieu.so_phieu} sang sổ cái')
    return redirect('phieu_thu_list')


def phieu_thu_chuyen_so_cai(request, pk):
    phieu = get_object_or_404(PhieuThu, pk=pk)
    if request.method == 'POST':
        if str(phieu.trang_thai or '').strip() == '2':
            messages.info(request, f'Phiếu thu {phieu.so_phieu} đã ở trạng thái Chuyển sổ cái.')
            return redirect('phieu_thu_list')

        if phieu.hoa_don:
            da_thu_moi = Decimal(phieu.hoa_don.da_thu or 0) + Decimal(phieu.tong_thu or 0)
            phieu.hoa_don.da_thu = da_thu_moi
            phieu.hoa_don.con_no = Decimal(phieu.hoa_don.tong_cong or 0) - da_thu_moi
            phieu.hoa_don.save(update_fields=['da_thu', 'con_no'])

        phieu.trang_thai = '2'
        phieu.save(update_fields=['trang_thai'])
        messages.success(request, f'Đã chuyển phiếu thu {phieu.so_phieu} sang sổ cái')
    return redirect('phieu_thu_list')


def hoa_don_ban_list(request):
    items = _hoa_don_ban_filtered_queryset(request)[:100]
    current_period = get_current_accounting_period()
    context = {
        'items': items,
        'current_period': current_period,
        'q': request.GET.get('q', '').strip(),
        'ma_giao_dich_filter': request.GET.get('ma_giao_dich', '').strip(),
        'trang_thai_filter': request.GET.get('trang_thai', '').strip(),
        'tu_ngay': request.GET.get('tu_ngay', '').strip(),
        'den_ngay': request.GET.get('den_ngay', '').strip(),
        'so_hd_tu': request.GET.get('so_hd_tu', '').strip(),
        'so_hd_den': request.GET.get('so_hd_den', '').strip(),
        'ma_kh': request.GET.get('ma_kh', '').strip(),
        'ma_hang': request.GET.get('ma_hang', '').strip(),
        'tk_vat_tu': request.GET.get('tk_vat_tu', '').strip(),
        'tk_gia_von': request.GET.get('tk_gia_von', '').strip(),
        'tk_doanh_thu': request.GET.get('tk_doanh_thu', '').strip(),
        'page_title': 'Hóa đơn bán hàng',
        'active_menu': 'hoa_don_ban',
    }
    return render(request, 'ban_hang/hoa_don_ban_list.html', context)


def hoa_don_ban_them(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                hoa_don = _save_hoa_don_from_request(request)
                if str(hoa_don.trang_thai or '').strip() in ('2', '3'):
                    post_to_ledger('hoa_don_ban', hoa_don.id, user=request.user)
        except (ValueError, LedgerPostingError) as exc:
            messages.error(request, str(exc))
            return redirect('hoa_don_ban_them')

        messages.success(request, f'Đã tạo hóa đơn {hoa_don.so_hoa_don}')
        return redirect('hoa_don_ban_detail', pk=hoa_don.pk)

    return render(request, 'ban_hang/hoa_don_ban_form.html', _build_hoa_don_context(request))


def hoa_don_ban_export_data(request):
    rows = []
    for hoa_don in _hoa_don_ban_filtered_queryset(request).prefetch_related('chi_tiet__hang_hoa', 'chi_tiet__kho'):
        for ct in hoa_don.chi_tiet.all():
            rows.append({
                'so_hoa_don': hoa_don.so_hoa_don,
                'ngay_lap': hoa_don.ngay_lap,
                'ngay_hach_toan': hoa_don.ngay_hach_toan,
                'ma_giao_dich': hoa_don.get_ma_giao_dich_display(),
                'ma_kh': hoa_don.khach_hang.ma_kh if hoa_don.khach_hang else '',
                'ten_kh': hoa_don.ten_kh,
                'dia_chi': hoa_don.dia_chi,
                'so_dien_thoai': hoa_don.so_dien_thoai,
                'mst': hoa_don.mst,
                'nguoi_mua_hang': hoa_don.nguoi_mua_hang,
                'ma_hang': ct.hang_hoa.ma_hang,
                'ten_hang': ct.hang_hoa.ten_hang,
                'ma_kho': ct.kho.ma_kho,
                'ten_kho': ct.kho.ten_kho,
                'so_luong': int(Decimal(ct.so_luong or 0)),
                'gia_ban': ct.gia_ban,
                'ty_le_chiet_khau': ct.ty_le_chiet_khau,
                'thue_suat': ct.thue_suat,
                'tk_no': hoa_don.tk_no,
                'tk_co': hoa_don.tk_co,
                'tk_vat_tu': ct.tk_vat_tu,
                'tk_gia_von': ct.tk_gia_von,
                'tk_doanh_thu': ct.tk_doanh_thu,
                'dien_giai': hoa_don.dien_giai,
                'trang_thai': hoa_don.get_trang_thai_display(),
            })

    wb = _export_hoa_don_workbook('DANH SÁCH HÓA ĐƠN BÁN HÀNG', rows)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=hoa_don_ban.xlsx'
    wb.save(response)
    return response


def hoa_don_ban_export_template(request):
    wb = _export_hoa_don_workbook('MẪU NHẬP HÓA ĐƠN BÁN HÀNG', [])
    ws = wb.active
    ws['A7'] = 'Nhập dữ liệu từ dòng 7 trở đi, giữ nguyên các cột đã có.'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=mau_hoa_don_ban.xlsx'
    wb.save(response)
    return response


def hoa_don_ban_import_excel(request):
    if request.method != 'POST':
        return redirect('hoa_don_ban_list')

    uploaded = request.FILES.get('excel_file')
    if not uploaded:
        messages.error(request, 'Vui lòng chọn tệp Excel.')
        return redirect('hoa_don_ban_list')

    try:
        wb = load_workbook(uploaded)
    except Exception:
        messages.error(request, 'Không đọc được tệp Excel.')
        return redirect('hoa_don_ban_list')

    ws = wb.active
    header_row = _find_header_row(ws, [header for header, _ in HOA_DON_EXCEL_HEADERS])
    if not header_row:
        messages.error(request, 'Không tìm thấy dòng tiêu đề hợp lệ trong tệp Excel.')
        return redirect('hoa_don_ban_list')

    header_map = {}
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(row=header_row, column=col).value or '').strip()
        if value:
            header_map[value] = col

    required = ['Số hóa đơn', 'Mã khách', 'Mã hàng', 'Mã kho', 'Số lượng', 'Giá bán']
    missing = [label for label in required if label not in header_map]
    if missing:
        messages.error(request, f'Thiếu cột bắt buộc: {", ".join(missing)}')
        return redirect('hoa_don_ban_list')

    grouped_rows = {}
    for row in range(header_row + 1, ws.max_row + 1):
        so_hoa_don = str(ws.cell(row=row, column=header_map['Số hóa đơn']).value or '').strip()
        if not so_hoa_don:
            continue
        grouped_rows.setdefault(so_hoa_don, []).append(row)

    created = updated = skipped = 0
    with transaction.atomic():
        for so_hoa_don, row_numbers in grouped_rows.items():
            first_row = row_numbers[0]
            kh_code = str(ws.cell(row=first_row, column=header_map['Mã khách']).value or '').strip()
            kh = KhachHang.objects.filter(ma_kh=kh_code).first()
            if not kh:
                skipped += 1
                continue

            hoa_don = HoaDonBan.objects.filter(so_hoa_don=so_hoa_don).select_related('khach_hang').first()
            if hoa_don:
                updated += 1
                if str(hoa_don.trang_thai or '').strip() in ('2', '3'):
                    _restore_ton_kho_from_hoa_don(hoa_don)
                hoa_don.chi_tiet.all().delete()
            else:
                hoa_don = HoaDonBan(so_hoa_don=so_hoa_don)
                created += 1

            hoa_don.ma_giao_dich = str(ws.cell(row=first_row, column=header_map.get('Mã giao dịch', 0)).value or '1')[:5] if header_map.get('Mã giao dịch') else '1'
            hoa_don.ngay_lap = _parse_date(ws.cell(row=first_row, column=header_map.get('Ngày lập', 1)).value)
            hoa_don.ngay_hach_toan = _parse_date(ws.cell(row=first_row, column=header_map.get('Ngày hạch toán', 2)).value)
            hoa_don.ma_ngoai_te = 'VND'
            hoa_don.ty_gia = Decimal('1')
            hoa_don.khach_hang = kh
            hoa_don.ten_kh = str(ws.cell(row=first_row, column=header_map.get('Tên khách', 0)).value or kh.ten_kh)
            hoa_don.dia_chi = str(ws.cell(row=first_row, column=header_map.get('Địa chỉ', 0)).value or kh.dia_chi)
            hoa_don.so_dien_thoai = str(ws.cell(row=first_row, column=header_map.get('SĐT', 0)).value or kh.so_dien_thoai)
            hoa_don.mst = str(ws.cell(row=first_row, column=header_map.get('MST', 0)).value or kh.ma_so_thue)
            hoa_don.nguoi_mua_hang = str(ws.cell(row=first_row, column=header_map.get('Người mua hàng', 0)).value or '')
            hoa_don.tk_no = str(ws.cell(row=first_row, column=header_map.get('TK nợ', 0)).value or '131').strip() if header_map.get('TK nợ') else '131'
            hoa_don.tk_co = str(ws.cell(row=first_row, column=header_map.get('TK có đối ứng', 0)).value or '').strip() if header_map.get('TK có đối ứng') else '511'
            if hoa_don.tk_no == '131' and not hoa_don.tk_co:
                hoa_don.tk_co = '511'
            if not hoa_don.tk_co:
                hoa_don.tk_co = '511'
            hoa_don.dien_giai = str(ws.cell(row=first_row, column=header_map.get('Diễn giải', 0)).value or '')
            hoa_don.trang_thai = '1'
            hoa_don.save()

            for row in row_numbers:
                hang_code = str(ws.cell(row=row, column=header_map['Mã hàng']).value or '').strip()
                kho_code = str(ws.cell(row=row, column=header_map['Mã kho']).value or '').strip()
                hang = HangHoa.objects.filter(ma_hang=hang_code).first()
                kho = Kho.objects.filter(ma_kho=kho_code).first()
                if not hang or not kho:
                    continue
                HoaDonBan_CT.objects.create(
                    hoa_don=hoa_don,
                    hang_hoa=hang,
                    kho=kho,
                    so_luong=_parse_decimal(ws.cell(row=row, column=header_map['Số lượng']).value),
                    gia_ban=_parse_decimal(ws.cell(row=row, column=header_map['Giá bán']).value),
                    ty_le_chiet_khau=_parse_decimal(ws.cell(row=row, column=header_map.get('CK %', 0)).value if header_map.get('CK %') else 0),
                    thue_suat=_parse_decimal(ws.cell(row=row, column=header_map.get('Thuế %', 0)).value if header_map.get('Thuế %') else 10),
                    tk_vat_tu=str(ws.cell(row=row, column=header_map.get('TK vật tư', 0)).value or ''),
                    tk_gia_von=str(ws.cell(row=row, column=header_map.get('TK giá vốn', 0)).value or ''),
                    tk_doanh_thu=str(ws.cell(row=row, column=header_map.get('TK doanh thu', 0)).value or ''),
                )
            hoa_don.tinh_tong()
            _sync_cong_no_from_hoa_don(hoa_don)

    messages.success(request, f'Đã nhập hóa đơn: tạo mới {created}, cập nhật {updated}, bỏ qua {skipped}.')
    return redirect('hoa_don_ban_list')


def hoa_don_ban_detail(request, pk):
    hoa_don = get_object_or_404(HoaDonBan.objects.select_related('khach_hang', 'don_ban'), pk=pk)
    chi_tiet = hoa_don.chi_tiet.select_related('hang_hoa', 'kho')
    current_period = get_current_accounting_period()
    hoa_don_trong_ky = bool(
        current_period
        and hoa_don.ngay_lap
        and current_period.tu_ngay <= hoa_don.ngay_lap <= current_period.den_ngay
    )
    return render(request, 'ban_hang/hoa_don_ban_detail.html', {
        'page_title': f'Hóa đơn {hoa_don.so_hoa_don}',
        'active_menu': 'hoa_don_ban',
        'hoa_don': hoa_don,
        'chi_tiet': chi_tiet,
        'current_period': current_period,
        'hoa_don_trong_ky': hoa_don_trong_ky,
    })

def hoa_don_ban_copy(request, pk):
    source = get_object_or_404(HoaDonBan.objects.select_related('khach_hang', 'don_ban'), pk=pk)
    with transaction.atomic():
        copy_hd = HoaDonBan.objects.create(
            ma_giao_dich=source.ma_giao_dich,
            so_hoa_don=_gen_so_hoa_don(),
            ngay_lap=date.today(),
            ngay_hach_toan=date.today(),
            ma_ngoai_te=source.ma_ngoai_te,
            ty_gia=source.ty_gia,
            khach_hang=source.khach_hang,
            ten_kh=source.ten_kh,
            dia_chi=source.dia_chi,
            so_dien_thoai=source.so_dien_thoai,
            mst=source.mst,
            nguoi_mua_hang=source.nguoi_mua_hang,
            tk_no=source.tk_no,
            tk_co=source.tk_co,
            dien_giai=source.dien_giai,
            trang_thai='1',
            don_ban=source.don_ban,
            nguoi_tao=request.user if request.user.is_authenticated else None,
        )
        for ct in source.chi_tiet.all():
            HoaDonBan_CT.objects.create(
                hoa_don=copy_hd,
                hang_hoa=ct.hang_hoa,
                kho=ct.kho,
                so_luong=ct.so_luong,
                gia_ban=ct.gia_ban,
                ty_le_chiet_khau=ct.ty_le_chiet_khau,
                thue_suat=ct.thue_suat,
                tk_vat_tu=ct.tk_vat_tu,
                tk_gia_von=ct.tk_gia_von,
                tk_doanh_thu=ct.tk_doanh_thu,
            )
        copy_hd.tinh_tong()
        _sync_cong_no_from_hoa_don(copy_hd)
    messages.success(request, f'Đã sao chép hóa đơn thành {copy_hd.so_hoa_don}')
    return redirect('hoa_don_ban_detail', pk=copy_hd.pk)


def hoa_don_ban_sua(request, pk):
    hoa_don = get_object_or_404(HoaDonBan, pk=pk)
    if request.method == 'POST':
        try:
            ensure_accounting_period_open_for_dates([hoa_don.ngay_lap, hoa_don.ngay_hach_toan], 'hóa đơn bán hàng')
            with transaction.atomic():
                _save_hoa_don_from_request(request, hoa_don)
                if str(hoa_don.trang_thai or '').strip() in ('2', '3'):
                    post_to_ledger('hoa_don_ban', hoa_don.id, user=request.user)
        except (ValueError, InvalidOperation, LedgerPostingError) as exc:
            messages.error(request, str(exc))
            return redirect('hoa_don_ban_sua', pk=pk)
        messages.success(request, f'Đã cập nhật hóa đơn {hoa_don.so_hoa_don}')
        return redirect('hoa_don_ban_detail', pk=hoa_don.pk)

    return render(request, 'ban_hang/hoa_don_ban_form.html', {
        **_build_hoa_don_context(request, hoa_don),
        'chi_tiet': hoa_don.chi_tiet.select_related('hang_hoa', 'kho'),
    })


def hoa_don_ban_xoa(request, pk):
    hoa_don = get_object_or_404(HoaDonBan, pk=pk)
    if request.method == 'POST':
        don_lien_ket = hoa_don.don_ban
        try:
            ensure_accounting_period_open_for_dates([hoa_don.ngay_lap, hoa_don.ngay_hach_toan], 'hóa đơn bán hàng')
            with transaction.atomic():
                if str(hoa_don.trang_thai or '').strip() in ('2', '3'):
                    _restore_ton_kho_from_hoa_don(hoa_don)
                hoa_don.delete()
                _recompute_don_ban_from_linked_hoa_don(don_lien_ket)
            messages.success(request, f'Đã xóa hóa đơn {hoa_don.so_hoa_don}')
        except ProtectedError:
            messages.error(
                request,
                f'Không thể xóa hóa đơn {hoa_don.so_hoa_don} vì đang được tham chiếu bởi chứng từ khác (ví dụ: phiếu trả hàng).',
            )
    return redirect('hoa_don_ban_list')


@login_required
def hoa_don_ban_xoa_nhieu(request):
    if request.method != 'POST':
        return redirect('hoa_don_ban_list')

    ids = request.POST.getlist('ids[]') or request.POST.getlist('ids')
    if not ids:
        messages.error(request, 'Vui lòng chọn ít nhất 1 hóa đơn để xóa.')
        return redirect('hoa_don_ban_list')

    items = list(HoaDonBan.objects.filter(pk__in=ids).select_related('don_ban'))
    if not items:
        messages.error(request, 'Không tìm thấy hóa đơn cần xóa.')
        return redirect('hoa_don_ban_list')

    deleted = 0
    blocked_codes = []

    for hoa_don in items:
        don_lien_ket = hoa_don.don_ban
        try:
            with transaction.atomic():
                if str(hoa_don.trang_thai or '').strip() in ('2', '3'):
                    _restore_ton_kho_from_hoa_don(hoa_don)
                hoa_don.delete()
                _recompute_don_ban_from_linked_hoa_don(don_lien_ket)
            deleted += 1
        except ProtectedError:
            blocked_codes.append(hoa_don.so_hoa_don)

    if deleted:
        messages.success(request, f'Đã xóa {deleted} hóa đơn.')
    if blocked_codes:
        preview = ', '.join(blocked_codes[:5])
        suffix = '...' if len(blocked_codes) > 5 else ''
        messages.error(
            request,
            f'Không thể xóa {len(blocked_codes)} hóa đơn do đang được tham chiếu bởi chứng từ khác: {preview}{suffix}',
        )

    if not deleted and not blocked_codes:
        messages.error(request, 'Không có hóa đơn nào được xóa.')
    return redirect('hoa_don_ban_list')


def hoa_don_ban_chuyen_so_cai(request, pk):
    hoa_don = get_object_or_404(HoaDonBan, pk=pk)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                if str(hoa_don.trang_thai or '').strip() in ('2', '3'):
                    post_to_ledger('hoa_don_ban', hoa_don.id, user=request.user)
                    messages.info(request, f'Hóa đơn {hoa_don.so_hoa_don} đã ở trạng thái chuyển sổ cái và đã được ghi sổ.')
                    return redirect('hoa_don_ban_detail', pk=pk)
                hoa_don.trang_thai = '2'
                hoa_don.save(update_fields=['trang_thai'])
                _apply_ton_kho_for_hoa_don(hoa_don)
                _sync_cong_no_from_hoa_don(hoa_don)
                post_to_ledger('hoa_don_ban', hoa_don.id, user=request.user)
            messages.success(request, f'Đã chuyển sổ cái hóa đơn {hoa_don.so_hoa_don}')
        except LedgerPostingError as exc:
            messages.error(request, str(exc))
        except ValueError as exc:
            messages.error(request, str(exc))
    return redirect('hoa_don_ban_detail', pk=pk)


def hoa_don_ban_in(request, pk):
    hoa_don = get_object_or_404(HoaDonBan.objects.select_related('khach_hang', 'don_ban'), pk=pk)
    chi_tiet = hoa_don.chi_tiet.select_related('hang_hoa', 'kho')
    if request.GET.get('format') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'HoaDonBan'
        _style_export_sheet(ws, f'HÓA ĐƠN {hoa_don.so_hoa_don}', [
            ('STT', 'stt'),
            ('Mã hàng', 'ma_hang'),
            ('Tên hàng', 'ten_hang'),
            ('Kho', 'kho'),
            ('Số lượng', 'so_luong'),
            ('Giá bán', 'gia_ban'),
            ('CK %', 'ck'),
            ('Thuế %', 'thue'),
            ('Thành tiền', 'thanh_tien'),
        ])
        for index, ct in enumerate(chi_tiet, start=1):
            ws.cell(row=6 + index, column=1, value=index)
            ws.cell(row=6 + index, column=2, value=ct.hang_hoa.ma_hang)
            ws.cell(row=6 + index, column=3, value=ct.hang_hoa.ten_hang)
            ws.cell(row=6 + index, column=4, value=ct.kho.ma_kho)
            ws.cell(row=6 + index, column=5, value=float(ct.so_luong or 0))
            ws.cell(row=6 + index, column=6, value=float(ct.gia_ban or 0))
            ws.cell(row=6 + index, column=7, value=float(ct.ty_le_chiet_khau or 0))
            ws.cell(row=6 + index, column=8, value=float(ct.thue_suat or 0))
            ws.cell(row=6 + index, column=9, value=float(ct.thanh_tien or 0))
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={hoa_don.so_hoa_don}.xlsx'
        wb.save(response)
        return response

    return render(request, 'ban_hang/hoa_don_ban_print.html', {
        'hoa_don': hoa_don,
        'chi_tiet': chi_tiet,
        'page_title': f'In hóa đơn {hoa_don.so_hoa_don}',
        'active_menu': 'hoa_don_ban',
    })


def gia_ban_list(request):
    q = request.GET.get('q', '').strip()
    trang_thai = request.GET.get('trang_thai', '').strip()

    items = PhieuGiaBan.objects.select_related('nhom_hang', 'nguoi_lap').prefetch_related('chi_tiet__hang_hoa__don_vi_tinh')
    if q:
        items = items.filter(
            Q(ma_phieu__icontains=q)
            | Q(nhom_hang__ma_nhom__icontains=q)
            | Q(nhom_hang__ten_nhom__icontains=q)
            | Q(chi_tiet__hang_hoa__ma_hang__icontains=q)
            | Q(chi_tiet__hang_hoa__ten_hang__icontains=q)
        ).distinct()
    if trang_thai:
        items = items.filter(trang_thai_duyet=trang_thai)

    rows = []
    for phieu in items.order_by('-ngay_lap', '-ngay_tao'):
        for ct in phieu.chi_tiet.all():
            rows.append({
                'phieu_id': phieu.pk,
                'phieu__trang_thai_duyet': phieu.trang_thai_duyet,
                'phieu__ma_phieu': phieu.ma_phieu,
                'phieu__ngay_hieu_luc': phieu.ngay_hieu_luc,
                'phieu__loai_tien_te': phieu.loai_tien_te,
                'hang_hoa__ma_hang': ct.hang_hoa.ma_hang,
                'hang_hoa__ten_hang': ct.hang_hoa.ten_hang,
                'hang_hoa__don_vi_tinh__ten': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa.don_vi_tinh else '',
                'gia_ban_chuan': ct.gia_ban_chuan,
            })

    return render(request, 'ban_hang/gia_ban_list.html', {
        'rows': rows,
        'q': q,
        'trang_thai': trang_thai,
        'page_title': 'Quản lý giá bán',
        'active_menu': 'gia_ban',
    })


def gia_ban_them(request):
    return _gia_ban_form(request)


@login_required
def gia_ban_detail(request, pk):
    phieu = get_object_or_404(
        PhieuGiaBan.objects.select_related('nhom_hang', 'nguoi_lap').prefetch_related('chi_tiet__hang_hoa__don_vi_tinh', 'bang_chiet_khau'),
        pk=pk,
    )
    return render(request, 'ban_hang/gia_ban_detail.html', {
        'phieu': phieu,
        'chi_tiet': phieu.chi_tiet.all(),
        'bang_ck': phieu.bang_chiet_khau.all(),
        'page_title': f'Chi tiết phiếu giá bán {phieu.ma_phieu}',
        'active_menu': 'gia_ban',
    })


def _gen_ma_phieu_gia_ban():
    last = PhieuGiaBan.objects.order_by('-id').first()
    if not last:
        return 'GB1'
    digits = ''.join(ch for ch in last.ma_phieu if ch.isdigit())
    next_num = int(digits or 0) + 1
    return f'GB{next_num}'


def _gia_ban_validate_and_save(request, phieu=None):
    data = request.POST
    nhom_hang_id = data.get('nhom_hang')
    nhom_hang = NhomHang.objects.filter(pk=nhom_hang_id).first() if nhom_hang_id else None

    ma_phieu = data.get('ma_phieu') or _gen_ma_phieu_gia_ban()
    ngay_lap = _parse_date(data.get('ngay_lap'))
    ngay_hieu_luc = _parse_date(data.get('ngay_hieu_luc'))
    default_bien_do = nhom_hang.bien_do_loi_nhuan if nhom_hang else Decimal('10')
    bien_do_loi_nhuan = _parse_decimal(data.get('bien_do_loi_nhuan'), default_bien_do)
    loai_tien_te = data.get('loai_tien_te', 'VND')
    ghi_chu = data.get('ghi_chu', '')

    if phieu is None and PhieuGiaBan.objects.filter(ma_phieu=ma_phieu).exists():
        raise ValueError('Mã phiếu giá bán đã tồn tại')

    hang_ids = data.getlist('hang_id[]')
    gia_vons = data.getlist('gia_von[]')
    gia_ban_chuans = data.getlist('gia_ban_chuan[]')
    ck_tu_list = data.getlist('ck_tu[]')
    ck_den_list = data.getlist('ck_den[]')
    ck_pt_list = data.getlist('ck_pt[]')

    detail_rows = []
    seen_hang = set()
    for idx, hang_id in enumerate(hang_ids):
        if not hang_id:
            continue
        hang = HangHoa.objects.select_related('don_vi_tinh', 'nhom_hang').filter(pk=hang_id).first()
        if not hang:
            raise ValueError('Vui lòng kiểm tra lại thông tin các trường dữ liệu bắt buộc')

        gia_von = _parse_decimal(gia_vons[idx] if idx < len(gia_vons) else 0)
        gia_ban_chuan = _parse_decimal(gia_ban_chuans[idx] if idx < len(gia_ban_chuans) else 0)

        if gia_von < 0 or gia_ban_chuan < 0:
            raise ValueError('Vui lòng kiểm tra lại thông tin các trường dữ liệu bắt buộc')
        if gia_ban_chuan < gia_von:
            raise ValueError('Giá bán mới không được nhỏ hơn giá vốn')
        if hang.pk in seen_hang:
            raise ValueError('Vui lòng kiểm tra lại thông tin các trường dữ liệu bắt buộc')
        seen_hang.add(hang.pk)

        detail_rows.append({
            'hang': hang,
            'gia_von': gia_von,
            'gia_ban_chuan': gia_ban_chuan,
        })

    if not detail_rows:
        raise ValueError('Vui lòng kiểm tra lại thông tin các trường dữ liệu bắt buộc')

    # Nhóm hàng không bắt buộc nhập tay: tự suy ra từ mặt hàng đầu tiên.
    if not nhom_hang:
        nhom_hang = detail_rows[0]['hang'].nhom_hang
    if not nhom_hang and phieu and phieu.nhom_hang_id:
        nhom_hang = phieu.nhom_hang
    if not nhom_hang:
        raise ValueError('Không xác định được nhóm hàng cho phiếu giá bán')

    discount_rows = []
    ranges = []
    for idx in range(max(len(ck_tu_list), len(ck_den_list), len(ck_pt_list))):
        tu = ck_tu_list[idx] if idx < len(ck_tu_list) else ''
        den = ck_den_list[idx] if idx < len(ck_den_list) else ''
        pt = ck_pt_list[idx] if idx < len(ck_pt_list) else ''
        if not tu and not den and not pt:
            continue
        tu_int = int(_parse_decimal(tu, 0))
        den_int = int(_parse_decimal(den, 0))
        pt_dec = _parse_decimal(pt, Decimal('0'))
        if tu_int <= 0 or den_int <= 0 or tu_int > den_int or pt_dec < 0:
            raise ValueError('Khoảng số lượng chiết khấu không hợp lệ')
        for used_tu, used_den in ranges:
            if not (den_int < used_tu or used_den < tu_int):
                raise ValueError('Khoảng số lượng chiết khấu không hợp lệ')
        ranges.append((tu_int, den_int))
        discount_rows.append({'tu': tu_int, 'den': den_int, 'pt': pt_dec})

    with transaction.atomic():
        phieu = phieu or PhieuGiaBan()
        phieu.ma_phieu = ma_phieu
        phieu.ngay_lap = ngay_lap
        phieu.ngay_hieu_luc = ngay_hieu_luc
        phieu.nguoi_lap = request.user
        phieu.nhom_hang = nhom_hang
        phieu.bien_do_loi_nhuan = bien_do_loi_nhuan
        phieu.loai_tien_te = loai_tien_te
        phieu.trang_thai_duyet = data.get('trang_thai_duyet', phieu.trang_thai_duyet or '1')
        if phieu.trang_thai_duyet not in ('0', '1'):
            phieu.trang_thai_duyet = '1'
        phieu.ghi_chu = ghi_chu
        phieu.save()

        phieu.chi_tiet.all().delete()
        phieu.bang_chiet_khau.all().delete()

        for row in detail_rows:
            PhieuGiaBan_CT.objects.create(
                phieu=phieu,
                hang_hoa=row['hang'],
                gia_von=row['gia_von'],
                gia_ban_chuan=row['gia_ban_chuan'],
            )
        for row in discount_rows:
            PhieuGiaBanChietKhau.objects.create(
                phieu=phieu,
                tu_so_luong=row['tu'],
                den_so_luong=row['den'],
                phan_tram_chiet_khau=row['pt'],
            )

    return phieu


def _gia_ban_form(request, phieu=None):
    if request.method == 'POST':
        try:
            phieu = _gia_ban_validate_and_save(request, phieu)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(request.path)

        messages.success(request, 'Thêm giá bán thành công' if request.path.endswith('/them/') else 'Cập nhật phiếu giá bán thành công')
        return redirect('gia_ban_list')

    context = {
        'phieu': phieu,
        'editing': bool(phieu),
        'nhom_list': NhomHang.objects.all().order_by('ten_nhom'),
        'hang_list': HangHoa.objects.select_related('don_vi_tinh', 'nhom_hang').order_by('ma_hang'),
        'ma_phieu_default': phieu.ma_phieu if phieu else _gen_ma_phieu_gia_ban(),
        'today': date.today(),
        'page_title': 'Sửa phiếu giá bán' if phieu else 'Thêm phiếu giá bán',
        'active_menu': 'gia_ban',
        'chi_tiet': phieu.chi_tiet.select_related('hang_hoa__don_vi_tinh') if phieu else [],
        'bang_ck': phieu.bang_chiet_khau.all() if phieu else [],
    }
    return render(request, 'ban_hang/gia_ban_form.html', context)


def gia_ban_export_template(request):
    return _view_not_ready('xuất mẫu bảng giá bán')


def gia_ban_import_excel(request):
    return _view_not_ready('nhập bảng giá bán từ Excel')


def gia_ban_copy(request, pk):
    source = get_object_or_404(PhieuGiaBan.objects.select_related('nhom_hang'), pk=pk)
    clone = PhieuGiaBan.objects.create(
        ma_phieu=_gen_ma_phieu_gia_ban(),
        ngay_lap=date.today(),
        ngay_hieu_luc=source.ngay_hieu_luc,
        nguoi_lap=request.user,
        nhom_hang=source.nhom_hang,
        bien_do_loi_nhuan=source.bien_do_loi_nhuan,
        loai_tien_te=source.loai_tien_te,
        trang_thai_duyet=source.trang_thai_duyet if source.trang_thai_duyet in ('0', '1') else '1',
        ghi_chu=source.ghi_chu,
    )
    for ct in source.chi_tiet.all():
        PhieuGiaBan_CT.objects.create(
            phieu=clone,
            hang_hoa=ct.hang_hoa,
            gia_von=ct.gia_von,
            gia_ban_chuan=ct.gia_ban_chuan,
        )
    for ck in source.bang_chiet_khau.all():
        PhieuGiaBanChietKhau.objects.create(
            phieu=clone,
            tu_so_luong=ck.tu_so_luong,
            den_so_luong=ck.den_so_luong,
            phan_tram_chiet_khau=ck.phan_tram_chiet_khau,
        )
    messages.success(request, f'Đã sao chép phiếu giá bán thành {clone.ma_phieu}')
    return redirect('gia_ban_sua', pk=clone.pk)


def gia_ban_sua(request, pk):
    phieu = get_object_or_404(PhieuGiaBan, pk=pk)
    return _gia_ban_form(request, phieu)


def gia_ban_xoa(request, pk):
    phieu = get_object_or_404(PhieuGiaBan, pk=pk)
    if request.method == 'POST':
        phieu.delete()
        messages.success(request, 'Đã xóa phiếu giá bán')


def _get_active_phieu_gia_for_hang(hang):
    # Tương thích dữ liệu cũ: trước đây trạng thái lưu bằng chuỗi duyệt.
    active_statuses = ['1', 'da_duyet', 'cho_duyet']
    return (
        PhieuGiaBan.objects
        .filter(trang_thai_duyet__in=active_statuses, ngay_hieu_luc__lte=date.today(), chi_tiet__hang_hoa=hang)
        .select_related('nhom_hang')
        .prefetch_related('chi_tiet', 'bang_chiet_khau')
        .order_by('-ngay_hieu_luc', '-ngay_cap_nhat', '-id')
        .first()
    )


# ─── PHIẾU GIAO HÀNG ────────────────────────────────────────────────────────

def _gen_so_phieu_giao_hang():
    return _next_code_from_queryset_values(
        PhieuGiaoHang.objects.filter(so_phieu__istartswith='PGH').values_list('so_phieu', flat=True),
        'PGH',
    )


@login_required
def phieu_giao_hang_list(request):
    q = request.GET.get('q', '').strip()
    trang_thai = request.GET.get('trang_thai', '').strip()
    tu_ngay = request.GET.get('tu_ngay', '').strip()
    den_ngay = request.GET.get('den_ngay', '').strip()

    items = PhieuGiaoHang.objects.select_related('khach_hang', 'hoa_don_goc')
    if q:
        items = items.filter(
            Q(so_phieu__icontains=q) | Q(ten_kh__icontains=q)
            | Q(khach_hang__ma_kh__icontains=q) | Q(khach_hang__ten_kh__icontains=q)
        )
    if trang_thai:
        items = items.filter(trang_thai=trang_thai)
    if tu_ngay:
        items = items.filter(ngay_lap__gte=tu_ngay)
    if den_ngay:
        items = items.filter(ngay_lap__lte=den_ngay)

    items = items.order_by('-ngay_lap', '-ngay_tao')

    context = {
        'items': items,
        'q': q,
        'trang_thai_filter': trang_thai,
        'tu_ngay': tu_ngay,
        'den_ngay': den_ngay,
        'page_title': 'Phiếu giao hàng',
        'active_menu': 'phieu_giao_hang',
    }
    return render(request, 'ban_hang/phieu_giao_hang_list.html', context)


@login_required
def phieu_giao_hang_them(request):
    ke_thua_id = request.GET.get('ke_thua_hoa_don') or request.POST.get('ke_thua_hoa_don')

    if request.method == 'POST':
        data = request.POST
        kh_id = data.get('khach_hang') or None
        kh = KhachHang.objects.filter(pk=kh_id).first() if kh_id else None
        if not kh:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin các trường dữ liệu bắt buộc')
            return redirect('phieu_giao_hang_them')

        hoa_don_goc_id = data.get('hoa_don_goc') or None
        hoa_don_goc = HoaDonBan.objects.filter(pk=hoa_don_goc_id).first() if hoa_don_goc_id else None

        so_phieu_input = (data.get('so_phieu') or '').strip()
        if not so_phieu_input:
            so_phieu_input = _gen_so_phieu_giao_hang()
        if PhieuGiaoHang.objects.filter(so_phieu=so_phieu_input).exists():
            so_phieu_input = _gen_so_phieu_giao_hang()

        phieu = PhieuGiaoHang.objects.create(
            so_phieu=so_phieu_input,
            ngay_lap=_parse_date(data.get('ngay_lap')),
            khach_hang=kh,
            ten_kh=kh.ten_kh,
            nguoi_nhan=data.get('nguoi_nhan', ''),
            dien_giai=data.get('dien_giai', ''),
            trang_thai=data.get('trang_thai', 'cho_giao'),
            hoa_don_goc=hoa_don_goc,
            nguoi_tao=request.user,
        )

        hang_ids = data.getlist('hang_id[]')
        kho_ids = data.getlist('kho_id[]')
        vi_tris = data.getlist('vi_tri[]')
        so_luongs = data.getlist('so_luong[]')
        don_gias = data.getlist('don_gia[]')
        ngay_giaos = data.getlist('ngay_giao[]')
        loai_vcs = data.getlist('loai_van_chuyen[]')
        ghi_chus = data.getlist('ghi_chu[]')

        so_dong_hop_le = 0
        for i in range(len(hang_ids)):
            hang_id = hang_ids[i] if i < len(hang_ids) else ''
            kho_id = kho_ids[i] if i < len(kho_ids) else ''
            if not hang_id or not kho_id:
                continue
            sl = _normalize_decimal(so_luongs[i] if i < len(so_luongs) else 0, 2)
            if sl <= 0:
                continue
            PhieuGiaoHang_CT.objects.create(
                phieu=phieu,
                hang_hoa_id=hang_id,
                kho_id=kho_id,
                vi_tri=vi_tris[i] if i < len(vi_tris) else '',
                so_luong=sl,
                don_gia=_normalize_decimal(don_gias[i] if i < len(don_gias) else 0, 0),
                ngay_giao=_parse_date(ngay_giaos[i] if i < len(ngay_giaos) else None),
                loai_van_chuyen=loai_vcs[i] if i < len(loai_vcs) else 'tu_giao',
                ghi_chu=ghi_chus[i] if i < len(ghi_chus) else '',
            )
            so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            phieu.delete()
            messages.error(request, 'Phiếu giao hàng phải có ít nhất 1 dòng hàng hóa hợp lệ')
            return redirect('phieu_giao_hang_them')

        phieu.tinh_tong()
        messages.success(request, f'Đã tạo phiếu giao hàng {phieu.so_phieu}')
        return redirect('phieu_giao_hang_list')

    ke_thua_hoa_don = None
    ke_thua_chi_tiet = []
    if ke_thua_id:
        ke_thua_hoa_don = HoaDonBan.objects.select_related('khach_hang').filter(pk=ke_thua_id).first()
        if ke_thua_hoa_don:
            ke_thua_chi_tiet = list(ke_thua_hoa_don.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh', 'kho'))

    # Xu ly copy_from: neu co tham so copy_from, lay du lieu phieu giao hang cu de dien vao form moi
    copy_from_id = request.GET.get('copy_from', '').strip()
    phieu_copy = None
    chi_tiet_copy = []
    if copy_from_id and copy_from_id.isdigit() and not ke_thua_id:
        phieu_copy = PhieuGiaoHang.objects.select_related('khach_hang', 'hoa_don_goc').filter(pk=copy_from_id).first()
        if phieu_copy:
            chi_tiet_copy = list(phieu_copy.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh', 'kho'))

    context = {
        'editing': False,
        'phieu': phieu_copy,  # dung lam data nguon dien san form (copy_from)
        'chi_tiet': chi_tiet_copy or ke_thua_chi_tiet,
        'ke_thua_hoa_don': ke_thua_hoa_don,
        'is_copy_mode': bool(phieu_copy),  # flag de template biet dang o che do copy
        'kh_list': KhachHang.objects.filter(Q(la_khach_hang=True) | Q(la_nhan_vien=True), trang_thai=True).order_by('ma_kh'),
        'hoa_don_list': HoaDonBan.objects.select_related('khach_hang').order_by('-ngay_lap', '-id')[:200],
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'hang_list': HangHoa.objects.filter(trang_thai='dang_ban').select_related('don_vi_tinh').order_by('ma_hang'),
        'so_phieu_default': _gen_so_phieu_giao_hang(),
        'today': date.today(),
        'page_title': 'Lập phiếu giao hàng (chép từ phiếu cũ)' if phieu_copy else 'Lập phiếu giao hàng',
        'active_menu': 'phieu_giao_hang',
    }
    return render(request, 'ban_hang/phieu_giao_hang_form.html', context)


@login_required
def phieu_giao_hang_sua(request, pk):
    phieu = get_object_or_404(PhieuGiaoHang, pk=pk)
    if phieu.trang_thai == 'da_giao':
        messages.error(request, 'Phiếu giao hàng đã hoàn tất, không thể chỉnh sửa')
        return redirect('phieu_giao_hang_list')

    if request.method == 'POST':
        data = request.POST
        kh_id = data.get('khach_hang') or None
        kh = KhachHang.objects.filter(pk=kh_id).first() if kh_id else phieu.khach_hang

        hoa_don_goc_id = data.get('hoa_don_goc') or None
        hoa_don_goc = HoaDonBan.objects.filter(pk=hoa_don_goc_id).first() if hoa_don_goc_id else phieu.hoa_don_goc

        phieu.ngay_lap = _parse_date(data.get('ngay_lap'))
        phieu.khach_hang = kh
        phieu.ten_kh = kh.ten_kh if kh else phieu.ten_kh
        phieu.nguoi_nhan = data.get('nguoi_nhan', '')
        phieu.dien_giai = data.get('dien_giai', '')
        phieu.trang_thai = data.get('trang_thai', phieu.trang_thai)
        phieu.hoa_don_goc = hoa_don_goc
        phieu.save()

        phieu.chi_tiet.all().delete()
        hang_ids = data.getlist('hang_id[]')
        kho_ids = data.getlist('kho_id[]')
        vi_tris = data.getlist('vi_tri[]')
        so_luongs = data.getlist('so_luong[]')
        don_gias = data.getlist('don_gia[]')
        ngay_giaos = data.getlist('ngay_giao[]')
        loai_vcs = data.getlist('loai_van_chuyen[]')
        ghi_chus = data.getlist('ghi_chu[]')

        so_dong_hop_le = 0
        for i in range(len(hang_ids)):
            hang_id = hang_ids[i] if i < len(hang_ids) else ''
            kho_id = kho_ids[i] if i < len(kho_ids) else ''
            if not hang_id or not kho_id:
                continue
            sl = _normalize_decimal(so_luongs[i] if i < len(so_luongs) else 0, 2)
            if sl <= 0:
                continue
            PhieuGiaoHang_CT.objects.create(
                phieu=phieu,
                hang_hoa_id=hang_id,
                kho_id=kho_id,
                vi_tri=vi_tris[i] if i < len(vi_tris) else '',
                so_luong=sl,
                don_gia=_normalize_decimal(don_gias[i] if i < len(don_gias) else 0, 0),
                ngay_giao=_parse_date(ngay_giaos[i] if i < len(ngay_giaos) else None),
                loai_van_chuyen=loai_vcs[i] if i < len(loai_vcs) else 'tu_giao',
                ghi_chu=ghi_chus[i] if i < len(ghi_chus) else '',
            )
            so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin các trường dữ liệu bắt buộc')
            return redirect('phieu_giao_hang_sua', pk=pk)

        phieu.tinh_tong()
        messages.success(request, f'Đã cập nhật phiếu giao hàng {phieu.so_phieu}')
        return redirect('phieu_giao_hang_list')

    context = {
        'editing': True,
        'phieu': phieu,
        'chi_tiet': list(phieu.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh', 'kho')),
        'ke_thua_hoa_don': phieu.hoa_don_goc,
        'kh_list': KhachHang.objects.filter(trang_thai=True).order_by('ma_kh'),
        'hoa_don_list': HoaDonBan.objects.select_related('khach_hang').order_by('-ngay_lap', '-id')[:200],
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'hang_list': HangHoa.objects.filter(trang_thai='dang_ban').select_related('don_vi_tinh').order_by('ma_hang'),
        'so_phieu_default': phieu.so_phieu,
        'today': date.today(),
        'page_title': f'Sửa phiếu giao hàng {phieu.so_phieu}',
        'active_menu': 'phieu_giao_hang',
    }
    return render(request, 'ban_hang/phieu_giao_hang_form.html', context)


@login_required
def phieu_giao_hang_xoa(request, pk):
    phieu = get_object_or_404(PhieuGiaoHang, pk=pk)
    if phieu.trang_thai == 'da_giao':
        messages.error(request, 'Phiếu giao hàng đã hoàn tất, không thể xóa')
        return redirect('phieu_giao_hang_list')
    if request.method == 'POST':
        so = phieu.so_phieu
        phieu.delete()
        messages.success(request, f'Xóa phiếu giao hàng {so} thành công')
    return redirect('phieu_giao_hang_list')


@login_required
def phieu_giao_hang_api_hoa_don(request, pk):
    """API: lấy chi tiết hóa đơn để kế thừa vào phiếu giao hàng."""
    hd = get_object_or_404(HoaDonBan.objects.select_related('khach_hang'), pk=pk)
    rows = []
    for ct in hd.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh', 'kho'):
        rows.append({
            'hang_id': ct.hang_hoa_id,
            'ma_hang': ct.hang_hoa.ma_hang,
            'ten_hang': ct.hang_hoa.ten_hang,
            'dvt': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa.don_vi_tinh else '',
            'kho_id': ct.kho_id,
            'ma_kho': ct.kho.ma_kho,
            'ten_kho': ct.kho.ten_kho,
            'so_luong': float(ct.so_luong),
            'don_gia': float(ct.gia_ban),
        })
    return JsonResponse({
        'ok': True,
        'so_hoa_don': hd.so_hoa_don,
        'khach_hang_id': hd.khach_hang_id,
        'ten_kh': hd.ten_kh or (hd.khach_hang.ten_kh if hd.khach_hang else ''),
        'rows': rows,
    })


# ─── HELPER GIÁ BÁN ─────────────────────────────────────────────────────────

def _resolve_gia_ban_hang_hoa(hang):
    gia_von = hang.get_gia_von() if hasattr(hang, 'get_gia_von') else 0
    bien_do = hang.nhom_hang.bien_do_loi_nhuan if hang.nhom_hang and hasattr(hang.nhom_hang, 'bien_do_loi_nhuan') else 0

    phieu_gia = _get_active_phieu_gia_for_hang(hang)
    if phieu_gia:
        ct_gia = phieu_gia.chi_tiet.filter(hang_hoa=hang).first()
        if ct_gia and ct_gia.gia_ban_chuan is not None:
            return Decimal(ct_gia.gia_ban_chuan or 0)

    return round(Decimal(gia_von or 0) * (Decimal('1') + Decimal(bien_do or 0) / Decimal('100')), 0)


def _resolve_chiet_khau_hang_hoa(hang, so_luong):
    sl = _parse_decimal(so_luong, Decimal('0'))
    if sl <= 0:
        return Decimal('0')

    phieu_gia = _get_active_phieu_gia_for_hang(hang)
    if not phieu_gia:
        return Decimal('0')

    muc_ck = (
        phieu_gia.bang_chiet_khau
        .filter(tu_so_luong__lte=sl, den_so_luong__gte=sl)
        .order_by('-tu_so_luong', '-id')
        .first()
    )
    return Decimal(muc_ck.phan_tram_chiet_khau or 0) if muc_ck else Decimal('0')


def gia_ban_hang_hoa_api(request):
    hang_id = request.GET.get('hang_id')
    so_luong = _parse_decimal(request.GET.get('so_luong') or 1, Decimal('1'))
    hang = HangHoa.objects.select_related('don_vi_tinh', 'nhom_hang').filter(pk=hang_id).first()
    if not hang:
        return JsonResponse({'error': 'Không tìm thấy'}, status=404)

    gia_von = hang.get_gia_von() if hasattr(hang, 'get_gia_von') else 0
    bien_do = hang.nhom_hang.bien_do_loi_nhuan if hang.nhom_hang and hasattr(hang.nhom_hang, 'bien_do_loi_nhuan') else 0
    gia_ban_chuan = _resolve_gia_ban_hang_hoa(hang)
    chiet_khau = _resolve_chiet_khau_hang_hoa(hang, so_luong)

    phieu_gia = _get_active_phieu_gia_for_hang(hang)

    return JsonResponse({
        'id': hang.pk,
        'ma_hang': hang.ma_hang,
        'ten_hang': hang.ten_hang,
        'don_vi_tinh': hang.don_vi_tinh.ten if hang.don_vi_tinh else '',
        'nhom_hang_id': hang.nhom_hang_id,
        'bien_do_loi_nhuan': float(bien_do or 0),
        'gia_von': float(gia_von or 0),
        'gia_ban_chuan': float(gia_ban_chuan or 0),
        'chiet_khau_phan_tram': float(chiet_khau or 0),
        'gia_ban_nguon': 'phieu_gia_ban' if phieu_gia else 'bien_do_nhom',
    })


_period_guard_fallbacks = {
    'don_ban_them': 'don_ban_list',
    'don_ban_sua': 'don_ban_detail',
    'don_ban_xoa': 'don_ban_list',
    'don_ban_xoa_nhieu': 'don_ban_list',
    'phieu_thu_them': 'phieu_thu_list',
    'phieu_thu_sua': 'phieu_thu_list',
    'phieu_thu_xoa': 'phieu_thu_list',
    'phieu_thu_xoa_nhieu': 'phieu_thu_list',
    'phieu_doi_tra_them': 'phieu_doi_tra_list',
    'phieu_doi_tra_sua': 'phieu_doi_tra_list',
    'hoa_don_ban_them': 'hoa_don_ban_list',
    'hoa_don_ban_sua': 'hoa_don_ban_detail',
    'hoa_don_ban_xoa': 'hoa_don_ban_list',
    'hoa_don_ban_xoa_nhieu': 'hoa_don_ban_list',
}

for _view_name, _fallback in _period_guard_fallbacks.items():
    if _view_name in globals():
        globals()[_view_name] = guard_accounting_period_error(_fallback)(globals()[_view_name])

