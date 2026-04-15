from __future__ import annotations

from io import BytesIO
from typing import Any

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook

from .revenue_report_serializers import RevenueReportSerializer
from .revenue_report_service import RevenueReportFilters, RevenueReportService


def _build_service_from_request(request):
    filters = RevenueReportFilters.from_payload(request.GET)
    return RevenueReportService(filters)


def _ok(payload: dict[str, Any]) -> JsonResponse:
    return JsonResponse({'ok': True, 'data': RevenueReportSerializer.to_json_ready(payload)})


def _bad_request(message: str) -> JsonResponse:
    return JsonResponse({'ok': False, 'error': message}, status=400)


@login_required
def revenue_summary_api(request):
    try:
        service = _build_service_from_request(request)
        group_by = (request.GET.get('group_by') or 'day').strip()
        data = service.get_revenue_summary(group_by=group_by)
        return _ok(data)
    except ValueError as exc:
        return _bad_request(str(exc))


@login_required
def revenue_by_customer_api(request):
    try:
        service = _build_service_from_request(request)
        return _ok(service.get_revenue_by_customer())
    except ValueError as exc:
        return _bad_request(str(exc))


@login_required
def revenue_by_product_api(request):
    try:
        service = _build_service_from_request(request)
        return _ok(service.get_revenue_by_product())
    except ValueError as exc:
        return _bad_request(str(exc))


@login_required
def revenue_by_salesperson_api(request):
    try:
        service = _build_service_from_request(request)
        return _ok(service.get_revenue_by_salesperson())
    except ValueError as exc:
        return _bad_request(str(exc))


def _write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(key, '') for key in headers])


def _export_summary_excel(data: dict[str, Any]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Revenue Summary'
    headers = ['period', 'gross_revenue', 'revenue_deduction', 'net_revenue', 'cogs', 'gross_profit']
    _write_sheet(ws, headers, RevenueReportSerializer.to_json_ready(data['rows']))

    ws_total = wb.create_sheet('Totals')
    _write_sheet(ws_total, list(data['totals'].keys()), [RevenueReportSerializer.to_json_ready(data['totals'])])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="revenue_summary.xlsx"'
    return response


@login_required
def revenue_export_excel(request):
    report_type = (request.GET.get('report_type') or 'summary').strip()
    try:
        service = _build_service_from_request(request)
    except ValueError as exc:
        return _bad_request(str(exc))

    if report_type == 'summary':
        group_by = (request.GET.get('group_by') or 'day').strip()
        data = service.get_revenue_summary(group_by=group_by)
        return _export_summary_excel(data)

    if report_type == 'customer':
        data = service.get_revenue_by_customer()
        filename = 'revenue_by_customer.xlsx'
    elif report_type == 'product':
        data = service.get_revenue_by_product()
        filename = 'revenue_by_product.xlsx'
    elif report_type == 'salesperson':
        data = service.get_revenue_by_salesperson()
        filename = 'revenue_by_salesperson.xlsx'
    else:
        return _bad_request('report_type khong hop le')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Revenue'
    rows = RevenueReportSerializer.to_json_ready(data['rows'])
    headers = list(rows[0].keys()) if rows else []
    if headers:
        _write_sheet(ws, headers, rows)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
