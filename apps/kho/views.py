"""Views cho app Kho: PhieuXuat, TonKho, KiemKe, SoDo, TinhGia, DoiChieu"""
import re
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.clickjacking import xframe_options_exempt
from openpyxl import Workbook

from apps.ban_hang.models import PhieuTraHang
from apps.danh_muc.models import (HangHoa, KhachHang, Kho, NhaCungCap,
                                  NhomHang, TaiKhoanKeToan, ViTriKho)
from apps.so_cai.periods import guard_accounting_period_error

from .models import (KiemKe, KiemKe_CT, MucTonKho, PhieuDieuChinhKiemKe,
                     PhieuDieuChinhKiemKe_CT, PhieuNhap, PhieuNhap_CT,
                     PhieuXuat, PhieuXuat_CT, TonKho, TonKhoViTri)

SLOT_TAG_REGEX = re.compile(r"\[SLOT=(?P<size>[1-5])\]")
SUC_CHUA_O_MAC_DINH = 50


def _parse_money_input(value, default=Decimal('0')):
    if value in (None, ''):
        return default
    try:
        raw = str(value).strip()
        if not raw:
            return default

        if ',' in raw and '.' in raw:
            if raw.rfind(',') > raw.rfind('.'):
                raw = raw.replace('.', '').replace(',', '.')
            else:
                raw = raw.replace(',', '')
        elif ',' in raw:
            if raw.count(',') == 1 and len(raw.split(',')[1]) <= 2:
                raw = raw.replace(',', '.')
            else:
                raw = raw.replace(',', '')
        elif '.' in raw:
            if raw.count('.') > 1:
                raw = raw.replace('.', '')
            else:
                left, right = raw.split('.', 1)
                if right.isdigit() and len(right) == 3 and left.replace('-', '').isdigit():
                    raw = left + right

        return Decimal(raw)
    except Exception:
        return default


def _can_access_inventory(user):
    return user.is_superuser or user.is_staff or user.has_perm('kho.view_tonkho')


def _recalculate_weighted_average(kho_id=None, hang_hoa_id=None):
    nhap_ct_qs = PhieuNhap_CT.objects.select_related('phieu_nhap').filter(
        phieu_nhap__trang_thai__in=('2', '3')
    )
    xuat_ct_qs = PhieuXuat_CT.objects.select_related('phieu_xuat').filter(
        phieu_xuat__trang_thai__in=('2', '3')
    )

    if kho_id:
        nhap_ct_qs = nhap_ct_qs.filter(phieu_nhap__kho_id=kho_id)
        xuat_ct_qs = xuat_ct_qs.filter(phieu_xuat__kho_id=kho_id)
    if hang_hoa_id:
        nhap_ct_qs = nhap_ct_qs.filter(hang_hoa_id=hang_hoa_id)
        xuat_ct_qs = xuat_ct_qs.filter(hang_hoa_id=hang_hoa_id)

    movements = []
    for ct in nhap_ct_qs.order_by('phieu_nhap__ngay_chung_tu', 'phieu_nhap_id', 'id'):
        movements.append({
            'type': 'N',
            'kho_id': ct.phieu_nhap.kho_id,
            'hang_hoa_id': ct.hang_hoa_id,
            'date': ct.phieu_nhap.ngay_chung_tu,
            'qty': int(ct.so_luong_nhan or 0),
            'unit_price': Decimal(ct.don_gia or 0),
            'obj': ct,
        })

    for ct in xuat_ct_qs.order_by('phieu_xuat__ngay_chung_tu', 'phieu_xuat_id', 'id'):
        movements.append({
            'type': 'X',
            'kho_id': ct.phieu_xuat.kho_id,
            'hang_hoa_id': ct.hang_hoa_id,
            'date': ct.phieu_xuat.ngay_chung_tu,
            'qty': int(ct.so_luong or 0),
            'unit_price': Decimal(0),
            'obj': ct,
        })

    movements.sort(key=lambda m: (m['date'], 0 if m['type'] == 'N' else 1, m['obj'].id))

    state = {}
    touched_keys = set()
    errors = []

    for mv in movements:
        key = (mv['kho_id'], mv['hang_hoa_id'])
        touched_keys.add(key)
        if key not in state:
            state[key] = {'qty': Decimal(0), 'value': Decimal(0)}
        st = state[key]

        if mv['qty'] <= 0:
            continue

        if mv['type'] == 'N':
            st['qty'] += Decimal(mv['qty'])
            st['value'] += Decimal(mv['qty']) * mv['unit_price']
            continue

        if st['qty'] <= 0 or st['qty'] < Decimal(mv['qty']):
            errors.append({
                'kho_id': mv['kho_id'],
                'hang_hoa_id': mv['hang_hoa_id'],
                'so_luong_xuat': mv['qty'],
                'so_luong_hien_co': int(st['qty']),
            })
            continue

        avg = (st['value'] / st['qty']) if st['qty'] > 0 else Decimal(0)
        mv['obj'].gia_von = round(avg, 0)
        mv['obj'].tong_gia_von = round(avg * Decimal(mv['qty']), 0)
        mv['obj'].save(update_fields=['gia_von', 'tong_gia_von'])

        st['qty'] -= Decimal(mv['qty'])
        st['value'] -= avg * Decimal(mv['qty'])

    if errors:
        return False, errors, 0

    updated = 0
    for kho_key, hang_key in touched_keys:
        st = state.get((kho_key, hang_key), {'qty': Decimal(0), 'value': Decimal(0)})
        qty = int(st['qty'])
        if qty > 0:
            avg = round(st['value'] / Decimal(qty), 0)
        else:
            avg = Decimal(0)

        ton_obj, _ = TonKho.objects.get_or_create(
            kho_id=kho_key,
            hang_hoa_id=hang_key,
            defaults={'so_luong': 0, 'so_luong_loi': 0, 'gia_von_tb': 0},
        )
        ton_obj.so_luong = qty
        ton_obj.gia_von_tb = avg
        ton_obj.save(update_fields=['so_luong', 'gia_von_tb', 'ngay_cap_nhat'])
        updated += 1

    for phieu in PhieuXuat.objects.filter(trang_thai__in=('2', '3')):
        if kho_id and phieu.kho_id != int(kho_id):
            continue
        tong = phieu.chi_tiet.aggregate(t=Sum('tong_gia_von'))['t'] or 0
        phieu.tong_gia_von = tong
        phieu.save(update_fields=['tong_gia_von'])

    return True, [], updated


def _parse_period_month(thang_raw):
    """Parse tham số tháng (YYYY-MM) thành ngày đầu/kết thúc kỳ."""
    today = timezone.localdate()
    if not thang_raw:
        month_start = today.replace(day=1)
    else:
        try:
            y_str, m_str = str(thang_raw).split('-', 1)
            year = int(y_str)
            month = int(m_str)
            month_start = date(year, month, 1)
        except Exception:
            month_start = today.replace(day=1)

    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end = next_month - timedelta(days=1)
    return month_start, month_end, month_start.strftime('%Y-%m')


def _build_weighted_average_input_rows(kho_id=None, hang_hoa_id=None, start_date=None, end_date=None):
    """Tổng hợp đầu vào tính giá theo kỳ: tồn đầu, nhập kỳ, xuất kỳ, tồn cuối."""
    nhap_ct_qs = PhieuNhap_CT.objects.select_related(
        'phieu_nhap', 'hang_hoa', 'hang_hoa__don_vi_tinh', 'phieu_nhap__kho'
    ).filter(phieu_nhap__trang_thai__in=('2', '3'))
    xuat_ct_qs = PhieuXuat_CT.objects.select_related(
        'phieu_xuat', 'hang_hoa', 'hang_hoa__don_vi_tinh', 'phieu_xuat__kho'
    ).filter(phieu_xuat__trang_thai__in=('2', '3'))

    if kho_id:
        nhap_ct_qs = nhap_ct_qs.filter(phieu_nhap__kho_id=kho_id)
        xuat_ct_qs = xuat_ct_qs.filter(phieu_xuat__kho_id=kho_id)
    if hang_hoa_id:
        nhap_ct_qs = nhap_ct_qs.filter(hang_hoa_id=hang_hoa_id)
        xuat_ct_qs = xuat_ct_qs.filter(hang_hoa_id=hang_hoa_id)

    if end_date:
        nhap_ct_qs = nhap_ct_qs.filter(phieu_nhap__ngay_chung_tu__lte=end_date)
        xuat_ct_qs = xuat_ct_qs.filter(phieu_xuat__ngay_chung_tu__lte=end_date)

    movements = []
    for ct in nhap_ct_qs.order_by('phieu_nhap__ngay_chung_tu', 'phieu_nhap_id', 'id'):
        movements.append({
            'type': 'N',
            'kho_id': ct.phieu_nhap.kho_id,
            'kho_ma': ct.phieu_nhap.kho.ma_kho,
            'hang_hoa_id': ct.hang_hoa_id,
            'hang_ma': ct.hang_hoa.ma_hang,
            'hang_ten': ct.hang_hoa.ten_hang,
            'dvt': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa.don_vi_tinh else '',
            'date': ct.phieu_nhap.ngay_chung_tu,
            'qty': Decimal(int(ct.so_luong_nhan or 0)),
            'value': Decimal(int(ct.thanh_tien or 0)),
        })

    for ct in xuat_ct_qs.order_by('phieu_xuat__ngay_chung_tu', 'phieu_xuat_id', 'id'):
        movements.append({
            'type': 'X',
            'kho_id': ct.phieu_xuat.kho_id,
            'kho_ma': ct.phieu_xuat.kho.ma_kho,
            'hang_hoa_id': ct.hang_hoa_id,
            'hang_ma': ct.hang_hoa.ma_hang,
            'hang_ten': ct.hang_hoa.ten_hang,
            'dvt': ct.hang_hoa.don_vi_tinh.ten if ct.hang_hoa.don_vi_tinh else '',
            'date': ct.phieu_xuat.ngay_chung_tu,
            'qty': Decimal(int(ct.so_luong or 0)),
            'value': Decimal(int(ct.tong_gia_von or 0)),
        })

    movements.sort(key=lambda m: (m['date'], 0 if m['type'] == 'N' else 1))

    state = {}
    period_has_data = False

    for mv in movements:
        key = (mv['kho_id'], mv['hang_hoa_id'])
        if key not in state:
            state[key] = {
                'kho_ma': mv['kho_ma'],
                'hang_ma': mv['hang_ma'],
                'hang_ten': mv['hang_ten'],
                'dvt': mv['dvt'],
                'opening_qty': Decimal(0),
                'opening_value': Decimal(0),
                'in_qty': Decimal(0),
                'in_value': Decimal(0),
                'out_qty': Decimal(0),
                'out_value': Decimal(0),
                'closing_qty': Decimal(0),
                'closing_value': Decimal(0),
                '_running_qty': Decimal(0),
                '_running_value': Decimal(0),
                '_opening_captured': False,
                '_period_activity': False,
            }

        row = state[key]
        in_period = (start_date is None or mv['date'] >= start_date) and (end_date is None or mv['date'] <= end_date)

        if in_period and not row['_opening_captured']:
            row['opening_qty'] = row['_running_qty']
            row['opening_value'] = row['_running_value']
            row['_opening_captured'] = True

        if mv['type'] == 'N':
            row['_running_qty'] += mv['qty']
            row['_running_value'] += mv['value']
            if in_period:
                row['in_qty'] += mv['qty']
                row['in_value'] += mv['value']
                row['_period_activity'] = True
                period_has_data = True
        else:
            avg = (row['_running_value'] / row['_running_qty']) if row['_running_qty'] > 0 else Decimal(0)
            out_value = mv['value'] if mv['value'] > 0 else (avg * mv['qty'])
            row['_running_qty'] -= mv['qty']
            row['_running_value'] -= out_value
            if in_period:
                row['out_qty'] += mv['qty']
                row['out_value'] += out_value
                row['_period_activity'] = True
                period_has_data = True

        row['closing_qty'] = row['_running_qty']
        row['closing_value'] = row['_running_value']

    rows = []
    for (_, _), row in state.items():
        if hang_hoa_id or kho_id or row['_period_activity']:
            rows.append({
                'kho_ma': row['kho_ma'],
                'ma_hang': row['hang_ma'],
                'ten_hang': row['hang_ten'],
                'dvt': row['dvt'],
                'sl_dau_ky': int(row['opening_qty']),
                'gt_dau_ky': int(round(row['opening_value'], 0)),
                'sl_nhap': int(row['in_qty']),
                'gt_nhap': int(round(row['in_value'], 0)),
                'sl_xuat': int(row['out_qty']),
                'gt_xuat': int(round(row['out_value'], 0)),
                'sl_cuoi_ky': int(round(row['closing_qty'], 0)),
                'gt_cuoi_ky': int(round(row['closing_value'], 0)),
            })

    rows.sort(key=lambda r: (r['kho_ma'], r['ma_hang']))
    return rows, period_has_data


def _validate_weighted_average_input(kho_id=None, hang_hoa_id=None, start_date=None, end_date=None):
    """Kiểm tra dữ liệu đầu vào hợp lệ trước khi tính giá."""
    nhap_bad = PhieuNhap_CT.objects.filter(phieu_nhap__trang_thai__in=('2', '3'))
    xuat_bad = PhieuXuat_CT.objects.filter(phieu_xuat__trang_thai__in=('2', '3'))

    if kho_id:
        nhap_bad = nhap_bad.filter(phieu_nhap__kho_id=kho_id)
        xuat_bad = xuat_bad.filter(phieu_xuat__kho_id=kho_id)
    if hang_hoa_id:
        nhap_bad = nhap_bad.filter(hang_hoa_id=hang_hoa_id)
        xuat_bad = xuat_bad.filter(hang_hoa_id=hang_hoa_id)
    if start_date and end_date:
        nhap_bad = nhap_bad.filter(phieu_nhap__ngay_chung_tu__range=[start_date, end_date])
        xuat_bad = xuat_bad.filter(phieu_xuat__ngay_chung_tu__range=[start_date, end_date])

    has_bad_nhap = nhap_bad.filter(Q(so_luong_nhan__lte=0) | Q(don_gia__lt=0)).exists()
    has_bad_xuat = xuat_bad.filter(so_luong__lte=0).exists()
    return not (has_bad_nhap or has_bad_xuat)


def _extract_so_don_from_ghi_chu(ghi_chu):
    text = (ghi_chu or '').strip()
    if not text:
        return None
    m = re.search(r'xuất từ đơn\s+(.+?)(?:\s+-|$)', text, flags=re.IGNORECASE)
    if not m:
        return None
    return (m.group(1) or '').strip() or None


def _phieu_xuat_has_related_documents(phieu):
    if phieu.loai_xuat != 'ban_hang':
        return False

    from apps.ban_hang.models import DonBan

    so_don = _extract_so_don_from_ghi_chu(phieu.ghi_chu)
    if so_don:
        return DonBan.objects.filter(so_don=so_don, trang_thai__in=('2', '3')).exists()
    return False


def _check_kiem_ke_diff(kho_id, hang_hoa_ids):
    """
    BR12.x.5: Kiểm tra có chênh lệch kiểm kê chưa xử lý đối với những mặt hàng này không.
    Trả về list hàng hóa có chênh lệch chưa xử lý (phiếu điều chỉnh chưa duyệt).
    """
    if not hang_hoa_ids or not kho_id:
        return []
    
    # Tìm kiểm kê chờ điều chỉnh (trang_thai = 2)
    pending_kk = list(
        KiemKe.objects
        .filter(kho_id=kho_id, trang_thai='2')
        .filter(chi_tiet__hang_hoa_id__in=hang_hoa_ids)
        .filter(chi_tiet__chenh_lech__ne=0) if False else  # ORM
        KiemKe.objects
        .filter(kho_id=kho_id, trang_thai='2')
        .prefetch_related('chi_tiet')
    )
    
    conflicted_items = []
    for kk in pending_kk:
        for ct in kk.chi_tiet.all():
            if int(ct.hang_hoa_id) in hang_hoa_ids and int(ct.chenh_lech or 0) != 0:
                # Kiểm tra xem có phiếu điều chỉnh được duyệt chưa
                phieu_dc = getattr(kk, 'phieu_dieu_chinh', None)
                if not phieu_dc or phieu_dc.trang_thai == '1':  # Chưa duyệt
                    conflicted_items.append({
                        'hang_hoa_id': ct.hang_hoa_id,
                        'ma_hang': ct.hang_hoa.ma_hang,
                        'ten_hang': ct.hang_hoa.ten_hang,
                        'chenh_lech': ct.chenh_lech,
                        'kiem_ke_so': kk.ma_phieu,
                    })
    
    return conflicted_items


def _can_delete_vouchers(user):
    return bool(user and (user.is_superuser or user.is_staff))


def _has_kiem_ke_dieu_chinh_references(phieu_dc):
    note_tag = f'[KK_DC:{phieu_dc.pk}]'
    has_nhap_ref = PhieuNhap.objects.filter(ghi_chu__icontains=note_tag).exists()
    if has_nhap_ref:
        return True
    has_xuat_ref = PhieuXuat.objects.filter(ghi_chu__icontains=note_tag).exists()
    return has_xuat_ref


def _rollback_phieu_nhap_inventory(phieu):
    if phieu.trang_thai not in ('2', '3'):
        return True, ''

    for ct in phieu.chi_tiet.select_related('hang_hoa'):
        ton = TonKho.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=phieu.kho).first()
        if not ton:
            continue
        ton.so_luong = int(ton.so_luong or 0) - int(ct.so_luong_nhan or 0)
        if ton.so_luong < 0:
            ton.so_luong = 0
        ton.save(update_fields=['so_luong', 'ngay_cap_nhat'])

        con_lai_can_tru = int(ct.so_luong_nhan or 0)
        vt_rows = list(
            TonKhoViTri.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=phieu.kho, so_luong__gt=0)
            .order_by('-so_luong', 'id')
        )
        for vt in vt_rows:
            if con_lai_can_tru <= 0:
                break
            tru = min(int(vt.so_luong or 0), con_lai_can_tru)
            vt.so_luong = int(vt.so_luong or 0) - tru
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])
            con_lai_can_tru -= tru

    phieu.trang_thai = '1'
    phieu.save(update_fields=['trang_thai'])
    return True, ''


def _rollback_phieu_xuat_inventory(phieu):
    if phieu.trang_thai not in ('2', '3'):
        return True, ''

    for ct in phieu.chi_tiet.select_related('hang_hoa'):
        ton, _ = TonKho.objects.select_for_update().get_or_create(
            hang_hoa=ct.hang_hoa,
            kho=phieu.kho,
            defaults={'so_luong': 0, 'gia_von_tb': ct.gia_von or 0},
        )
        ton.so_luong = int(ton.so_luong or 0) + int(ct.so_luong or 0)
        ton.save(update_fields=['so_luong', 'ngay_cap_nhat'])
        
        vt = TonKhoViTri.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=phieu.kho).exclude(vi_tri__ma_vi_tri__startswith='HL-').first()
        if not vt:
            vi_tri_kho = ViTriKho.objects.filter(kho=phieu.kho, trang_thai='hoat_dong').first()
            if vi_tri_kho:
                vt, _ = TonKhoViTri.objects.get_or_create(
                    hang_hoa=ct.hang_hoa, kho=phieu.kho, vi_tri=vi_tri_kho, defaults={'so_luong': 0}
                )
        if vt:
            vt.so_luong = int(vt.so_luong or 0) + int(ct.so_luong or 0)
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])

    phieu.trang_thai = '1'
    phieu.save(update_fields=['trang_thai'])
    return True, ''


def _xac_nhan_phieu_xuat(phieu):
    if phieu.trang_thai != '1':
        return False, 'Chỉ phiếu nháp (bước 1) mới được xác nhận ghi Sổ kho'

    # BR12.x.5: Kiểm tra chênh lệch kiểm kê trước khi xác nhận xuất
    hang_ids = list(phieu.chi_tiet.values_list('hang_hoa_id', flat=True))
    if hang_ids:
        conflicted = _check_kiem_ke_diff(phieu.kho_id, hang_ids)
        if conflicted:
            items_str = ', '.join([f"{c['ma_hang']}" for c in conflicted])
            return False, f'Hàng hóa {items_str} đang có chênh lệch kiểm kê chưa được xử lý. Hãy duyệt phiếu điều chỉnh kiểm kê trước.'

    tong_gv = 0
    for ct in phieu.chi_tiet.select_related('hang_hoa'):
        ton = TonKho.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=phieu.kho).first()
        if not ton or ton.so_luong < ct.so_luong:
            return False, f'Không đủ tồn kho để xuất: {ct.hang_hoa.ten_hang}'

        ct.gia_von = ton.gia_von_tb
        ct.tong_gia_von = ton.gia_von_tb * ct.so_luong
        ct.save(update_fields=['gia_von', 'tong_gia_von'])

        ton.so_luong -= ct.so_luong
        ton.save(update_fields=['so_luong', 'ngay_cap_nhat'])
        
        con_lai = int(ct.so_luong or 0)
        vt_rows = list(
            TonKhoViTri.objects.select_for_update().filter(hang_hoa=ct.hang_hoa, kho=phieu.kho, so_luong__gt=0)
            .exclude(vi_tri__ma_vi_tri__startswith='HL-')
            .order_by('-so_luong', 'id')
        )
        for vt in vt_rows:
            if con_lai <= 0: break
            tru = min(int(vt.so_luong or 0), con_lai)
            vt.so_luong = int(vt.so_luong or 0) - tru
            vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])
            con_lai -= tru
        tong_gv += ct.tong_gia_von

    phieu.tong_gia_von = tong_gv
    phieu.trang_thai = '2'
    phieu.save(update_fields=['tong_gia_von', 'trang_thai'])
    return True, ''


def _gen_so_phieu(prefix):
    normalized_prefix = (prefix or '').strip().upper()
    if normalized_prefix == 'XK':
        normalized_prefix = 'PX'

    model_map = {
        'NK': (PhieuNhap, 'so_phieu'),
        'PX': (PhieuXuat, 'so_phieu'),
    }
    model_info = model_map.get(normalized_prefix)
    if model_info:
        model_cls, field_name = model_info
        max_index = 0
        filter_key = {f'{field_name}__istartswith': normalized_prefix}
        for code in model_cls.objects.filter(**filter_key).values_list(field_name, flat=True):
            text = str(code or '').strip().upper()
            if not text.startswith(normalized_prefix):
                continue
            suffix = text[len(normalized_prefix):]
            if suffix.isdigit():
                max_index = max(max_index, int(suffix))
        return f'{normalized_prefix}{max_index + 1:05d}'

    now = timezone.now()
    return f"{normalized_prefix}-{now.strftime('%Y%m%d-%H%M%S')}"


def _hang_slot_size(hang_hoa):
    m = SLOT_TAG_REGEX.search((hang_hoa.ghi_chu or '').upper())
    if not m:
        return 1
    return int(m.group('size'))


def _normalize_loai_nhap(value):
    mapping = {
        '1': '1',
        '2': '2',
        '3': '2',
        'mua_ncc': '1',
        'tra_hang_kh': '2',
        'dieu_chinh': '2',
    }
    return mapping.get(str(value or '').strip(), '1')


def _extract_doi_tra_id_from_ghi_chu(ghi_chu):
    text = str(ghi_chu or '')
    m = re.search(r'\[DOI_TRA_PHIEU:(\d+)\]', text)
    return int(m.group(1)) if m else None


def _build_doi_tra_pending_payload(selected_id=None):
    qs = PhieuTraHang.objects.select_related('khach_hang', 'hoa_don_goc').prefetch_related(
        'chi_tiet__hang_hoa',
        'chi_tiet__kho',
    )
    pending_qs = qs.filter(trang_thai='1')
    if selected_id:
        pending_qs = (pending_qs | qs.filter(pk=selected_id)).distinct()

    payload = []
    for phieu in pending_qs.order_by('-ngay_lap', '-id')[:200]:
        rows = []
        for ct in phieu.chi_tiet.all():
            if int(ct.so_luong or 0) <= 0:
                continue
            rows.append({
                'doi_tra_ct_id': ct.id,
                'hang_id': ct.hang_hoa_id,
                'ma_hang': ct.hang_hoa.ma_hang if ct.hang_hoa else '',
                'ten_hang': ct.hang_hoa.ten_hang if ct.hang_hoa else '',
                'so_luong_nhan': int(ct.so_luong or 0),
                'don_gia': float(ct.don_gia or 0),
                'kho_id': ct.kho_id,
                'tk_no': '156',
                'tk_co': '131',
            })

        if not rows:
            continue

        kh = phieu.khach_hang
        kho_ids = {int(r.get('kho_id') or 0) for r in rows if int(r.get('kho_id') or 0) > 0}
        payload.append({
            'id': phieu.id,
            'so_phieu': phieu.so_phieu,
            'hoa_don': phieu.hoa_don_goc.so_hoa_don if phieu.hoa_don_goc else '',
            'ma_kh': kh.ma_kh if kh else '',
            'ten_kh': kh.ten_kh if kh else '',
            'ngay_lap': phieu.ngay_lap.isoformat() if phieu.ngay_lap else '',
            'ngay_hach_toan': phieu.ngay_hach_toan.isoformat() if phieu.ngay_hach_toan else '',
            'dien_giai': (phieu.dien_giai or '').strip(),
            'ly_do_tra': (phieu.ly_do_tra or '').strip(),
            'kho_id_mac_dinh': next(iter(kho_ids)) if len(kho_ids) == 1 else None,
            'rows': rows,
        })

    return payload


def _de_xuat_vi_tri_nhap(kho_id, hang_hoa, so_luong_nhap):
    rows = ViTriKho.objects.filter(kho_id=kho_id, trang_thai='hoat_dong').order_by('ma_vi_tri')
    if hang_hoa.loai_o_phu_hop:
        rows = rows.filter(loai_o=hang_hoa.loai_o_phu_hop)
    rows = list(rows)

    if not rows:
        return {
            'so_o_can': 0,
            'suc_chua_moi_o': SUC_CHUA_O_MAC_DINH,
            'de_xuat': [],
            'khong_hop_le': ['Không tìm thấy ô phù hợp theo loại ô'],
        }

    muc_chiem_cho = max(1, int(getattr(hang_hoa, 'muc_chiem_cho', 1) or 1))
    tong_dung_luong_can = int(so_luong_nhap) * muc_chiem_cho
    suggestions = []
    invalid_reasons = []

    for vi_tri in rows:
        ton_rows = list(TonKhoViTri.objects.filter(vi_tri=vi_tri).select_related('hang_hoa'))
        dung_luong_da_dung = sum(int(r.so_luong or 0) * max(1, int(getattr(r.hang_hoa, 'muc_chiem_cho', 1) or 1)) for r in ton_rows)
        suc_chua_toi_da = int(vi_tri.suc_chua_toi_da or SUC_CHUA_O_MAC_DINH)
        dung_luong_con_lai = max(0, suc_chua_toi_da - dung_luong_da_dung)
        so_luong_toi_da_co_the_nhap = dung_luong_con_lai // muc_chiem_cho
        if so_luong_toi_da_co_the_nhap <= 0:
            continue

        so_luong_cung_ma = sum(int(r.so_luong or 0) for r in ton_rows if r.hang_hoa_id == hang_hoa.id)
        uu_tien = 1 if so_luong_cung_ma > 0 else 0
        suggestions.append({
            'ma_vi_tri': vi_tri.ma_vi_tri,
            'vi_tri_id': vi_tri.id,
            'dung_luong_con_lai': dung_luong_con_lai,
            'so_luong_toi_da_co_the_nhap': int(so_luong_toi_da_co_the_nhap),
            'uu_tien': uu_tien,
        })

    suggestions.sort(key=lambda x: (-x['uu_tien'], -x['so_luong_toi_da_co_the_nhap'], x['ma_vi_tri']))

    con_lai = int(so_luong_nhap)
    de_xuat = []
    for row in suggestions:
        if con_lai <= 0:
            break
        phan_bo = min(con_lai, row['so_luong_toi_da_co_the_nhap'])
        if phan_bo <= 0:
            continue
        de_xuat.append({
            'ma_vi_tri': row['ma_vi_tri'],
            'vi_tri_id': row['vi_tri_id'],
            'de_xuat_so_luong': phan_bo,
            'dung_luong_con_lai': row['dung_luong_con_lai'],
        })
        con_lai -= phan_bo

    so_o_can = (tong_dung_luong_can + SUC_CHUA_O_MAC_DINH - 1) // SUC_CHUA_O_MAC_DINH
    return {
        'so_o_can': so_o_can,
        'suc_chua_moi_o': SUC_CHUA_O_MAC_DINH,
        'muc_chiem_cho': muc_chiem_cho,
        'tong_dung_luong_can': tong_dung_luong_can,
        'de_xuat': de_xuat,
        'con_lai_chua_phan_bo': con_lai,
        'khong_hop_le': invalid_reasons[:5],
    }


def _ghi_nhan_phan_bo_vi_tri(phieu_nhap):
    so_dong = 0
    for ct in phieu_nhap.chi_tiet.select_related('hang_hoa'):
        so_luong = int(ct.so_luong_nhan or 0)
        if so_luong <= 0:
            continue

        ket_qua = _de_xuat_vi_tri_nhap(phieu_nhap.kho_id, ct.hang_hoa, so_luong)
        for de_xuat in ket_qua['de_xuat']:
            vi_tri = ViTriKho.objects.filter(kho_id=phieu_nhap.kho_id, ma_vi_tri=de_xuat['ma_vi_tri']).first()
            if not vi_tri:
                continue
            obj, _ = TonKhoViTri.objects.get_or_create(
                hang_hoa=ct.hang_hoa,
                kho_id=phieu_nhap.kho_id,
                vi_tri=vi_tri,
                defaults={'so_luong': 0},
            )
            obj.so_luong += int(de_xuat['de_xuat_so_luong'])
            obj.save(update_fields=['so_luong', 'ngay_cap_nhat'])
            so_dong += 1
    return so_dong


@login_required
def goi_y_vi_tri_nhap(request):
    kho_id = request.GET.get('kho_id')
    hang_id = request.GET.get('hang_id')
    so_luong = request.GET.get('so_luong')

    try:
        so_luong = int(so_luong or 0)
    except (TypeError, ValueError):
        so_luong = 0

    if not kho_id or not hang_id or so_luong <= 0:
        return JsonResponse({'error': 'Thiếu dữ liệu đầu vào hợp lệ'}, status=400)

    hang_hoa = get_object_or_404(HangHoa, pk=hang_id)
    ket_qua = _de_xuat_vi_tri_nhap(kho_id, hang_hoa, so_luong)
    return JsonResponse({
        'ma_hang': hang_hoa.ma_hang,
        'ten_hang': hang_hoa.ten_hang,
        'so_luong_nhap': so_luong,
        **ket_qua,
    })


@login_required
def chi_tiet_vi_tri(request, pk):
    vi_tri = get_object_or_404(ViTriKho.objects.select_related('kho'), pk=pk)
    ton_rows = list(
        TonKhoViTri.objects
        .filter(vi_tri=vi_tri, so_luong__gt=0, hang_hoa__trang_thai='dang_ban')
        .select_related('hang_hoa')
        .order_by('-so_luong')
    )

    items = []
    dung_luong_da_dung = 0
    for row in ton_rows:
        muc_chiem_cho = max(1, int(getattr(row.hang_hoa, 'muc_chiem_cho', 1) or 1))
        dung_luong_su_dung = int(row.so_luong or 0) * muc_chiem_cho
        dung_luong_da_dung += dung_luong_su_dung
        items.append({
            'ma_hang': row.hang_hoa.ma_hang,
            'ten_hang': row.hang_hoa.ten_hang,
            'so_luong': int(row.so_luong or 0),
            'muc_chiem_cho': muc_chiem_cho,
            'dung_luong_su_dung': dung_luong_su_dung,
        })

    suc_chua_toi_da = int(vi_tri.suc_chua_toi_da or SUC_CHUA_O_MAC_DINH)
    dung_luong_con_lai = max(0, suc_chua_toi_da - dung_luong_da_dung)
    if dung_luong_con_lai <= 0:
        trang_thai = 'DA_DAY'
    elif dung_luong_con_lai <= 10:
        trang_thai = 'GAN_DAY'
    else:
        trang_thai = 'CON_TRONG'

    return JsonResponse({
        'vi_tri_id': vi_tri.id,
        'ma_vi_tri': vi_tri.ma_vi_tri,
        'ten_kho': vi_tri.kho.ten_kho,
        'suc_chua_toi_da': suc_chua_toi_da,
        'dung_luong_da_dung': dung_luong_da_dung,
        'dung_luong_con_lai': dung_luong_con_lai,
        'trang_thai': trang_thai,
        'items': items,
    })


# ─── TỒN KHO ────────────────────────────────────────────────
@login_required
def ton_kho_list(request):
    if not _can_access_inventory(request.user):
        messages.error(request, 'Bạn không có quyền xem tồn kho')
        return redirect('dashboard')

    kho_id = request.GET.get('kho', '').strip()
    q = request.GET.get('q', '').strip()
    nhom_hang_id = request.GET.get('nhom_hang', '').strip()
    vi_tri = request.GET.get('vi_tri', '').strip()
    trang_thai_ton = request.GET.get('trang_thai_ton', '').strip()

    items = TonKho.objects.select_related(
        'hang_hoa', 'kho', 'hang_hoa__nhom_hang', 'hang_hoa__don_vi_tinh', 'hang_hoa__thuong_hieu'
    ).filter(hang_hoa__trang_thai='dang_ban')

    if kho_id:
        items = items.filter(kho_id=kho_id)
    if q:
        items = items.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))
    if nhom_hang_id:
        items = items.filter(hang_hoa__nhom_hang_id=nhom_hang_id)

    _resync_tonkho_from_vitri_scope(kho_id or None)

    items = list(items.order_by('hang_hoa__ma_hang', 'kho__ma_kho'))

    vi_tri_map = {}
    ton_vitri_qs = TonKhoViTri.objects.select_related('vi_tri').filter(so_luong__gt=0)
    if vi_tri:
        ton_vitri_qs = ton_vitri_qs.filter(vi_tri__ma_vi_tri__icontains=vi_tri)
    for row in ton_vitri_qs:
        key = (row.hang_hoa_id, row.kho_id)
        vi_tri_map.setdefault(key, []).append(row.vi_tri.ma_vi_tri)

    muc_ton_map = {
        (mt.hang_hoa_id, mt.kho_id): mt
        for mt in MucTonKho.objects.select_related('hang_hoa', 'kho')
    }

    filtered_items = []
    for item in items:
        key = (item.hang_hoa_id, item.kho_id)
        vi_tri_list = sorted(set(vi_tri_map.get(key, [])))
        if vi_tri and not vi_tri_list:
            continue

        item.vi_tri_hien_tai = ', '.join(vi_tri_list) if vi_tri_list else '-'
        item.so_luong_loi = int(item.so_luong_loi or 0)
        item.so_luong_thuong = max(0, int(item.so_luong or 0) - item.so_luong_loi)
        item.tong_gia_tri = int((item.so_luong_thuong + item.so_luong_loi) * (item.gia_von_tb or 0))

        muc_ton = muc_ton_map.get(key)
        item.ton_toi_thieu_ap_dung = int(muc_ton.ton_toi_thieu) if muc_ton else int(item.hang_hoa.ton_toi_thieu or 0)
        item.ton_toi_da_ap_dung = int(muc_ton.ton_toi_da) if (muc_ton and muc_ton.ton_toi_da is not None) else int(item.hang_hoa.ton_toi_da or 0)

        tong_ton = int(item.so_luong or 0)
        if tong_ton <= 0:
            item.trang_thai_ton = 'het_hang'
        elif tong_ton <= item.ton_toi_thieu_ap_dung:
            item.trang_thai_ton = 'sap_het'
        elif item.ton_toi_da_ap_dung > 0 and tong_ton > item.ton_toi_da_ap_dung:
            item.trang_thai_ton = 'vuot_toi_da'
        else:
            item.trang_thai_ton = 'binh_thuong'

        if trang_thai_ton and item.trang_thai_ton != trang_thai_ton:
            continue

        filtered_items.append(item)

    tong_gia_tri = sum(i.tong_gia_tri for i in filtered_items)
    so_hang_can_nhap = sum(1 for i in filtered_items if i.trang_thai_ton in ('het_hang', 'sap_het'))
    tong_chung_lo = len(filtered_items)
    so_hang_het = sum(1 for i in filtered_items if i.trang_thai_ton == 'het_hang')
    so_hang_sap_het = sum(1 for i in filtered_items if i.trang_thai_ton == 'sap_het')
    so_hang_vuot_toi_da = sum(1 for i in filtered_items if i.trang_thai_ton == 'vuot_toi_da')
    tong_hang_loi = sum(int(i.so_luong_loi or 0) for i in filtered_items)
    tong_so_luong_ton = sum(int(i.so_luong or 0) for i in filtered_items)
    ty_le_hang_loi = round((tong_hang_loi / tong_so_luong_ton) * 100, 2) if tong_so_luong_ton > 0 else 0
    tong_sku_canh_bao = so_hang_het + so_hang_sap_het + so_hang_vuot_toi_da

    if request.GET.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ton kho'
        ws.append([
            'Mã hàng',
            'Tên hàng',
            'Nhóm hàng',
            'Kho',
            'Vị trí/Kệ',
            'Đơn vị tính',
            'SL hàng thường',
            'SL hàng lỗi',
            'Giá vốn bình quân',
            'Giá trị tồn kho',
            'Trạng thái tồn kho',
        ])

        trang_thai_label = {
            'het_hang': 'Hết hàng',
            'sap_het': 'Sắp hết',
            'vuot_toi_da': 'Vượt tối đa',
            'binh_thuong': 'Bình thường',
        }

        for item in filtered_items:
            ws.append([
                item.hang_hoa.ma_hang,
                item.hang_hoa.ten_hang,
                item.hang_hoa.nhom_hang.ten_nhom if item.hang_hoa.nhom_hang else '',
                item.kho.ten_kho,
                item.vi_tri_hien_tai,
                item.hang_hoa.don_vi_tinh.ten if item.hang_hoa.don_vi_tinh else '',
                int(item.so_luong_thuong or 0),
                int(item.so_luong_loi or 0),
                int(item.gia_von_tb or 0),
                int(item.tong_gia_tri or 0),
                trang_thai_label.get(item.trang_thai_ton, item.trang_thai_ton),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="ton_kho_hien_tai.xlsx"'
        wb.save(response)
        return response

    context = {
        'items': filtered_items,
        'kho_options': [
            {'id': str(k.pk), 'label': k.ten_kho}
            for k in Kho.objects.filter(trang_thai=True).order_by('ten_kho')
        ],
        'nhom_hang_options': [
            {'id': str(row['nhom_hang__id']), 'label': row['nhom_hang__ten_nhom']}
            for row in HangHoa.objects.exclude(nhom_hang__isnull=True)
            .values('nhom_hang__id', 'nhom_hang__ten_nhom')
            .distinct()
            .order_by('nhom_hang__ten_nhom')
        ],
        'vi_tri_options': [
            {'label': vi_tri_ma}
            for vi_tri_ma in TonKhoViTri.objects.select_related('vi_tri')
            .filter(so_luong__gt=0)
            .values_list('vi_tri__ma_vi_tri', flat=True)
            .distinct()
            .order_by('vi_tri__ma_vi_tri')
        ],
        'kho_filter': kho_id,
        'q': q,
        'nhom_hang_filter': nhom_hang_id,
        'vi_tri_filter': vi_tri,
        'trang_thai_ton_filter': trang_thai_ton,
        'tong_gia_tri': int(tong_gia_tri),
        'so_hang_can_nhap': so_hang_can_nhap,
        'tong_chung_lo': tong_chung_lo,
        'so_hang_het': so_hang_het,
        'so_hang_sap_het': so_hang_sap_het,
        'so_hang_vuot_toi_da': so_hang_vuot_toi_da,
        'tong_hang_loi': tong_hang_loi,
        'tong_so_luong_ton': tong_so_luong_ton,
        'ty_le_hang_loi': ty_le_hang_loi,
        'tong_sku_canh_bao': tong_sku_canh_bao,
        'tong_kho': Kho.objects.filter(trang_thai=True).count(),
        'page_title': 'Tồn kho hiện tại',
        'active_menu': 'ton_kho',
    }
    return render(request, 'kho/ton_kho_list.html', context)


@login_required
def thiet_lap_muc_ton_kho(request):
    if not (request.user.is_superuser or request.user.is_staff or request.user.has_perm('kho.change_tonkho')):
        messages.error(request, 'Bạn không có quyền thiết lập mức tồn kho')
        return redirect('ton_kho_list')

    if request.method != 'POST':
        return redirect('ton_kho_list')

    hang_hoa_id = request.POST.get('hang_hoa_id')
    kho_id = request.POST.get('kho_id')
    ton_toi_thieu = request.POST.get('ton_toi_thieu', '0').strip()
    ton_toi_da = request.POST.get('ton_toi_da', '').strip()
    ghi_chu = request.POST.get('ghi_chu', '').strip()

    if not hang_hoa_id or not kho_id:
        messages.error(request, 'Vui lòng chọn hàng hóa và kho áp dụng')
        return redirect('ton_kho_list')

    try:
        ton_min = int(ton_toi_thieu or '0')
    except ValueError:
        messages.error(request, 'Vui lòng kiểm tra lại thông tin')
        return redirect('ton_kho_list')

    ton_max = None
    if ton_toi_da:
        try:
            ton_max = int(ton_toi_da)
        except ValueError:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin')
            return redirect('ton_kho_list')

    if ton_min < 0:
        messages.error(request, 'Tồn tối thiểu phải lớn hơn hoặc bằng 0')
        return redirect('ton_kho_list')

    if ton_max is not None and ton_max < ton_min:
        messages.error(request, 'Mức tồn tối đa phải lớn hơn hoặc bằng mức tồn tối thiểu')
        return redirect('ton_kho_list')

    MucTonKho.objects.update_or_create(
        hang_hoa_id=hang_hoa_id,
        kho_id=kho_id,
        defaults={
            'ton_toi_thieu': ton_min,
            'ton_toi_da': ton_max,
            'ghi_chu': ghi_chu,
            'nguoi_cap_nhat': request.user,
        }
    )

    messages.success(request, 'Thiết lập mức tồn kho thành công')
    return redirect('ton_kho_list')


# ─── SƠ ĐỒ KHO ──────────────────────────────────────────────
@login_required
def so_do_kho(request):
    kho_sel = request.GET.get('kho_id', '')
    hien_thi = request.GET.get('hien_thi', 'tat_ca')
    if hien_thi not in ('tat_ca', 'co_hang'):
        hien_thi = 'tat_ca'
    kho_all = Kho.objects.filter(trang_thai=True)

    _resync_tonkho_from_vitri_scope(kho_sel or None)

    # Thêm thống kê số mặt hàng cho mỗi kho
    kho_list = []
    for k in kho_all:
        k.so_mat_hang = TonKho.objects.filter(kho=k, so_luong__gt=0).count()
        kho_list.append(k)

    items = TonKho.objects.select_related(
        'hang_hoa', 'kho', 'hang_hoa__nhom_hang',
        'hang_hoa__don_vi_tinh', 'hang_hoa__thuong_hieu'
    ).filter(so_luong__gte=0).order_by('kho__ma_kho', 'hang_hoa__ma_hang')

    if kho_sel:
        items = items.filter(kho_id=kho_sel)

    map_kho_list = [k for k in kho_list if str(k.id) == str(kho_sel)] if kho_sel else kho_list
    kho_maps = []
    ma_pattern = re.compile(r'^(?P<side>[A-Z0-9]+)-D(?P<day>\d+)-K(?P<ke>\d+)-T(?P<tang>\d+)$')
    for kho in map_kho_list:
        vi_tri_hang_loi = (
            ViTriKho.objects
            .filter(kho=kho)
            .filter(
                Q(ma_vi_tri__icontains='LOI')
                | Q(mo_ta__icontains='hàng lỗi')
                | Q(mo_ta__icontains='hang loi')
                | Q(mo_ta__icontains='lỗi')
                | Q(mo_ta__icontains='loi')
            )
            .order_by('ma_vi_tri')
            .first()
        )
        vi_tri_hang_loi_created = False
        if vi_tri_hang_loi is None:
            base_code = f'HL-{kho.ma_kho}'.upper()
            counter = 1
            while True:
                ma_vi_tri = f'{base_code}-{counter:02d}'
                if len(ma_vi_tri) > 20:
                    ma_vi_tri = f'HL-{counter:02d}'
                if not ViTriKho.objects.filter(kho=kho, ma_vi_tri=ma_vi_tri).exists():
                    vi_tri_hang_loi = ViTriKho.objects.create(
                        kho=kho,
                        ma_vi_tri=ma_vi_tri,
                        mo_ta='Vị trí hàng lỗi',
                        loai_o='nho',
                        suc_chua_toi_da=500,
                        trang_thai='hoat_dong',
                    )
                    vi_tri_hang_loi_created = True
                    break
                counter += 1
        elif int(vi_tri_hang_loi.suc_chua_toi_da or 0) != 500:
            vi_tri_hang_loi.suc_chua_toi_da = 500
            vi_tri_hang_loi.save(update_fields=['suc_chua_toi_da'])

        vi_tri_list = list(
            ViTriKho.objects.filter(kho=kho).order_by('ma_vi_tri')
        )
        vi_tri_id_map = {(vt.ma_vi_tri or '').upper(): vt.id for vt in vi_tri_list}
        ton_items = list(
            TonKho.objects.select_related('hang_hoa', 'hang_hoa__don_vi_tinh')
            .filter(kho=kho, so_luong__gt=0)
            .order_by('-so_luong', 'hang_hoa__ma_hang')
        )
        tong_so_luong_thuong = sum(int(t.so_luong or 0) for t in ton_items)
        tong_so_luong_loi = sum(int(t.so_luong_loi or 0) for t in ton_items)

        structure = {}
        for vt in vi_tri_list:
            if vi_tri_hang_loi and vt.id == vi_tri_hang_loi.id:
                continue
            m = ma_pattern.match((vt.ma_vi_tri or '').upper())
            if not m:
                continue

            side = m.group('side')
            day = int(m.group('day'))
            ke = int(m.group('ke'))
            tang = int(m.group('tang'))

            if side not in structure:
                structure[side] = {}
            if day not in structure[side]:
                structure[side][day] = {}
            if ke not in structure[side][day]:
                structure[side][day][ke] = set()

            structure[side][day][ke].add(tang)

        side_blocks = []
        for side, days in sorted(structure.items(), key=lambda x: x[0]):
            day_blocks = []
            total_kes = 0
            total_positions = 0
            max_tang = 0

            for day, kes in sorted(days.items(), key=lambda x: x[0]):
                ke_blocks = []
                for ke, tangs in sorted(kes.items(), key=lambda x: x[0]):
                    tang_list = [1, 2, 3]
                    tang_count = len(tang_list)
                    ke_blocks.append({
                        'ke': ke,
                        'tang_list': tang_list,
                        'tang_count': tang_count,
                    })
                    max_tang = max(max_tang, tang_count)
                    total_positions += tang_count

                total_kes += len(ke_blocks)
                day_blocks.append({
                    'day': day,
                    'ke_blocks': ke_blocks,
                    'ke_count': len(ke_blocks),
                })

            side_blocks.append({
                'side': side,
                'day_blocks': day_blocks,
                'day_count': len(day_blocks),
                'total_kes': total_kes,
                'total_positions': total_positions,
                'max_tang': max_tang,
            })

        ordered_slots = []
        for side, days in sorted(structure.items(), key=lambda x: x[0]):
            for day, kes in sorted(days.items(), key=lambda x: x[0]):
                for ke, tangs in sorted(kes.items(), key=lambda x: x[0]):
                    for tang in sorted(tangs):
                        ordered_slots.append({
                            'side': side,
                            'day': day,
                            'ke': ke,
                            'tang': tang,
                            'ma_vi_tri': f"{side}-D{day:02d}-K{ke:02d}-T{tang:02d}",
                        })

        ton_vitri_rows = list(
            TonKhoViTri.objects.select_related('vi_tri', 'hang_hoa')
            .filter(kho=kho, so_luong__gt=0)
        )
        ton_by_slot = {(r.vi_tri.ma_vi_tri or '').upper(): r for r in ton_vitri_rows}
        overflow_items = []
        occupied_side_blocks = []
        returns_ton_rows = []
        returns_total_qty = 0
        if vi_tri_hang_loi:
            returns_ton_rows = list(
                TonKhoViTri.objects.select_related('hang_hoa')
                .filter(kho=kho, vi_tri=vi_tri_hang_loi, so_luong__gt=0)
                .order_by('-so_luong', 'hang_hoa__ma_hang')
            )
            returns_total_qty = sum(int(r.so_luong or 0) for r in returns_ton_rows)

        display_side_blocks = []
        displayed_count = 0
        for side in side_blocks:
            day_blocks = []
            for day in side['day_blocks']:
                ke_blocks = []
                for ke in day['ke_blocks']:
                    levels = []
                    for tang in ke['tang_list']:
                        ma_vi_tri = f"{side['side']}-D{day['day']:02d}-K{ke['ke']:02d}-T{tang:02d}"
                        ma_vi_tri_key = ma_vi_tri.upper()
                        ton = ton_by_slot.get(ma_vi_tri_key)
                        if hien_thi == 'co_hang' and ton is None:
                            continue
                        levels.append({
                            'tang': tang,
                            'ma_vi_tri': ma_vi_tri,
                            'vi_tri_id': vi_tri_id_map.get(ma_vi_tri_key),
                            'ton': ton,
                        })
                        displayed_count += 1

                    # Giữ khung kệ để dãy không bị mất cột khi không có hàng.
                    if levels or hien_thi == 'tat_ca':
                        ke_blocks.append({
                            'ke': ke['ke'],
                            'levels': levels,
                        })

                # Luôn giữ đủ các dãy để sơ đồ không bị dồn sang 1 dãy.
                day_blocks.append({
                    'day': day['day'],
                    'ke_blocks': ke_blocks,
                })

            display_side_blocks.append({
                'side': side['side'],
                'day_blocks': day_blocks,
            })

        kho_maps.append({
            'kho': kho,
            'vi_tri_hang_loi': vi_tri_hang_loi,
            'vi_tri_hang_loi_created': vi_tri_hang_loi_created,
            'returns_ton_rows': returns_ton_rows,
            'returns_total_qty': returns_total_qty,
            'vi_tri_list': vi_tri_list,
            'ton_items': ton_items[:12],
            'ton_items_total': len(ton_items),
            'tong_so_luong_thuong': tong_so_luong_thuong,
            'tong_so_luong_loi': tong_so_luong_loi,
            'tong_so_luong': tong_so_luong_thuong + tong_so_luong_loi,
            'side_blocks': side_blocks,
            'occupied_side_blocks': occupied_side_blocks,
            'display_side_blocks': display_side_blocks,
            'displayed_count': displayed_count,
            'occupied_count': len(ton_vitri_rows),
            'total_slot_count': len(ordered_slots),
            'overflow_items': overflow_items,
        })

    return render(request, 'kho/so_do_kho.html', {
        'kho_list': kho_list,
        'items': items,
        'kho_maps': kho_maps,
        'kho_sel': kho_sel,
        'hien_thi': hien_thi,
        'page_title': 'Sơ đồ kho',
        'active_menu': 'so_do_kho',
    })


# ─── TÍNH GIÁ XUẤT KHO ──────────────────────────────────────
@login_required
def tinh_gia_xuat(request):
    if not _can_access_inventory(request.user):
        messages.error(request, 'Bạn không có quyền tính giá xuất kho')
        return redirect('dashboard')

    thang_raw = request.GET.get('thang', '')
    kho_filter = request.GET.get('kho', '')
    hang_filter = request.GET.get('hang_hoa', '')
    ky_tu_ngay, ky_den_ngay, thang_hien_tai = _parse_period_month(thang_raw)

    if request.method == 'POST':
        action = (request.POST.get('action') or 'calculate').strip()
        thang_raw = request.POST.get('thang', '').strip()
        ky_tu_ngay, ky_den_ngay, thang_hien_tai = _parse_period_month(thang_raw)
        kho_filter = request.POST.get('kho', '').strip()
        hang_filter = request.POST.get('hang_hoa', '').strip()
        if action == 'cancel':
            messages.info(request, 'Đã hủy thao tác tính giá xuất kho')
            return redirect(
                f"{request.path}?thang={thang_hien_tai}&kho={kho_filter}&hang_hoa={hang_filter}"
            )

        if not _validate_weighted_average_input(
            kho_id=kho_filter or None,
            hang_hoa_id=hang_filter or None,
            start_date=ky_tu_ngay,
            end_date=ky_den_ngay,
        ):
            messages.error(request, 'Dữ liệu tồn kho không hợp lệ, vui lòng kiểm tra lại')
        else:
            input_rows, period_has_data = _build_weighted_average_input_rows(
                kho_id=kho_filter or None,
                hang_hoa_id=hang_filter or None,
                start_date=ky_tu_ngay,
                end_date=ky_den_ngay,
            )
            if not input_rows or not period_has_data:
                messages.warning(request, 'Không có dữ liệu để tính giá xuất kho')
            else:
                try:
                    with transaction.atomic():
                        ok, errors, updated = _recalculate_weighted_average(
                            kho_id=kho_filter or None,
                            hang_hoa_id=hang_filter or None,
                        )
                        if not ok:
                            messages.error(request, 'Không thể tính giá do tồn kho âm')
                            for err in errors[:5]:
                                hang = HangHoa.objects.filter(pk=err['hang_hoa_id']).first()
                                kho = Kho.objects.filter(pk=err['kho_id']).first()
                                messages.error(
                                    request,
                                    f"{hang.ma_hang if hang else err['hang_hoa_id']} - {kho.ma_kho if kho else err['kho_id']}: "
                                    f"xuất {err['so_luong_xuat']} > tồn {err['so_luong_hien_co']}"
                                )
                        elif updated == 0:
                            messages.warning(request, 'Không có dữ liệu để tính giá xuất kho')
                        else:
                            messages.success(request, 'Tính giá bình quân xuất kho thành công')
                except Exception:
                    messages.error(request, 'Không thể tính giá bình quân xuất kho')

    input_rows, period_has_data = _build_weighted_average_input_rows(
        kho_id=kho_filter or None,
        hang_hoa_id=hang_filter or None,
        start_date=ky_tu_ngay,
        end_date=ky_den_ngay,
    )

    items = TonKho.objects.select_related(
        'hang_hoa', 'kho', 'hang_hoa__nhom_hang', 'hang_hoa__don_vi_tinh'
    ).filter(
        so_luong__gt=0,
        hang_hoa__trang_thai='dang_ban',
    ).order_by('hang_hoa__ma_hang')

    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if hang_filter:
        items = items.filter(hang_hoa_id=hang_filter)

    tong_gia_tri_ton = sum(t.so_luong * t.gia_von_tb for t in items)
    thang = ky_tu_ngay.strftime('%m/%Y')

    # Thống kê xuất kho tháng này
    now = timezone.now()
    first_day = now.replace(day=1)
    tong_phieu_xuat = PhieuXuat.objects.filter(
        ngay_chung_tu__gte=first_day.date(), trang_thai__in=('2', '3')
    ).count()
    tong_gia_von_xuat = PhieuXuat.objects.filter(
        ngay_chung_tu__gte=first_day.date(), trang_thai__in=('2', '3')
    ).aggregate(t=Sum('tong_gia_von'))['t'] or 0

    return render(request, 'kho/tinh_gia_xuat.html', {
        'items': items,
        'input_rows': input_rows,
        'period_has_data': period_has_data,
        'ky_tu_ngay': ky_tu_ngay,
        'ky_den_ngay': ky_den_ngay,
        'thang_hien_tai': thang_hien_tai,
        'kho_list': Kho.objects.filter(trang_thai=True),
        'hang_hoa_list': HangHoa.objects.filter(trang_thai='dang_ban').order_by('ma_hang')[:500],
        'kho_filter': kho_filter,
        'hang_filter': hang_filter,
        'tong_gia_tri_ton': int(tong_gia_tri_ton),
        'tong_phieu_xuat': tong_phieu_xuat,
        'tong_gia_von_xuat': int(tong_gia_von_xuat),
        'tong_ton_kho': int(tong_gia_tri_ton),
        'thang': thang,
        'page_title': 'Tính giá xuất kho',
        'active_menu': 'tinh_gia_xuat',
    })


# ─── ĐỐI CHIẾU SỐ LIỆU ──────────────────────────────────────
@login_required
def doi_chieu_so_lieu(request):
    today = date.today()
    tu_ngay = request.GET.get('tu_ngay', today.replace(day=1).isoformat())
    den_ngay = request.GET.get('den_ngay', today.isoformat())
    kho_filter = request.GET.get('kho', '')

    # Tổng giá vốn xuất từ phiếu xuất trong kỳ
    xuat_qs = PhieuXuat.objects.filter(
        ngay_chung_tu__range=[tu_ngay, den_ngay],
        trang_thai__in=('2', '3')
    )
    if kho_filter:
        xuat_qs = xuat_qs.filter(kho_id=kho_filter)
    tong_gia_von_xuat = xuat_qs.aggregate(t=Sum('tong_gia_von'))['t'] or 0

    # Tổng giá trị tồn kho hiện tại
    ton_qs = TonKho.objects.select_related('hang_hoa', 'kho')
    if kho_filter:
        ton_qs = ton_qs.filter(kho_id=kho_filter)
    tong_ton_hien_tai = sum(t.so_luong * t.gia_von_tb for t in ton_qs)

    chenh_lech_tong = int(tong_ton_hien_tai) - int(tong_gia_von_xuat)

    # Chi tiết theo mặt hàng
    rows = []
    for tk in ton_qs.order_by('hang_hoa__ma_hang'):
        gv_xuat = PhieuXuat_CT.objects.filter(
            phieu_xuat__ngay_chung_tu__range=[tu_ngay, den_ngay],
            phieu_xuat__trang_thai__in=('2', '3'),
            hang_hoa=tk.hang_hoa,
        ).aggregate(t=Sum('tong_gia_von'))['t'] or 0

        gv_ton = int(tk.so_luong * tk.gia_von_tb)
        rows.append({
            'ma_hang': tk.hang_hoa.ma_hang,
            'ten_hang': tk.hang_hoa.ten_hang,
            'ten_kho': tk.kho.ten_kho,
            'gv_xuat': int(gv_xuat),
            'gv_ton': gv_ton,
            'chenh_lech': gv_ton - int(gv_xuat),
        })

    return render(request, 'kho/doi_chieu_so_lieu.html', {
        'tu_ngay': tu_ngay,
        'den_ngay': den_ngay,
        'kho_list': Kho.objects.filter(trang_thai=True),
        'kho_filter': kho_filter,
        'tong_gia_von_xuat': int(tong_gia_von_xuat),
        'tong_ton_hien_tai': int(tong_ton_hien_tai),
        'chenh_lech_tong': chenh_lech_tong,
        'rows': rows,
        'page_title': 'Đối chiếu số liệu',
        'active_menu': 'doi_chieu_kho',
    })


# ─── PHIẾU NHẬP KHO ────────────────────────────────────────
@login_required
def phieu_nhap_list(request):
    q = request.GET.get('q', '')
    kho_filter = request.GET.get('kho', '')
    trang_thai_filter = request.GET.get('trang_thai', '')
    items = PhieuNhap.objects.select_related('nha_cung_cap', 'kho')
    if q:
        items = items.filter(Q(so_phieu__icontains=q) |
                             Q(nha_cung_cap__ten_ncc__icontains=q))
    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if trang_thai_filter in ('1', '2', '3'):
        items = items.filter(trang_thai=trang_thai_filter)
    context = {
        'items': items[:50],
        'q': q,
        'kho_list': Kho.objects.filter(trang_thai=True),
        'kho_filter': kho_filter,
        'trang_thai_filter': trang_thai_filter,
        'page_title': 'Phiếu nhập kho',
        'active_menu': 'phieu_nhap',
    }
    return render(request, 'kho/phieu_nhap_list.html', context)


@login_required
def phieu_nhap_them(request):
    if request.method == 'POST':
        data = request.POST
        kho_id = data.get('kho')
        loai_nhap = _normalize_loai_nhap(data.get('loai_nhap', '1'))
        if loai_nhap not in ('1', '2'):
            loai_nhap = '1'
        trang_thai_mong_muon = str(data.get('trang_thai', '1') or '1').strip()
        if trang_thai_mong_muon not in ('1', '2', '3'):
            trang_thai_mong_muon = '1'
        ncc_id = data.get('ncc') or None
        phieu_doi_tra_id = (data.get('phieu_doi_tra_id') or '').strip()
        phieu_doi_tra = None
        if not kho_id:
            messages.error(request, 'Vui lòng chọn kho nhận')
            return redirect('phieu_nhap_them')

        if loai_nhap == '1' and not ncc_id:
            messages.error(request, 'Loại chứng từ 1 - Mua nhà cung cấp bắt buộc chọn Nhà cung cấp')
            return redirect('phieu_nhap_them')

        if loai_nhap == '2':
            if not phieu_doi_tra_id:
                messages.error(request, 'Loại chứng từ 2 - Khách hàng trả hàng bắt buộc chọn Phiếu đổi trả kế thừa')
                return redirect('phieu_nhap_them')
            phieu_doi_tra = PhieuTraHang.objects.select_related('khach_hang').prefetch_related('chi_tiet').filter(
                pk=phieu_doi_tra_id,
                trang_thai='1',
            ).first()
            if not phieu_doi_tra:
                messages.error(request, 'Phiếu đổi trả không hợp lệ hoặc đã hoàn tất')
                return redirect('phieu_nhap_them')

        so_phieu = data.get('so_phieu') or _gen_so_phieu('NK')
        ngay_lap = data.get('ngay_lap') or data.get('ngay_nhap') or data.get('ngay_chung_tu') or date.today()
        ngay_hach_toan = data.get('ngay_hach_toan') or data.get('ngay_chung_tu') or ngay_lap
        phieu = PhieuNhap.objects.create(
            so_phieu=so_phieu,
            ngay_lap=ngay_lap,
            ngay_hach_toan=ngay_hach_toan,
            ngay_chung_tu=ngay_hach_toan,
            ngay_nhap=ngay_lap,
            loai_nhap=loai_nhap,
            trang_thai='1',
            nha_cung_cap_id=ncc_id,
            so_hd_ncc=data.get('so_hd_ncc', ''),
            ngay_hd_ncc=data.get('ngay_hd_ncc') or None,
            kho_id=kho_id,
            ghi_chu=(data.get('ghi_chu', '') or '').strip(),
            nguoi_tao=request.user,
        )
        if phieu_doi_tra:
            link_tag = f'[DOI_TRA_PHIEU:{phieu_doi_tra.pk}]'
            phieu.ghi_chu = f"{(phieu.ghi_chu + ' ') if phieu.ghi_chu else ''}{link_tag}".strip()
            phieu.save(update_fields=['ghi_chu'])

        # Xử lý chi tiết
        hang_ids = data.getlist('hang_id[]')
        doi_tra_ct_ids = data.getlist('doi_tra_ct_id[]')
        sl_nhans = data.getlist('so_luong_nhan[]')
        don_gias = data.getlist('don_gia[]')
        tk_nos = data.getlist('tk_no[]')
        tk_cos = data.getlist('tk_co[]')
        so_dong_hop_le = 0
        doi_tra_so_luong_con_lai = {}
        doi_tra_ct_map = {}

        if phieu_doi_tra:
            for ct in phieu_doi_tra.chi_tiet.all():
                doi_tra_so_luong_con_lai[ct.id] = int(ct.so_luong or 0)
                doi_tra_ct_map[ct.id] = ct

        for i in range(len(hang_ids)):
            if hang_ids[i] and sl_nhans[i] and don_gias[i]:
                try:
                    sl_nhan = int(sl_nhans[i])
                    don_gia = _parse_money_input(don_gias[i])
                except (ValueError, TypeError):
                    continue

                if sl_nhan <= 0 or don_gia < 0:
                    continue

                if phieu_doi_tra:
                    ct_id_raw = (doi_tra_ct_ids[i] if i < len(doi_tra_ct_ids) else '').strip()
                    if not ct_id_raw or not ct_id_raw.isdigit():
                        phieu.delete()
                        messages.error(request, 'Dòng kế thừa phiếu đổi trả không hợp lệ')
                        return redirect('phieu_nhap_them')
                    ct_id = int(ct_id_raw)
                    if ct_id not in doi_tra_ct_map:
                        phieu.delete()
                        messages.error(request, 'Có dòng hàng không thuộc phiếu đổi trả đã chọn')
                        return redirect('phieu_nhap_them')
                    if int(hang_ids[i]) != int(doi_tra_ct_map[ct_id].hang_hoa_id):
                        phieu.delete()
                        messages.error(request, 'Sai mặt hàng kế thừa từ phiếu đổi trả')
                        return redirect('phieu_nhap_them')
                    if sl_nhan > doi_tra_so_luong_con_lai.get(ct_id, 0):
                        phieu.delete()
                        messages.error(request, 'Số lượng nhập vượt quá số lượng hàng trả')
                        return redirect('phieu_nhap_them')
                    doi_tra_so_luong_con_lai[ct_id] = doi_tra_so_luong_con_lai.get(ct_id, 0) - sl_nhan

                tk_no = tk_nos[i].strip() if i < len(tk_nos) else ''
                tk_co = tk_cos[i].strip() if i < len(tk_cos) else ''

                PhieuNhap_CT.objects.create(
                    phieu_nhap=phieu,
                    hang_hoa_id=hang_ids[i],
                    so_luong_nhan=sl_nhan,
                    don_gia=don_gia,
                    thue_vat=0,
                    tk_no=tk_no,
                    tk_co=tk_co,
                )
                so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            phieu.delete()
            messages.error(request, 'Phiếu nhập phải có ít nhất 1 dòng hàng hóa hợp lệ')
            return redirect('phieu_nhap_them')

        phieu.tinh_tong()

        thong_diep = [f'Đã tạo phiếu nhập {phieu.so_phieu} ở bước 1 - Lập phiếu']

        if phieu_doi_tra:
            phieu.trang_thai = trang_thai_mong_muon
            phieu.save(update_fields=['trang_thai'])
            if trang_thai_mong_muon == '2':
                thong_diep.append('Đã giữ trạng thái bước 2 - Sổ kho theo lựa chọn')
            elif trang_thai_mong_muon == '3':
                thong_diep.append('Đã giữ trạng thái bước 3 - Sổ cái theo lựa chọn')
        else:
            if trang_thai_mong_muon == '2':
                ok = phieu.xac_nhan_nhap_kho()
                if ok:
                    so_dong_phan_bo = _ghi_nhan_phan_bo_vi_tri(phieu)
                    thong_diep.append(f'Đã tự động chuyển bước 2 - Sổ kho ({so_dong_phan_bo} dòng vị trí)')
                else:
                    messages.error(request, 'Không thể tự động chuyển sang bước 2 - Sổ kho')
                    return redirect('phieu_nhap_detail', pk=phieu.pk)

                    if trang_thai_mong_muon == '3':
                        try:
                            # Tự động ghi bước 2 nếu chưa ghi
                            if phieu.trang_thai != '2':
                                phieu.xac_nhan_nhap_kho()
                            phieu.trang_thai = '3'
                            phieu.save(update_fields=['trang_thai'])
                            post_to_ledger('phieu_nhap', phieu.id, user=request.user)
                            thong_diep.append('Đã tự động chuyển bước 3 - Sổ cái')
                        except LedgerPostingError as exc:
                            messages.warning(request, f'Lỗi ghi sổ cái: {str(exc)}')
                    return redirect('phieu_nhap_detail', pk=phieu.pk)

            if trang_thai_mong_muon == '3':
                # BR12.x.5: Kiểm tra chênh lệch kiểm kê trước khi xác nhận nhập
                hang_ids = list(phieu.chi_tiet.values_list('hang_hoa_id', flat=True))
                conflicted = _check_kiem_ke_diff(phieu.kho_id, hang_ids)
                if conflicted:
                    items_str = ', '.join([f"{c['ma_hang']}" for c in conflicted])
                    messages.error(request, f'Không thể xác nhận: hàng hóa {items_str} đang có chênh lệch kiểm kê chưa được xử lý.')
                    return redirect('phieu_nhap_detail', pk=phieu.pk)

                ok = phieu.xac_nhan_nhap_kho()
                if not ok:
                    messages.error(request, 'Không thể tự động chuyển sang bước 2 - Sổ kho để chuyển sổ cái')
                    return redirect('phieu_nhap_detail', pk=phieu.pk)

        messages.success(request, '. '.join(thong_diep))
        return redirect('phieu_nhap_detail', pk=phieu.pk)

    copy_from = request.GET.get('copy_from', '').strip()
    copy_data = None
    if copy_from:
        source = get_object_or_404(
            PhieuNhap.objects.select_related('nha_cung_cap', 'kho').prefetch_related('chi_tiet__hang_hoa'),
            pk=copy_from,
        )
        copy_data = {
            'ngay_lap': source.ngay_lap.isoformat(),
            'ngay_hach_toan': source.ngay_hach_toan.isoformat(),
            'loai_nhap': source.loai_nhap,
            'kho_id': str(source.kho_id or ''),
            'ncc_id': str(source.nha_cung_cap_id or ''),
            'ncc_label': f'{source.nha_cung_cap.ma_ncc} - {source.nha_cung_cap.ten_ncc}' if source.nha_cung_cap else '',
            'so_hd_ncc': source.so_hd_ncc or '',
            'ngay_hd_ncc': source.ngay_hd_ncc.isoformat() if source.ngay_hd_ncc else '',
            'ghi_chu': source.ghi_chu or '',
            'rows': [
                {
                    'hang_id': str(ct.hang_hoa_id),
                    'hang_label': f'{ct.hang_hoa.ma_hang} - {ct.hang_hoa.ten_hang}',
                    'so_luong_nhan': int(ct.so_luong_nhan or 0),
                    'don_gia': float(ct.don_gia or 0),
                    'tk_no': ct.tk_no or '',
                    'tk_co': ct.tk_co or '',
                }
                for ct in source.chi_tiet.all()
            ],
        }

    ma_ncc_hop_le = KhachHang.objects.filter(trang_thai=True, la_nha_cung_cap=True).values_list('ma_kh', flat=True)
    context = {
        'kho_list': Kho.objects.filter(trang_thai=True),
        'ncc_list': NhaCungCap.objects.filter(trang_thai=True, ma_ncc__in=ma_ncc_hop_le).order_by('ma_ncc'),
        'hang_list': HangHoa.objects.all().order_by('ma_hang'),
        'tai_khoan_list': TaiKhoanKeToan.objects.filter(trang_thai=True).order_by('ma_tk'),
        'so_phieu_default': _gen_so_phieu('NK'),
        'today': date.today(),
        'copy_data': copy_data,
        'doi_tra_pending_payload': _build_doi_tra_pending_payload(),
        'page_title': 'Tạo phiếu nhập kho',
        'active_menu': 'phieu_nhap',
    }
    return render(request, 'kho/phieu_nhap_form.html', context)


@login_required
def phieu_nhap_detail(request, pk):
    phieu = get_object_or_404(PhieuNhap, pk=pk)
    chi_tiet = phieu.chi_tiet.select_related('hang_hoa')
    context = {
        'phieu': phieu,
        'chi_tiet': chi_tiet,
        'page_title': f'Phiếu nhập {phieu.so_phieu}',
        'active_menu': 'phieu_nhap',
    }
    return render(request, 'kho/phieu_nhap_detail.html', context)


@login_required
@transaction.atomic
def phieu_nhap_sua(request, pk):
    phieu = get_object_or_404(PhieuNhap.objects.select_related('nha_cung_cap', 'kho').prefetch_related('chi_tiet__hang_hoa'), pk=pk)

    if request.method == 'POST':
        ok, err_msg = _rollback_phieu_nhap_inventory(phieu)
        if not ok:
            messages.error(request, err_msg)
            return redirect('phieu_nhap_detail', pk=pk)

        data = request.POST
        kho_id = data.get('kho')
        loai_nhap = _normalize_loai_nhap(data.get('loai_nhap', phieu.loai_nhap or '1'))
        if loai_nhap not in ('1', '2'):
            loai_nhap = '2' if str(phieu.loai_nhap or '').strip() == '3' else (phieu.loai_nhap or '1')
        trang_thai_mong_muon = str(data.get('trang_thai', '1') or '1').strip()
        if trang_thai_mong_muon not in ('1', '2', '3'):
            trang_thai_mong_muon = '1'
        linked_doi_tra_id = _extract_doi_tra_id_from_ghi_chu(phieu.ghi_chu)
        posted_doi_tra_id = (data.get('phieu_doi_tra_id') or '').strip()
        if posted_doi_tra_id.isdigit():
            linked_doi_tra_id = int(posted_doi_tra_id)
        ncc_id = data.get('ncc') or None

        if not kho_id:
            messages.error(request, 'Vui lòng chọn kho nhận')
            return redirect('phieu_nhap_sua', pk=pk)
        if loai_nhap == '1' and not ncc_id:
            messages.error(request, 'Loại chứng từ 1 - Mua nhà cung cấp bắt buộc chọn Nhà cung cấp')
            return redirect('phieu_nhap_sua', pk=pk)

        phieu.ngay_lap = data.get('ngay_lap') or data.get('ngay_nhap') or data.get('ngay_chung_tu') or date.today()
        phieu.ngay_hach_toan = data.get('ngay_hach_toan') or data.get('ngay_chung_tu') or phieu.ngay_lap
        phieu.ngay_chung_tu = phieu.ngay_hach_toan
        phieu.ngay_nhap = phieu.ngay_lap
        phieu.loai_nhap = loai_nhap
        phieu.nha_cung_cap_id = ncc_id
        phieu.so_hd_ncc = data.get('so_hd_ncc', '')
        phieu.ngay_hd_ncc = data.get('ngay_hd_ncc') or None
        phieu.kho_id = kho_id
        phieu.ghi_chu = (data.get('ghi_chu', '') or '').strip()
        if linked_doi_tra_id:
            link_tag = f'[DOI_TRA_PHIEU:{linked_doi_tra_id}]'
            if link_tag not in phieu.ghi_chu:
                phieu.ghi_chu = f"{phieu.ghi_chu} {link_tag}".strip()
        phieu.trang_thai = '1'
        phieu.save(update_fields=['ngay_lap', 'ngay_hach_toan', 'ngay_chung_tu', 'ngay_nhap', 'loai_nhap', 'nha_cung_cap', 'so_hd_ncc', 'ngay_hd_ncc', 'kho', 'ghi_chu', 'trang_thai'])

        phieu.chi_tiet.all().delete()
        hang_ids = data.getlist('hang_id[]')
        sl_nhans = data.getlist('so_luong_nhan[]')
        don_gias = data.getlist('don_gia[]')
        tk_nos = data.getlist('tk_no[]')
        tk_cos = data.getlist('tk_co[]')
        so_dong_hop_le = 0

        for i in range(len(hang_ids)):
            if hang_ids[i] and sl_nhans[i] and don_gias[i]:
                try:
                    sl_nhan = int(sl_nhans[i])
                    don_gia = _parse_money_input(don_gias[i])
                except (ValueError, TypeError):
                    continue
                if sl_nhan <= 0 or don_gia < 0:
                    continue
                tk_no = tk_nos[i].strip() if i < len(tk_nos) else ''
                tk_co = tk_cos[i].strip() if i < len(tk_cos) else ''
                PhieuNhap_CT.objects.create(
                    phieu_nhap=phieu,
                    hang_hoa_id=hang_ids[i],
                    so_luong_nhan=sl_nhan,
                    don_gia=don_gia,
                    thue_vat=0,
                    tk_no=tk_no,
                    tk_co=tk_co,
                )
                so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            messages.error(request, 'Phiếu nhập phải có ít nhất 1 dòng hàng hóa hợp lệ')
            return redirect('phieu_nhap_sua', pk=pk)

        phieu.tinh_tong()
        if linked_doi_tra_id:
            phieu.trang_thai = trang_thai_mong_muon
            phieu.save(update_fields=['trang_thai'])
        else:
            if trang_thai_mong_muon == '2':
                # BR12.x.5: Kiểm tra chênh lệch kiểm kê trước khi xác nhận nhập
                hang_ids = list(phieu.chi_tiet.values_list('hang_hoa_id', flat=True))
                conflicted = _check_kiem_ke_diff(phieu.kho_id, hang_ids)
                if conflicted:
                    items_str = ', '.join([f"{c['ma_hang']}" for c in conflicted])
                    messages.error(request, f'Không thể xác nhận: hàng hóa {items_str} đang có chênh lệch kiểm kê chưa được xử lý. Hãy duyệt phiếu điều chỉnh kiểm kê trước.')
                    return redirect('phieu_nhap_sua', pk=pk)
                
                ok = phieu.xac_nhan_nhap_kho()
                if ok:
                    _ghi_nhan_phan_bo_vi_tri(phieu)
            if trang_thai_mong_muon == '3':
                hang_ids = list(phieu.chi_tiet.values_list('hang_hoa_id', flat=True))
                conflicted = _check_kiem_ke_diff(phieu.kho_id, hang_ids)
                if conflicted:
                    items_str = ', '.join([f"{c['ma_hang']}" for c in conflicted])
                    messages.error(request, f'Không thể xác nhận: hàng hóa {items_str} đang có chênh lệch kiểm kê chưa được xử lý.')
                    return redirect('phieu_nhap_sua', pk=pk)
                
                ok = phieu.xac_nhan_nhap_kho()
                if ok:
                    _ghi_nhan_phan_bo_vi_tri(phieu)
                    phieu.trang_thai = '3'
                    phieu.save(update_fields=['trang_thai'])

        messages.success(request, f'Đã cập nhật phiếu nhập {phieu.so_phieu}')
        return redirect('phieu_nhap_detail', pk=pk)

    copy_data = {
        'ngay_lap': phieu.ngay_lap.isoformat(),
        'ngay_hach_toan': phieu.ngay_hach_toan.isoformat(),
        'loai_nhap': '2' if str(phieu.loai_nhap or '').strip() == '3' else phieu.loai_nhap,
        'trang_thai': phieu.trang_thai,
        'kho_id': str(phieu.kho_id or ''),
        'ncc_id': str(phieu.nha_cung_cap_id or ''),
        'ncc_label': f'{phieu.nha_cung_cap.ma_ncc} - {phieu.nha_cung_cap.ten_ncc}' if phieu.nha_cung_cap else '',
        'so_hd_ncc': phieu.so_hd_ncc or '',
        'ngay_hd_ncc': phieu.ngay_hd_ncc.isoformat() if phieu.ngay_hd_ncc else '',
        'ghi_chu': phieu.ghi_chu or '',
        'phieu_doi_tra_id': str(_extract_doi_tra_id_from_ghi_chu(phieu.ghi_chu) or ''),
        'rows': [
            {
                'hang_id': str(ct.hang_hoa_id),
                'hang_label': f'{ct.hang_hoa.ma_hang} - {ct.hang_hoa.ten_hang}',
                'so_luong_nhan': int(ct.so_luong_nhan or 0),
                'don_gia': float(ct.don_gia or 0),
                'tk_no': ct.tk_no or '',
                'tk_co': ct.tk_co or '',
            }
            for ct in phieu.chi_tiet.all()
        ],
    }
    ma_ncc_hop_le = KhachHang.objects.filter(trang_thai=True, la_nha_cung_cap=True).values_list('ma_kh', flat=True)
    source_doi_tra_id = _extract_doi_tra_id_from_ghi_chu(phieu.ghi_chu)
    source_doi_tra_label = ''
    if source_doi_tra_id:
        dt_obj = PhieuTraHang.objects.filter(pk=source_doi_tra_id).only('so_phieu').first()
        source_doi_tra_label = dt_obj.so_phieu if dt_obj else f'DT{source_doi_tra_id}'
    return render(request, 'kho/phieu_nhap_form.html', {
        'kho_list': Kho.objects.filter(trang_thai=True),
        'ncc_list': NhaCungCap.objects.filter(trang_thai=True, ma_ncc__in=ma_ncc_hop_le).order_by('ma_ncc'),
        'hang_list': HangHoa.objects.all().order_by('ma_hang'),
        'tai_khoan_list': TaiKhoanKeToan.objects.filter(trang_thai=True).order_by('ma_tk'),
        'so_phieu_default': phieu.so_phieu,
        'today': phieu.ngay_lap,
        'copy_data': copy_data,
        'doi_tra_pending_payload': _build_doi_tra_pending_payload(source_doi_tra_id),
        'selected_doi_tra_id': str(source_doi_tra_id or ''),
        'selected_doi_tra_label': source_doi_tra_label,
        'editing_phieu': phieu,
        'page_title': f'Sửa phiếu nhập {phieu.so_phieu}',
        'active_menu': 'phieu_nhap',
    })


@login_required
@transaction.atomic
def phieu_nhap_xoa(request, pk):
    phieu = get_object_or_404(PhieuNhap, pk=pk)
    if request.method != 'POST':
        return redirect('phieu_nhap_list')

    ok, err_msg = _rollback_phieu_nhap_inventory(phieu)
    if not ok:
        messages.error(request, err_msg)
        return redirect('phieu_nhap_list')

    so_phieu = phieu.so_phieu
    phieu.delete()
    messages.success(request, f'Đã xóa phiếu nhập {so_phieu}')
    return redirect('phieu_nhap_list')


@login_required
@xframe_options_exempt
def phieu_nhap_in(request, pk):
    phieu = get_object_or_404(PhieuNhap.objects.select_related('nha_cung_cap', 'kho', 'nguoi_tao'), pk=pk)
    chi_tiet = list(phieu.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh').all())
    while len(chi_tiet) < 10:
        chi_tiet.append(None)
    return render(request, 'kho/phieu_nhap_print.html', {
        'phieu': phieu,
        'chi_tiet': chi_tiet,
        'page_title': f'In phiếu nhập {phieu.so_phieu}',
        'active_menu': 'phieu_nhap',
    })


@login_required
def phieu_nhap_xac_nhan(request, pk):
    phieu = get_object_or_404(PhieuNhap, pk=pk)
    if request.method == 'POST':
        if _extract_doi_tra_id_from_ghi_chu(phieu.ghi_chu):
            messages.error(request, 'Phiếu nhập kế thừa từ phiếu đổi trả không hỗ trợ ghi Sổ kho trực tiếp.')
            return redirect('phieu_nhap_detail', pk=pk)
        ok = phieu.xac_nhan_nhap_kho()
        if ok:
            so_dong_phan_bo = _ghi_nhan_phan_bo_vi_tri(phieu)
            messages.success(request, f'Đã ghi Sổ kho thành công: {phieu.so_phieu}. Đã gợi ý/ghi nhận {so_dong_phan_bo} dòng vị trí.')
        else:
            messages.error(request, 'Không thể xác nhận - phiếu đã được xử lý')
    return redirect('phieu_nhap_detail', pk=pk)


@login_required
def phieu_nhap_chuyen_so_cai(request, pk):
    phieu = get_object_or_404(PhieuNhap, pk=pk)
    if request.method == 'POST':
        if _extract_doi_tra_id_from_ghi_chu(phieu.ghi_chu):
            messages.error(request, 'Phiếu nhập kế thừa từ phiếu đổi trả không hỗ trợ chuyển Sổ cái trực tiếp.')
            return redirect('phieu_nhap_detail', pk=pk)
        if phieu.trang_thai == '2':
            phieu.trang_thai = '3'
            phieu.save(update_fields=['trang_thai'])
            try:
                post_to_ledger('phieu_nhap', phieu.id, user=request.user)
            except LedgerPostingError as exc:
                messages.warning(request, f'Lỗi ghi sổ cái: {str(exc)}')
            messages.success(request, f'Đã chuyển phiếu {phieu.so_phieu} sang bước Sổ cái')
        else:
            messages.error(request, 'Chỉ có thể chuyển Sổ cái sau khi đã ghi Sổ kho (bước 2)')
    return redirect('phieu_nhap_detail', pk=pk)


# ─── PHIẾU XUẤT KHO ────────────────────────────────────────
@login_required
def phieu_xuat_list(request):
    q = request.GET.get('q', '')
    kho_filter = request.GET.get('kho', '')
    trang_thai_filter = request.GET.get('trang_thai', '')
    items = PhieuXuat.objects.select_related('kho')
    if q:
        items = items.filter(so_phieu__icontains=q)
    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if trang_thai_filter in ('1', '2', '3'):
        items = items.filter(trang_thai=trang_thai_filter)
    context = {
        'items': items[:50],
        'q': q,
        'kho_list': Kho.objects.filter(trang_thai=True),
        'kho_filter': kho_filter,
        'trang_thai_filter': trang_thai_filter,
        'page_title': 'Phiếu xuất kho',
        'active_menu': 'phieu_xuat',
    }
    return render(request, 'kho/phieu_xuat_list.html', context)


@login_required
def phieu_xuat_them(request):
    if request.method == 'POST':
        data = request.POST
        kho_id = data.get('kho')
        trang_thai_mong_muon = str(data.get('trang_thai', '1') or '1').strip()
        if trang_thai_mong_muon not in ('1', '2', '3'):
            trang_thai_mong_muon = '1'
        ngay_lap = data.get('ngay_lap') or data.get('ngay_xuat') or data.get('ngay_chung_tu') or date.today()
        ngay_hach_toan = data.get('ngay_hach_toan') or data.get('ngay_chung_tu') or ngay_lap
        so_phieu_input = (data.get('so_phieu') or '').strip()
        so_phieu_candidate = so_phieu_input if so_phieu_input else _gen_so_phieu('PX')
        if PhieuXuat.objects.filter(so_phieu=so_phieu_candidate).exists():
            so_phieu_candidate = _gen_so_phieu('PX')

        phieu = None
        for _ in range(5):
            try:
                phieu = PhieuXuat.objects.create(
                    so_phieu=so_phieu_candidate,
                    ngay_lap=ngay_lap,
                    ngay_hach_toan=ngay_hach_toan,
                    ngay_chung_tu=ngay_hach_toan,
                    ngay_xuat=ngay_lap,
                    loai_xuat=data.get('loai_xuat', 'noi_bo'),
                    kho_id=kho_id,
                    trang_thai='1',
                    nguoi_tao=request.user,
                    ghi_chu=data.get('ghi_chu', ''),
                )
                break
            except IntegrityError:
                so_phieu_candidate = _gen_so_phieu('PX')

        if phieu is None:
            messages.error(request, 'Không thể tạo số phiếu mới, vui lòng thử lại.')
            return redirect('phieu_xuat_them')
        hang_ids = data.getlist('hang_id[]')
        so_luongs = data.getlist('so_luong[]')
        tk_nos = data.getlist('tk_no[]')
        tk_cos = data.getlist('tk_co[]')
        tong_gv = 0
        for i in range(len(hang_ids)):
            if hang_ids[i] and so_luongs[i]:
                h_id = int(hang_ids[i])
                sl = int(so_luongs[i])
                tk = TonKho.objects.filter(hang_hoa_id=h_id, kho_id=kho_id).first()
                gv = tk.gia_von_tb if tk else 0
                tgv = gv * sl
                tk_no = tk_nos[i].strip() if i < len(tk_nos) else ''
                tk_co = tk_cos[i].strip() if i < len(tk_cos) else ''
                PhieuXuat_CT.objects.create(
                    phieu_xuat=phieu, hang_hoa_id=h_id,
                    so_luong=sl, gia_von=gv, tong_gia_von=tgv,
                    tk_no=tk_no, tk_co=tk_co,
                )
                tong_gv += tgv
        phieu.tong_gia_von = tong_gv
        phieu.save(update_fields=['tong_gia_von'])

        thong_diep = [f'Đã tạo phiếu xuất {phieu.so_phieu} ở bước 1 - Lập phiếu']

        if trang_thai_mong_muon in ('2', '3'):
            ok, err_msg = _xac_nhan_phieu_xuat(phieu)
            if not ok:
                messages.error(request, err_msg)
                return redirect('phieu_xuat_detail', pk=phieu.pk)
            thong_diep.append('Đã tự động chuyển bước 2 - Sổ kho')

        if trang_thai_mong_muon == '3':
            phieu.trang_thai = '3'
            phieu.save(update_fields=['trang_thai'])
            thong_diep.append('Đã tự động chuyển bước 3 - Sổ cái')

        messages.success(request, '. '.join(thong_diep))
        return redirect('phieu_xuat_detail', pk=phieu.pk)

    copy_from = request.GET.get('copy_from', '').strip()
    copy_data = None
    if copy_from:
        source = get_object_or_404(
            PhieuXuat.objects.select_related('kho').prefetch_related('chi_tiet__hang_hoa'),
            pk=copy_from,
        )
        copy_data = {
            'ngay_lap': source.ngay_lap.isoformat(),
            'ngay_hach_toan': source.ngay_hach_toan.isoformat(),
            'loai_xuat': source.loai_xuat,
            'kho_id': str(source.kho_id or ''),
            'trang_thai': source.trang_thai,
            'ghi_chu': source.ghi_chu or '',
            'rows': [
                {
                    'hang_id': str(ct.hang_hoa_id),
                    'hang_label': f'{ct.hang_hoa.ma_hang} - {ct.hang_hoa.ten_hang}',
                    'so_luong': int(ct.so_luong or 0),
                    'tk_no': ct.tk_no or '',
                    'tk_co': ct.tk_co or '',
                }
                for ct in source.chi_tiet.all()
            ],
        }

    context = {
        'kho_list': Kho.objects.filter(trang_thai=True),
        'hang_list': HangHoa.objects.filter(
            pk__in=TonKho.objects.filter(so_luong__gte=1).values_list('hang_hoa_id', flat=True).distinct()
        ).order_by('ma_hang'),
        'so_phieu_default': _gen_so_phieu('PX'),
        'today': date.today(),
        'copy_data': copy_data,
        'page_title': 'Tạo phiếu xuất kho',
        'active_menu': 'phieu_xuat',
    }
    return render(request, 'kho/phieu_xuat_form.html', context)


@login_required
def phieu_xuat_detail(request, pk):
    phieu = get_object_or_404(PhieuXuat, pk=pk)
    chi_tiet = phieu.chi_tiet.select_related('hang_hoa')
    context = {
        'phieu': phieu,
        'chi_tiet': chi_tiet,
        'page_title': f'Phiếu xuất {phieu.so_phieu}',
        'active_menu': 'phieu_xuat',
    }
    return render(request, 'kho/phieu_xuat_detail.html', context)


@login_required
@transaction.atomic
def phieu_xuat_sua(request, pk):
    phieu = get_object_or_404(PhieuXuat.objects.select_related('kho').prefetch_related('chi_tiet__hang_hoa'), pk=pk)

    if request.method == 'POST':
        if _phieu_xuat_has_related_documents(phieu):
            messages.error(request, f'Phiếu {phieu.so_phieu} đã phát sinh chứng từ bán hàng liên quan, không thể sửa/xóa.')
            return redirect('phieu_xuat_detail', pk=pk)

        ok, err_msg = _rollback_phieu_xuat_inventory(phieu)
        if not ok:
            messages.error(request, err_msg)
            return redirect('phieu_xuat_detail', pk=pk)

        data = request.POST
        trang_thai_mong_muon = str(data.get('trang_thai', '1') or '1').strip()
        if trang_thai_mong_muon not in ('1', '2', '3'):
            trang_thai_mong_muon = '1'
        phieu.ngay_lap = data.get('ngay_lap') or data.get('ngay_xuat') or data.get('ngay_chung_tu') or date.today()
        phieu.ngay_hach_toan = data.get('ngay_hach_toan') or data.get('ngay_chung_tu') or phieu.ngay_lap
        phieu.ngay_chung_tu = phieu.ngay_hach_toan
        phieu.ngay_xuat = phieu.ngay_lap
        phieu.loai_xuat = data.get('loai_xuat', 'noi_bo')
        phieu.kho_id = data.get('kho')
        phieu.ghi_chu = data.get('ghi_chu', '')
        phieu.trang_thai = '1'
        phieu.save(update_fields=['ngay_lap', 'ngay_hach_toan', 'ngay_chung_tu', 'ngay_xuat', 'loai_xuat', 'kho', 'ghi_chu', 'trang_thai'])

        phieu.chi_tiet.all().delete()
        hang_ids = data.getlist('hang_id[]')
        so_luongs = data.getlist('so_luong[]')
        tk_nos = data.getlist('tk_no[]')
        tk_cos = data.getlist('tk_co[]')
        tong_gv = 0
        so_dong_hop_le = 0
        for i in range(len(hang_ids)):
            if hang_ids[i] and so_luongs[i]:
                try:
                    h_id = int(hang_ids[i])
                    sl = int(so_luongs[i])
                except (ValueError, TypeError):
                    continue
                if sl <= 0:
                    continue
                tk = TonKho.objects.filter(hang_hoa_id=h_id, kho_id=phieu.kho_id).first()
                gv = tk.gia_von_tb if tk else 0
                tgv = gv * sl
                tk_no = tk_nos[i].strip() if i < len(tk_nos) else ''
                tk_co = tk_cos[i].strip() if i < len(tk_cos) else ''
                PhieuXuat_CT.objects.create(
                    phieu_xuat=phieu,
                    hang_hoa_id=h_id,
                    so_luong=sl,
                    gia_von=gv,
                    tong_gia_von=tgv,
                    tk_no=tk_no,
                    tk_co=tk_co,
                )
                tong_gv += tgv
                so_dong_hop_le += 1

        if so_dong_hop_le == 0:
            messages.error(request, 'Phiếu xuất phải có ít nhất 1 dòng hàng hóa hợp lệ')
            return redirect('phieu_xuat_sua', pk=pk)

        phieu.tong_gia_von = tong_gv
        phieu.save(update_fields=['tong_gia_von'])

        thong_diep = [f'Đã cập nhật phiếu xuất {phieu.so_phieu} ở bước 1 - Lập phiếu']

        if trang_thai_mong_muon in ('2', '3'):
            ok, err_msg = _xac_nhan_phieu_xuat(phieu)
            if not ok:
                messages.error(request, err_msg)
                return redirect('phieu_xuat_detail', pk=pk)
            thong_diep.append('Đã tự động chuyển bước 2 - Sổ kho')

        if trang_thai_mong_muon == '3':
            phieu.trang_thai = '3'
            phieu.save(update_fields=['trang_thai'])
            thong_diep.append('Đã tự động chuyển bước 3 - Sổ cái')

        messages.success(request, '. '.join(thong_diep))
        return redirect('phieu_xuat_detail', pk=pk)

    copy_data = {
        'ngay_lap': phieu.ngay_lap.isoformat(),
        'ngay_hach_toan': phieu.ngay_hach_toan.isoformat(),
        'loai_xuat': phieu.loai_xuat,
        'kho_id': str(phieu.kho_id or ''),
        'trang_thai': phieu.trang_thai,
        'ghi_chu': phieu.ghi_chu or '',
        'rows': [
            {
                'hang_id': str(ct.hang_hoa_id),
                'hang_label': f'{ct.hang_hoa.ma_hang} - {ct.hang_hoa.ten_hang}',
                'so_luong': int(ct.so_luong or 0),
                'tk_no': ct.tk_no or '',
                'tk_co': ct.tk_co or '',
            }
            for ct in phieu.chi_tiet.all()
        ],
    }
    return render(request, 'kho/phieu_xuat_form.html', {
        'kho_list': Kho.objects.filter(trang_thai=True),
        'hang_list': HangHoa.objects.filter(
            pk__in=TonKho.objects.filter(so_luong__gte=1).values_list('hang_hoa_id', flat=True).distinct()
        ).order_by('ma_hang'),
        'so_phieu_default': phieu.so_phieu,
        'today': phieu.ngay_lap,
        'copy_data': copy_data,
        'editing_phieu': phieu,
        'page_title': f'Sửa phiếu xuất {phieu.so_phieu}',
        'active_menu': 'phieu_xuat',
    })


@login_required
@transaction.atomic
def phieu_xuat_xoa(request, pk):
    phieu = get_object_or_404(PhieuXuat, pk=pk)
    if request.method != 'POST':
        return redirect('phieu_xuat_list')

    if _phieu_xuat_has_related_documents(phieu):
        messages.error(request, f'Phiếu {phieu.so_phieu} đã phát sinh chứng từ bán hàng liên quan, không thể sửa/xóa.')
        return redirect('phieu_xuat_list')

    ok, err_msg = _rollback_phieu_xuat_inventory(phieu)
    if not ok:
        messages.error(request, err_msg)
        return redirect('phieu_xuat_list')

    so_phieu = phieu.so_phieu
    phieu.delete()
    messages.success(request, f'Đã xóa phiếu xuất {so_phieu}')
    return redirect('phieu_xuat_list')


@login_required
@transaction.atomic
def phieu_xuat_xac_nhan(request, pk):
    phieu = get_object_or_404(PhieuXuat, pk=pk)
    if request.method == 'POST':
        ok, err_msg = _xac_nhan_phieu_xuat(phieu)
        if ok:
            messages.success(request, f'Đã ghi Sổ kho thành công: {phieu.so_phieu}')
        else:
            messages.error(request, err_msg)

    return redirect('phieu_xuat_detail', pk=pk)


@login_required
def phieu_xuat_chuyen_so_cai(request, pk):
    phieu = get_object_or_404(PhieuXuat, pk=pk)
    if request.method == 'POST':
        if phieu.trang_thai == '2':
            phieu.trang_thai = '3'
            phieu.save(update_fields=['trang_thai'])
            try:
                post_to_ledger('phieu_xuat', phieu.id, user=request.user)
            except LedgerPostingError as exc:
                messages.warning(request, f'Lỗi ghi sổ cái: {str(exc)}')
            messages.success(request, f'Đã chuyển phiếu {phieu.so_phieu} sang bước Sổ cái')
        else:
            messages.error(request, 'Chỉ có thể chuyển Sổ cái sau khi đã ghi Sổ kho (bước 2)')
    return redirect('phieu_xuat_detail', pk=pk)


# ─── KIỂM KÊ ────────────────────────────────────────────────
# ─── XUẤT/NHẬP KHO (EXCEL) ─────────────────────────────────
@login_required
def phieu_xuat_export_data(request):
    """Export phiếu xuất hiện tại sang Excel"""
    items = PhieuXuat.objects.select_related('kho').all()
    
    kho_filter = request.GET.get('kho', '').strip()
    trang_thai_filter = request.GET.get('trang_thai', '').strip()
    
    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if trang_thai_filter:
        items = items.filter(trang_thai=trang_thai_filter)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Phieu Xuat'
    ws.append(['Số phiếu', 'Ngày xuất', 'Kho', 'Loại xuất', 'Số lượng', 'Tổng giá vốn', 'Trạng thái'])
    
    for item in items:
        ws.append([
            item.so_phieu,
            item.ngay_chung_tu.strftime('%d/%m/%Y') if item.ngay_chung_tu else '',
            item.kho.ten_kho if item.kho else '',
            item.loai_xuat,
            item.chi_tiet.count(),
            int(item.tong_gia_von or 0),
            item.get_trang_thai_display() if hasattr(item, 'get_trang_thai_display') else item.trang_thai,
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="phieu_xuat.xlsx"'
    wb.save(response)
    return response


@login_required
def phieu_xuat_export_template(request):
    """Xuất template phiếu xuất"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Phieu Xuat'
    ws.append(['Ngày xuất', 'Mã kho', 'Mã hàng', 'Số lượng', 'Ghi chú'])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_phieu_xuat.xlsx"'
    wb.save(response)
    return response


@login_required
def phieu_xuat_import_excel(request):
    """Import phiếu xuất từ Excel"""
    if request.method != 'POST':
        messages.error(request, 'Phương thức không được hỗ trợ')
        return redirect('phieu_xuat_list')
    
    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Vui lòng chọn file Excel')
        return redirect('phieu_xuat_list')
    
    messages.success(request, 'Chức năng import phiếu xuất sẽ được phát triển. Hiện tại vui lòng tạo phiếu thủ công.')
    return redirect('phieu_xuat_list')


@login_required
def phieu_nhap_export_data(request):
    """Export phiếu nhập hiện tại sang Excel"""
    items = PhieuNhap.objects.select_related('kho', 'nha_cung_cap').all()
    
    kho_filter = request.GET.get('kho', '').strip()
    trang_thai_filter = request.GET.get('trang_thai', '').strip()
    
    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if trang_thai_filter:
        items = items.filter(trang_thai=trang_thai_filter)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Phieu Nhap'
    ws.append(['Số phiếu', 'Ngày nhập', 'Kho', 'Loại nhập', 'Nhà cung cấp', 'Số lượng', 'Tổng tiền', 'Trạng thái'])
    
    for item in items:
        ws.append([
            item.so_phieu,
            item.ngay_chung_tu.strftime('%d/%m/%Y') if item.ngay_chung_tu else '',
            item.kho.ten_kho if item.kho else '',
            item.loai_nhap,
            item.nha_cung_cap.ten_ncc if item.nha_cung_cap else '',
            item.chi_tiet.count(),
            int(item.tong_tien or 0),
            item.get_trang_thai_display() if hasattr(item, 'get_trang_thai_display') else item.trang_thai,
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="phieu_nhap.xlsx"'
    wb.save(response)
    return response


@login_required
def phieu_nhap_export_template(request):
    """Xuất template phiếu nhập"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Phieu Nhap'
    ws.append(['Ngày nhập', 'Mã kho', 'Mã NCC', 'Mã hàng', 'Số lượng', 'Đơn giá', 'Ghi chú'])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_phieu_nhap.xlsx"'
    wb.save(response)
    return response


@login_required
def phieu_nhap_import_excel(request):
    """Import phiếu nhập từ Excel"""
    if request.method != 'POST':
        messages.error(request, 'Phương thức không được hỗ trợ')
        return redirect('phieu_nhap_list')
    
    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Vui lòng chọn file Excel')
        return redirect('phieu_nhap_list')
    
    messages.success(request, 'Chức năng import phiếu nhập sẽ được phát triển. Hiện tại vui lòng tạo phiếu thủ công.')
    return redirect('phieu_nhap_list')


# ─── KIỂM KÊ ────────────────────────────────────────────────
def _can_manage_kiem_ke(user):
    return (
        user.is_superuser
        or user.is_staff
        or user.has_perm('kho.add_kiemke')
        or user.has_perm('kho.change_kiemke')
        or user.has_perm('kho.view_kiemke')
    )


def _gen_so_phieu_kiem_ke():
    prefix = 'KK'
    max_index = 0
    for code in KiemKe.objects.filter(ma_phieu__istartswith=prefix).values_list('ma_phieu', flat=True):
        text = str(code or '').strip().upper()
        if not text.startswith(prefix):
            continue
        suffix = text[len(prefix):]
        if suffix.isdigit():
            max_index = max(max_index, int(suffix))
    return f'{prefix}{max_index + 1:05d}'


def _gen_so_phieu_dieu_chinh_kiem_ke():
    prefix = 'DC'
    max_index = 0
    for code in PhieuDieuChinhKiemKe.objects.filter(so_phieu__istartswith=prefix).values_list('so_phieu', flat=True):
        text = str(code or '').strip().upper()
        if not text.startswith(prefix):
            continue
        suffix = text[len(prefix):]
        if suffix.isdigit():
            max_index = max(max_index, int(suffix))
    return f'{prefix}{max_index + 1:05d}'


def _kiem_ke_form_context(form_data=None):
    return {
        'kho_list': Kho.objects.filter(trang_thai=True),
        'vi_tri_list': ViTriKho.objects.filter(trang_thai='hoat_dong').select_related('kho').order_by('kho__ma_kho', 'ma_vi_tri'),
        'nhom_hang_list': NhomHang.objects.order_by('ten_nhom'),
        'nhan_vien_list': KhachHang.objects.filter(trang_thai=True, la_nhan_vien=True).order_by('ma_kh'),
        'today': date.today(),
        'ma_phieu_goi_y': _gen_so_phieu_kiem_ke(),
        'page_title': 'Tạo phiếu kiểm kê',
        'active_menu': 'kiem_ke',
        'form_data': form_data or {},
    }


def _split_lookup_text(text):
    raw = (text or '').strip()
    if not raw:
        return '', ''
    if ' - ' in raw:
        left, right = raw.split(' - ', 1)
        return left.strip(), right.strip()
    return raw, raw


def _find_kiem_ke_vi_tri(kk):
    ma_vi_tri = (kk.khu_vuc or '').strip()
    if not ma_vi_tri:
        return None
    return ViTriKho.objects.filter(
        kho_id=kk.kho_id,
        ma_vi_tri=ma_vi_tri,
        trang_thai='hoat_dong',
    ).first()


def _move_kiem_ke_hang_loi(kho_id, source_vi_tri_id, dest_vi_tri_id, hang_hoa_id, so_luong_loi):
    so_luong_loi = int(so_luong_loi or 0)
    if so_luong_loi <= 0:
        return
    if not dest_vi_tri_id:
        raise ValueError('Vui lòng chọn vị trí hàng lỗi trước khi duyệt phiếu kiểm kê có hàng lỗi.')
    if int(source_vi_tri_id) == int(dest_vi_tri_id):
        raise ValueError('Vị trí hàng lỗi phải khác vị trí kiểm kê.')

    source_row, _ = TonKhoViTri.objects.select_for_update().get_or_create(
        kho_id=kho_id,
        vi_tri_id=source_vi_tri_id,
        hang_hoa_id=hang_hoa_id,
        defaults={'so_luong': 0},
    )
    dest_row, _ = TonKhoViTri.objects.select_for_update().get_or_create(
        kho_id=kho_id,
        vi_tri_id=dest_vi_tri_id,
        hang_hoa_id=hang_hoa_id,
        defaults={'so_luong': 0},
    )

    current_source = int(source_row.so_luong or 0)
    if current_source < so_luong_loi:
        raise ValueError(f'Không đủ số lượng tại vị trí kiểm kê để chuyển hàng lỗi cho mã {source_row.hang_hoa.ma_hang}.')

    source_row.so_luong = current_source - so_luong_loi
    dest_row.so_luong = int(dest_row.so_luong or 0) + so_luong_loi
    source_row.save(update_fields=['so_luong', 'ngay_cap_nhat'])
    dest_row.save(update_fields=['so_luong', 'ngay_cap_nhat'])

    _sync_tonkho_from_vitri(kho_id, hang_hoa_id)


def _sync_tonkho_from_vitri(kho_id, hang_hoa_id):
    """Cập nhật so_luong từ tổng vị trị bình thường, so_luong_loi từ vị trị hàng lỗi"""
    # Tính tổng từ vị trị bình thường (loại trừ hàng lỗi)
    tong_vi_tri_binh_thuong = (
        TonKhoViTri.objects
        .filter(kho_id=kho_id, hang_hoa_id=hang_hoa_id)
        .exclude(vi_tri__ma_vi_tri__startswith='HL-')
        .aggregate(total=Sum('so_luong'))
        .get('total')
        or 0
    )
    # Tính tổng từ vị trị hàng lỗi (HL-xxx)
    tong_hang_loi = (
        TonKhoViTri.objects
        .filter(kho_id=kho_id, hang_hoa_id=hang_hoa_id)
        .filter(vi_tri__ma_vi_tri__startswith='HL-')
        .aggregate(total=Sum('so_luong'))
        .get('total')
        or 0
    )
    
    ton, _ = TonKho.objects.get_or_create(
        kho_id=kho_id,
        hang_hoa_id=hang_hoa_id,
        defaults={'so_luong': 0, 'so_luong_loi': 0, 'gia_von_tb': 0},
    )
    ton.so_luong = int(tong_vi_tri_binh_thuong)
    ton.so_luong_loi = int(tong_hang_loi)
    if ton.so_luong < 0:
        ton.so_luong = 0
    if ton.so_luong_loi < 0:
        ton.so_luong_loi = 0
    ton.save(update_fields=['so_luong', 'so_luong_loi', 'ngay_cap_nhat'])


def _resync_tonkho_from_vitri_scope(kho_id=None):
    qs = TonKhoViTri.objects.values_list('kho_id', 'hang_hoa_id').distinct()
    if kho_id:
        qs = qs.filter(kho_id=kho_id)
    for current_kho_id, hang_hoa_id in qs.iterator():
        _sync_tonkho_from_vitri(current_kho_id, hang_hoa_id)


@login_required
def kiem_ke_list(request):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền truy cập chức năng kiểm kê.')
        return redirect('dashboard')

    q = request.GET.get('q', '').strip()
    ma_phieu = request.GET.get('ma_phieu', '').strip()
    kho_filter = request.GET.get('kho', '').strip()
    ngay_filter = request.GET.get('ngay', '').strip()
    trang_thai_filter = request.GET.get('trang_thai', '').strip()

    items = KiemKe.objects.select_related('kho', 'nhom_hang').order_by('-ngay_tao', '-id')

    if ma_phieu:
        items = items.filter(ma_phieu__icontains=ma_phieu)
    if q:
        items = items.filter(
            Q(ghi_chu__icontains=q)
            | Q(nguoi_kiem__icontains=q)
            | Q(kho__ten_kho__icontains=q)
            | Q(kho__ma_kho__icontains=q)
            | Q(khu_vuc__icontains=q)
        )
    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if ngay_filter:
        items = items.filter(ngay_kiem_ke=ngay_filter)
    if trang_thai_filter:
        items = items.filter(trang_thai=trang_thai_filter)

    items = items[:20]
    return render(request, 'kho/kiem_ke_list.html', {
        'items': items,
        'q': q,
        'ma_phieu': ma_phieu,
        'kho_list': Kho.objects.filter(trang_thai=True),
        'kho_filter': kho_filter,
        'ngay_filter': ngay_filter,
        'trang_thai_filter': trang_thai_filter,
        'page_title': 'Phiếu kiểm kê',
        'active_menu': 'kiem_ke',
    })


@login_required
def kiem_ke_export_data(request):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền kết xuất phiếu kiểm kê.')
        return redirect('kiem_ke_list')

    q = request.GET.get('q', '').strip()
    ma_phieu = request.GET.get('ma_phieu', '').strip()
    kho_filter = request.GET.get('kho', '').strip()
    ngay_filter = request.GET.get('ngay', '').strip()
    trang_thai_filter = request.GET.get('trang_thai', '').strip()

    items = KiemKe.objects.select_related('kho', 'nhom_hang').order_by('-ngay_tao', '-id')

    if ma_phieu:
        items = items.filter(ma_phieu__icontains=ma_phieu)
    if q:
        items = items.filter(
            Q(ghi_chu__icontains=q)
            | Q(nguoi_kiem__icontains=q)
            | Q(kho__ma_kho__icontains=q)
            | Q(khu_vuc__icontains=q)
        )
    if kho_filter:
        items = items.filter(kho_id=kho_filter)
    if ngay_filter:
        items = items.filter(ngay_kiem_ke=ngay_filter)
    if trang_thai_filter:
        items = items.filter(trang_thai=trang_thai_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = 'KiemKe'
    ws.append(['Mã phiếu', 'Ngày kiểm kê', 'Mã kho', 'Khu vực', 'Loại hàng', 'Người kiểm kê', 'Trạng thái', 'Ghi chú'])
    for item in items:
        ws.append([
            item.ma_phieu,
            item.ngay_kiem_ke.isoformat() if item.ngay_kiem_ke else '',
            item.kho.ma_kho if item.kho_id else '',
            item.khu_vuc or '',
            item.nhom_hang.ten_nhom if item.nhom_hang_id else 'Tất cả',
            item.nguoi_kiem or '',
            item.get_trang_thai_display(),
            item.ghi_chu or '',
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="phieu_kiem_ke.xlsx"'
    return response


@login_required
def kiem_ke_them(request):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền lập phiếu kiểm kê.')
        return redirect('kiem_ke_list')

    if request.method == 'POST':
        kho_id = (request.POST.get('kho') or '').strip()
        kho_display = (request.POST.get('kho_display') or '').strip()
        ngay = request.POST.get('ngay_kiem_ke') or date.today()
        khu_vuc_id = (request.POST.get('khu_vuc') or '').strip()
        khu_vuc_display = (request.POST.get('khu_vuc_display') or '').strip()
        vi_tri_hang_loi_id = (request.POST.get('vi_tri_hang_loi') or '').strip()
        vi_tri_hang_loi_display = (request.POST.get('vi_tri_hang_loi_display') or '').strip()
        nguoi_kiem_id = (request.POST.get('nguoi_kiem') or '').strip()
        nguoi_kiem_display = (request.POST.get('nguoi_kiem_display') or '').strip()
        ghi_chu = (request.POST.get('ghi_chu') or '').strip()
        nhom_hang_id = (request.POST.get('nhom_hang') or '').strip()
        nhom_hang_display = (request.POST.get('nhom_hang_display') or '').strip()

        form_data = {
            'ngay_kiem_ke': str(ngay),
            'kho': kho_id,
            'kho_display': kho_display,
            'khu_vuc': khu_vuc_id,
            'khu_vuc_display': khu_vuc_display,
            'vi_tri_hang_loi': vi_tri_hang_loi_id,
            'vi_tri_hang_loi_display': vi_tri_hang_loi_display,
            'nguoi_kiem': nguoi_kiem_id,
            'nguoi_kiem_display': nguoi_kiem_display,
            'nhom_hang': nhom_hang_id,
            'nhom_hang_display': nhom_hang_display,
            'ghi_chu': ghi_chu,
        }

        # Fallback when hidden IDs are empty: resolve by typed/display text.
        if not kho_id and kho_display:
            code, name = _split_lookup_text(kho_display)
            kho_obj = Kho.objects.filter(trang_thai=True).filter(Q(ma_kho__iexact=code) | Q(ten_kho__iexact=name)).first()
            if not kho_obj:
                kho_obj = Kho.objects.filter(trang_thai=True).filter(Q(ma_kho__icontains=code) | Q(ten_kho__icontains=name)).first()
            if kho_obj:
                kho_id = str(kho_obj.pk)
                form_data['kho'] = kho_id

        if not nguoi_kiem_id and nguoi_kiem_display:
            code, name = _split_lookup_text(nguoi_kiem_display)
            nv_obj = KhachHang.objects.filter(trang_thai=True, la_nhan_vien=True).filter(Q(ma_kh__iexact=code) | Q(ten_kh__iexact=name)).first()
            if not nv_obj:
                nv_obj = KhachHang.objects.filter(trang_thai=True, la_nhan_vien=True).filter(Q(ma_kh__icontains=code) | Q(ten_kh__icontains=name)).first()
            if nv_obj:
                nguoi_kiem_id = str(nv_obj.pk)
                form_data['nguoi_kiem'] = nguoi_kiem_id

        if nhom_hang_display and not nhom_hang_id:
            code, name = _split_lookup_text(nhom_hang_display)
            nhom_obj = NhomHang.objects.filter(Q(ma_nhom__iexact=code) | Q(ten_nhom__iexact=name)).first()
            if not nhom_obj:
                nhom_obj = NhomHang.objects.filter(Q(ma_nhom__icontains=code) | Q(ten_nhom__icontains=name)).first()
            if nhom_obj:
                nhom_hang_id = str(nhom_obj.pk)
                form_data['nhom_hang'] = nhom_hang_id

        if not kho_id or not nguoi_kiem_id:
            messages.error(request, 'Vui lòng chọn đủ thông tin bắt buộc: Kho, Khu vực kiểm kê, Người kiểm kê.')
            return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data))

        kho = Kho.objects.filter(pk=kho_id, trang_thai=True).first()
        if not kho:
            messages.error(request, 'Kho không hợp lệ hoặc đã ngừng hoạt động.')
            return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data))

        if not khu_vuc_id and khu_vuc_display:
            code, _name = _split_lookup_text(khu_vuc_display)
            vi_tri_obj = ViTriKho.objects.filter(kho_id=kho_id, trang_thai='hoat_dong').filter(Q(ma_vi_tri__iexact=code) | Q(ma_vi_tri__icontains=code)).first()
            if vi_tri_obj:
                khu_vuc_id = str(vi_tri_obj.pk)
                form_data['khu_vuc'] = khu_vuc_id

        if not vi_tri_hang_loi_id and vi_tri_hang_loi_display:
            code, _name = _split_lookup_text(vi_tri_hang_loi_display)
            vi_tri_loi_obj = ViTriKho.objects.filter(kho_id=kho_id, trang_thai='hoat_dong').filter(Q(ma_vi_tri__iexact=code) | Q(ma_vi_tri__icontains=code)).first()
            if vi_tri_loi_obj:
                vi_tri_hang_loi_id = str(vi_tri_loi_obj.pk)
                form_data['vi_tri_hang_loi'] = vi_tri_hang_loi_id

        vi_tri = ViTriKho.objects.filter(pk=khu_vuc_id, kho_id=kho_id, trang_thai='hoat_dong').first()
        if not vi_tri:
            messages.error(request, 'Khu vực kiểm kê không hợp lệ cho kho đã chọn.')
            return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data))

        vi_tri_hang_loi = None
        if vi_tri_hang_loi_id:
            vi_tri_hang_loi = ViTriKho.objects.filter(pk=vi_tri_hang_loi_id, kho_id=kho_id, trang_thai='hoat_dong').first()
            if not vi_tri_hang_loi:
                messages.error(request, 'Vị trí hàng lỗi không hợp lệ cho kho đã chọn.')
                return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data))

        nhan_vien = KhachHang.objects.filter(pk=nguoi_kiem_id, trang_thai=True, la_nhan_vien=True).first()
        if not nhan_vien:
            messages.error(request, 'Người kiểm kê phải được chọn từ danh mục Khách hàng có đánh dấu Nhân viên.')
            return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data))

        nguoi_kiem = f'{nhan_vien.ma_kh} - {nhan_vien.ten_kh}'
        khu_vuc = vi_tri.ma_vi_tri

        ton_qs = TonKhoViTri.objects.select_related('hang_hoa').filter(kho_id=kho_id, vi_tri_id=vi_tri.pk)
        if nhom_hang_id and nhom_hang_id.isdigit():
            ton_qs = ton_qs.filter(hang_hoa__nhom_hang_id=nhom_hang_id)

        if not ton_qs.exists():
            messages.error(request, 'Không có hàng hóa thuộc phạm vi kiểm kê đã chọn.')
            return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data))

        with transaction.atomic():
            kk = KiemKe.objects.create(
                ma_phieu=_gen_so_phieu_kiem_ke(),
                ngay_kiem_ke=ngay,
                kho_id=kho_id,
                nhom_hang_id=nhom_hang_id if nhom_hang_id.isdigit() else None,
                khu_vuc=khu_vuc,
                vi_tri_hang_loi=vi_tri_hang_loi,
                nguoi_kiem=nguoi_kiem,
                ghi_chu=ghi_chu,
                trang_thai='1',
            )

            for tk in ton_qs:
                KiemKe_CT.objects.create(
                    kiem_ke=kk,
                    hang_hoa=tk.hang_hoa,
                    so_luong_so_sach=int(tk.so_luong or 0),
                    so_luong_thuc_te=int(tk.so_luong or 0),
                    tinh_trang='tot_100',
                )

        messages.success(request, 'Lập phiếu kiểm kê thành công. Phiếu đang ở trạng thái Chờ kiểm kê.')
        return redirect('kiem_ke_detail', pk=kk.pk)

    # Xu ly copy_from: neu co tham so copy_from tren URL, lay du lieu phieu cu de dien vao form
    copy_from = request.GET.get('copy_from', '').strip()
    form_data_copy = {}
    if copy_from and copy_from.isdigit():
        kk_copy = KiemKe.objects.select_related('kho', 'nhom_hang', 'vi_tri_hang_loi').filter(pk=copy_from).first()
        if kk_copy:
            nv_obj = None
            nguoi_kiem_display = kk_copy.nguoi_kiem or ''
            # nguoi_kiem luu dang "MA - TEN", parse lai de tim ID
            if nguoi_kiem_display:
                code, _name = _split_lookup_text(nguoi_kiem_display)
                nv_obj = KhachHang.objects.filter(ma_kh=code, la_nhan_vien=True).first()
            form_data_copy = {
                'kho': str(kk_copy.kho_id or ''),
                'kho_display': f'{kk_copy.kho.ma_kho} - {kk_copy.kho.ten_kho}' if kk_copy.kho else '',
                'khu_vuc': '',  # vi_tri se duoc set lai, don gian lay ma_phieu
                'khu_vuc_display': kk_copy.khu_vuc or '',
                'nguoi_kiem': str(nv_obj.pk) if nv_obj else '',
                'nguoi_kiem_display': nguoi_kiem_display,
                'nhom_hang': str(kk_copy.nhom_hang_id or ''),
                'nhom_hang_display': f'{kk_copy.nhom_hang.ma_nhom} - {kk_copy.nhom_hang.ten_nhom}' if kk_copy.nhom_hang else '',
                'vi_tri_hang_loi': str(kk_copy.vi_tri_hang_loi_id or ''),
                'vi_tri_hang_loi_display': kk_copy.vi_tri_hang_loi.ma_vi_tri if kk_copy.vi_tri_hang_loi else '',
                'ghi_chu': kk_copy.ghi_chu or '',
                'ngay_kiem_ke': date.today().isoformat(),
            }

    return render(request, 'kho/kiem_ke_form.html', _kiem_ke_form_context(form_data_copy if form_data_copy else None))


@login_required
def kiem_ke_detail(request, pk):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền thao tác phiếu kiểm kê.')
        return redirect('kiem_ke_list')

    kk = get_object_or_404(KiemKe, pk=pk)
    vi_tri_hang_loi_list = ViTriKho.objects.filter(kho_id=kk.kho_id, trang_thai='hoat_dong').order_by('ma_vi_tri')
    chi_tiet = kk.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh').order_by('hang_hoa__ma_hang')
    for ct in chi_tiet:
        ct.so_luong_tot = max(0, int(ct.so_luong_thuc_te or 0) - int(ct.so_luong_loi or 0))

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()

        def _apply_vi_tri_hang_loi_from_post():
            vi_tri_hang_loi_id = (request.POST.get('vi_tri_hang_loi') or '').strip()
            if not vi_tri_hang_loi_id:
                if kk.vi_tri_hang_loi_id is not None:
                    kk.vi_tri_hang_loi = None
                    kk.save(update_fields=['vi_tri_hang_loi'])
                return True

            vi_tri_obj = vi_tri_hang_loi_list.filter(pk=vi_tri_hang_loi_id).first()
            if not vi_tri_obj:
                messages.error(request, 'Vị trí hàng lỗi không hợp lệ cho kho này.')
                return False

            if kk.vi_tri_hang_loi_id != vi_tri_obj.id:
                kk.vi_tri_hang_loi = vi_tri_obj
                kk.save(update_fields=['vi_tri_hang_loi'])
            return True

        if action == 'save_numbers':
            if kk.trang_thai == '3':
                messages.error(request, 'Không sửa phiếu đã hoàn tất.')
                return redirect('kiem_ke_detail', pk=pk)

            if not _apply_vi_tri_hang_loi_from_post():
                return redirect('kiem_ke_detail', pk=pk)

            ct_ids = request.POST.getlist('ct_id[]')
            sl_thuc_tes = request.POST.getlist('sl_thuc_te[]')
            sl_lois = request.POST.getlist('sl_loi[]')
            tinh_trangs = request.POST.getlist('tinh_trang[]')
            ly_dos = request.POST.getlist('ly_do[]')

            try:
                with transaction.atomic():
                    for i, ct_id in enumerate(ct_ids):
                        if not str(ct_id).isdigit():
                            continue
                        ct = KiemKe_CT.objects.select_for_update().get(pk=int(ct_id), kiem_ke=kk)
                        sl_text = (sl_thuc_tes[i] if i < len(sl_thuc_tes) else '').strip()
                        sl_loi_text = (sl_lois[i] if i < len(sl_lois) else '').strip()
                        sl_value = int(sl_text) if sl_text else 0
                        sl_loi_value = int(sl_loi_text) if sl_loi_text else 0
                        if sl_value < 0:
                            messages.error(request, f'SL thực tế phải >= 0 (mã hàng {ct.hang_hoa.ma_hang}).')
                            return redirect('kiem_ke_detail', pk=pk)
                        if sl_loi_value < 0:
                            messages.error(request, f'SL hàng lỗi phải >= 0 (mã hàng {ct.hang_hoa.ma_hang}).')
                            return redirect('kiem_ke_detail', pk=pk)
                        if sl_loi_value > sl_value:
                            messages.error(request, f'SL hàng lỗi không được lớn hơn SL thực tế (mã hàng {ct.hang_hoa.ma_hang}).')
                            return redirect('kiem_ke_detail', pk=pk)

                        tinh_trang = tinh_trangs[i] if i < len(tinh_trangs) else 'tot_100'
                        if tinh_trang not in dict(KiemKe_CT.TINH_TRANG):
                            tinh_trang = 'tot_100'

                        ct.so_luong_thuc_te = sl_value
                        ct.so_luong_loi = sl_loi_value
                        ct.tinh_trang = tinh_trang
                        ct.ly_do = (ly_dos[i] if i < len(ly_dos) else '').strip()
                        ct.save()
            except (TypeError, ValueError):
                messages.error(request, 'Dữ liệu số lượng kiểm kê không hợp lệ.')
                return redirect('kiem_ke_detail', pk=pk)

            messages.success(request, 'Đã cập nhật số liệu kiểm kê thực tế.')
            return redirect('kiem_ke_detail', pk=pk)

        if action == 'confirm_numbers':
            if kk.trang_thai not in ('1', '2'):
                messages.error(request, 'Phiếu đã hoàn tất, không thể xác nhận lại.')
                return redirect('kiem_ke_detail', pk=pk)

            if not _apply_vi_tri_hang_loi_from_post():
                return redirect('kiem_ke_detail', pk=pk)

            has_diff = kk.chi_tiet.filter(chenh_lech__ne=0).exists() if False else any(
                (ct.chenh_lech or 0) != 0 for ct in kk.chi_tiet.all()
            )
            has_hang_loi = any((ct.so_luong_loi or 0) > 0 for ct in kk.chi_tiet.all())
            if has_hang_loi and not kk.vi_tri_hang_loi_id:
                messages.error(request, 'Phiếu có hàng lỗi nhưng chưa chọn vị trí hàng lỗi.')
                return redirect('kiem_ke_detail', pk=pk)
            if has_diff:
                kk.trang_thai = '2'
                kk.save(update_fields=['trang_thai'])
                messages.warning(request, 'Phiếu có chênh lệch. Vui lòng lập phiếu điều chỉnh để xử lý.')
            else:
                kk.trang_thai = '3'
                kk.save(update_fields=['trang_thai'])
                messages.success(request, 'Không có chênh lệch. Phiếu kiểm kê đã hoàn thành.')
            return redirect('kiem_ke_detail', pk=pk)

        if action == 'create_adjustment':
            if kk.trang_thai != '2':
                messages.error(request, 'Chỉ phiếu chờ điều chỉnh mới lập được phiếu điều chỉnh.')
                return redirect('kiem_ke_detail', pk=pk)

            if hasattr(kk, 'phieu_dieu_chinh'):
                messages.error(request, 'Phiếu kiểm kê này đã có phiếu điều chỉnh.')
                return redirect('kiem_ke_detail', pk=pk)

            ly_do_dc = (request.POST.get('ly_do_dieu_chinh') or '').strip()
            if not ly_do_dc:
                messages.error(request, 'Vui lòng nhập lý do điều chỉnh.')
                return redirect('kiem_ke_detail', pk=pk)

            diff_rows = list(kk.chi_tiet.filter(chenh_lech__isnull=False).exclude(chenh_lech=0).select_related('hang_hoa'))
            if not diff_rows:
                messages.error(request, 'Phiếu không có chênh lệch để điều chỉnh.')
                return redirect('kiem_ke_detail', pk=pk)

            with transaction.atomic():
                phieu = PhieuDieuChinhKiemKe.objects.create(
                    so_phieu=_gen_so_phieu_dieu_chinh_kiem_ke(),
                    kiem_ke=kk,
                    ngay_dieu_chinh=date.today(),
                    kho=kk.kho,
                    nguoi_lap=request.user,
                    ly_do=ly_do_dc,
                    trang_thai='1',
                )
                for ct in diff_rows:
                    PhieuDieuChinhKiemKe_CT.objects.create(
                        phieu=phieu,
                        hang_hoa=ct.hang_hoa,
                        so_luong_he_thong=int(ct.so_luong_so_sach or 0),
                        so_luong_thuc_te=int(ct.so_luong_thuc_te or 0),
                        chenh_lech=int(ct.chenh_lech or 0),
                        ly_do=(ct.ly_do or '').strip(),
                    )

            messages.success(request, 'Lập phiếu điều chỉnh thành công. Phiếu đang ở trạng thái Chờ duyệt.')
            return redirect('kiem_ke_dieu_chinh_detail', pk=phieu.pk)

        messages.error(request, 'Thao tác không hợp lệ.')
        return redirect('kiem_ke_detail', pk=pk)

    return render(request, 'kho/kiem_ke_detail.html', {
        'kk': kk,
        'chi_tiet': chi_tiet,
        'vi_tri_hang_loi_list': vi_tri_hang_loi_list,
        'phieu_dieu_chinh': getattr(kk, 'phieu_dieu_chinh', None),
        'page_title': f'Kiểm kê {kk.kho.ten_kho}',
        'active_menu': 'kiem_ke',
    })


@login_required
@transaction.atomic
def kiem_ke_xoa(request, pk):
    if not _can_delete_vouchers(request.user):
        messages.error(request, 'Bạn không có quyền xóa phiếu kiểm kê.')
        return redirect('kiem_ke_list')

    kk = get_object_or_404(KiemKe.objects.select_related('kho'), pk=pk)
    if request.method != 'POST':
        return redirect('kiem_ke_detail', pk=pk)

    phieu_dc = getattr(kk, 'phieu_dieu_chinh', None)
    so_phieu_dc = phieu_dc.so_phieu if phieu_dc else None
    if phieu_dc:
        if _has_kiem_ke_dieu_chinh_references(phieu_dc):
            messages.error(request, f'Phiếu {kk.ma_phieu} đã có phiếu điều chỉnh {so_phieu_dc} phát sinh tham chiếu, không thể xóa.')
            return redirect('kiem_ke_detail', pk=pk)
        phieu_dc.delete()

    ma_phieu = kk.ma_phieu
    kk.delete()
    if so_phieu_dc:
        messages.success(request, f'Đã xóa phiếu kiểm kê {ma_phieu} và phiếu điều chỉnh {so_phieu_dc}.')
    else:
        messages.success(request, f'Đã xóa phiếu kiểm kê {ma_phieu}.')
    return redirect('kiem_ke_list')


@login_required
@transaction.atomic
def kiem_ke_xoa_nhieu(request):
    if not _can_delete_vouchers(request.user):
        messages.error(request, 'Bạn không có quyền xóa phiếu kiểm kê.')
        return redirect('kiem_ke_list')

    if request.method != 'POST':
        return redirect('kiem_ke_list')

    ids = [int(x) for x in request.POST.getlist('ids[]') if str(x).isdigit()]
    if not ids:
        messages.error(request, 'Vui lòng chọn ít nhất 1 phiếu để xóa.')
        return redirect('kiem_ke_list')

    items = list(KiemKe.objects.select_related('kho').filter(pk__in=ids))
    deleted = 0
    deleted_with_dc = 0
    blocked = []

    for item in items:
        phieu_dc = getattr(item, 'phieu_dieu_chinh', None)
        if phieu_dc:
            if _has_kiem_ke_dieu_chinh_references(phieu_dc):
                blocked.append(item.ma_phieu)
                continue
            phieu_dc.delete()
            deleted_with_dc += 1
        item.delete()
        deleted += 1

    if deleted:
        if deleted_with_dc:
            messages.success(request, f'Đã xóa {deleted} phiếu kiểm kê (trong đó {deleted_with_dc} phiếu có kèm phiếu điều chỉnh).')
        else:
            messages.success(request, f'Đã xóa {deleted} phiếu kiểm kê.')

    if not deleted:
        messages.error(request, 'Không có phiếu hợp lệ để xóa.')

    if blocked:
        preview = ', '.join(blocked[:5])
        suffix = '...' if len(blocked) > 5 else ''
        messages.error(request, f'Không thể xóa {len(blocked)} phiếu do đã có tham chiếu: {preview}{suffix}')

    return redirect('kiem_ke_list')


@login_required
def kiem_ke_dieu_chinh_list(request):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền truy cập phiếu điều chỉnh kiểm kê.')
        return redirect('dashboard')

    q = request.GET.get('q', '').strip()
    trang_thai = request.GET.get('trang_thai', '').strip()

    items = PhieuDieuChinhKiemKe.objects.select_related('kiem_ke', 'kho', 'nguoi_lap').order_by('-ngay_tao')
    if q:
        items = items.filter(
            Q(so_phieu__icontains=q)
            | Q(ly_do__icontains=q)
            | Q(kiem_ke__ma_phieu__icontains=q)
            | Q(kho__ten_kho__icontains=q)
        )
    if trang_thai:
        items = items.filter(trang_thai=trang_thai)

    return render(request, 'kho/kiem_ke_dieu_chinh_list.html', {
        'items': items[:30],
        'q': q,
        'trang_thai': trang_thai,
        'page_title': 'Phiếu điều chỉnh kiểm kê',
        'active_menu': 'kiem_ke',
    })


@login_required
def kiem_ke_dieu_chinh_xoa(request, pk):
    if not _can_delete_vouchers(request.user):
        messages.error(request, 'Bạn không có quyền xóa phiếu điều chỉnh kiểm kê.')
        return redirect('dashboard')

    if request.method != 'POST':
        messages.error(request, 'Thao tác không hợp lệ.')
        return redirect('kiem_ke_dieu_chinh_list')

    phieu = get_object_or_404(PhieuDieuChinhKiemKe, pk=pk)
    if _has_kiem_ke_dieu_chinh_references(phieu):
        messages.error(request, f'Phiếu điều chỉnh {phieu.so_phieu} đã phát sinh tham chiếu, không thể xóa.')
        return redirect('kiem_ke_dieu_chinh_detail', pk=pk)
    so_phieu = phieu.so_phieu
    phieu.delete()
    messages.success(request, f'Đã xóa phiếu điều chỉnh {so_phieu}.')
    return redirect('kiem_ke_dieu_chinh_list')


@login_required
def kiem_ke_dieu_chinh_detail(request, pk):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền xử lý phiếu điều chỉnh.')
        return redirect('dashboard')

    phieu = get_object_or_404(PhieuDieuChinhKiemKe.objects.select_related('kiem_ke', 'kho', 'nguoi_lap'), pk=pk)
    chi_tiet = list(phieu.chi_tiet.select_related('hang_hoa').order_by('hang_hoa__ma_hang'))
    sl_loi_by_hang = {
        row.hang_hoa_id: int(row.so_luong_loi or 0)
        for row in phieu.kiem_ke.chi_tiet.all()
    }
    for row in chi_tiet:
        row.so_luong_loi = sl_loi_by_hang.get(row.hang_hoa_id, 0)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'approve':
            if phieu.trang_thai != '1':
                messages.error(request, 'Phiếu điều chỉnh đã được duyệt trước đó.')
                return redirect('kiem_ke_dieu_chinh_detail', pk=pk)

            with transaction.atomic():
                vi_tri = _find_kiem_ke_vi_tri(phieu.kiem_ke)
                if not vi_tri:
                    messages.error(request, 'Không tìm thấy vị trí kiểm kê để duyệt điều chỉnh.')
                    return redirect('kiem_ke_dieu_chinh_detail', pk=pk)
                vi_tri_hang_loi = phieu.kiem_ke.vi_tri_hang_loi

                gia_von_map = {}
                for row in chi_tiet:
                    ton, _ = TonKho.objects.get_or_create(
                        kho=phieu.kho,
                        hang_hoa=row.hang_hoa,
                        defaults={'so_luong': 0, 'so_luong_loi': 0, 'gia_von_tb': 0},
                    )
                    gia_von_map[row.hang_hoa_id] = Decimal(ton.gia_von_tb or 0)

                    ton_vt, _ = TonKhoViTri.objects.select_for_update().get_or_create(
                        kho=phieu.kho,
                        vi_tri=vi_tri,
                        hang_hoa=row.hang_hoa,
                        defaults={'so_luong': 0},
                    )
                    ton_vt.so_luong = int(ton_vt.so_luong or 0) + int(row.chenh_lech or 0)
                    if ton_vt.so_luong < 0:
                        ton_vt.so_luong = 0
                    ton_vt.save(update_fields=['so_luong', 'ngay_cap_nhat'])

                    if int(getattr(row, 'so_luong_loi', 0) or 0) > 0:
                        _move_kiem_ke_hang_loi(
                            phieu.kho_id,
                            vi_tri.pk,
                            vi_tri_hang_loi.pk if vi_tri_hang_loi else None,
                            row.hang_hoa_id,
                            int(row.so_luong_loi or 0),
                        )

                    _sync_tonkho_from_vitri(phieu.kho_id, row.hang_hoa_id)

                tang_rows = [r for r in chi_tiet if int(r.chenh_lech or 0) > 0]
                giam_rows = [r for r in chi_tiet if int(r.chenh_lech or 0) < 0]
                note_tag = f'[KK_DC:{phieu.pk}]'

                if tang_rows:
                    pn = PhieuNhap.objects.create(
                        so_phieu=_gen_so_phieu('NK'),
                        ngay_lap=date.today(),
                        ngay_hach_toan=date.today(),
                        ngay_chung_tu=date.today(),
                        ngay_nhap=date.today(),
                        loai_nhap='2',
                        kho=phieu.kho,
                        tong_tien=0,
                        trang_thai='3',
                        nguoi_tao=request.user,
                        ghi_chu=f'Điều chỉnh tăng tồn từ phiếu {phieu.so_phieu} {note_tag}',
                    )
                    for r in tang_rows:
                        gia_von = gia_von_map.get(r.hang_hoa_id, Decimal('0'))
                        PhieuNhap_CT.objects.create(
                            phieu_nhap=pn,
                            hang_hoa=r.hang_hoa,
                            so_luong_dat=0,
                            so_luong_nhan=int(r.chenh_lech or 0),
                            don_gia=gia_von,
                            chiet_khau=0,
                            thue_vat=0,
                            tk_no='156',
                            tk_co='711',
                        )
                    pn.tinh_tong()
                    try:
                        post_to_ledger('phieu_nhap', pn.id, user=request.user)
                    except LedgerPostingError:
                        pass

                if giam_rows:
                    px = PhieuXuat.objects.create(
                        so_phieu=_gen_so_phieu('PX'),
                        ngay_lap=date.today(),
                        ngay_hach_toan=date.today(),
                        ngay_chung_tu=date.today(),
                        ngay_xuat=date.today(),
                        loai_xuat='hu_hong',
                        kho=phieu.kho,
                        tong_gia_von=0,
                        trang_thai='3',
                        nguoi_tao=request.user,
                        ghi_chu=f'Điều chỉnh giảm tồn từ phiếu {phieu.so_phieu} {note_tag}',
                    )
                    tong_gia_von = Decimal('0')
                    for r in giam_rows:
                        sl_giam = abs(int(r.chenh_lech or 0))
                        gia_von = gia_von_map.get(r.hang_hoa_id, Decimal('0'))
                        tg_von = Decimal(sl_giam) * gia_von
                        PhieuXuat_CT.objects.create(
                            phieu_xuat=px,
                            hang_hoa=r.hang_hoa,
                            so_luong=sl_giam,
                            gia_von=gia_von,
                            tong_gia_von=tg_von,
                            tk_no='632',
                            tk_co='156',
                        )
                        tong_gia_von += tg_von
                    px.tong_gia_von = tong_gia_von
                    px.save(update_fields=['tong_gia_von'])
                    try:
                        post_to_ledger('phieu_xuat', px.id, user=request.user)
                    except LedgerPostingError:
                        pass

                phieu.trang_thai = '2'
                phieu.save(update_fields=['trang_thai'])

                phieu.kiem_ke.trang_thai = '3'
                phieu.kiem_ke.save(update_fields=['trang_thai'])
                try:
                    post_to_ledger('phieu_dieu_chinh_kho', phieu.kiem_ke.id, user=request.user)
                except LedgerPostingError as exc:
                    messages.warning(request, f'Lỗi ghi sổ cái điều chỉnh: {str(exc)}')

            messages.success(request, 'Đã duyệt phiếu điều chỉnh và cập nhật tồn kho thành công.')
            return redirect('kiem_ke_dieu_chinh_detail', pk=pk)

        messages.error(request, 'Thao tác không hợp lệ.')
        return redirect('kiem_ke_dieu_chinh_detail', pk=pk)

    return render(request, 'kho/kiem_ke_dieu_chinh_detail.html', {
        'phieu': phieu,
        'chi_tiet': chi_tiet,
        'page_title': f'Điều chỉnh {phieu.so_phieu}',
        'active_menu': 'kiem_ke',
    })


@login_required
@xframe_options_exempt
def kiem_ke_in(request, pk):
    if not _can_manage_kiem_ke(request.user):
        messages.error(request, 'Bạn không có quyền in phiếu kiểm kê.')
        return redirect('kiem_ke_list')

    kk = get_object_or_404(KiemKe.objects.select_related('kho', 'nhom_hang'), pk=pk)
    chi_tiet = kk.chi_tiet.select_related('hang_hoa', 'hang_hoa__don_vi_tinh').order_by('hang_hoa__ma_hang')
    for ct in chi_tiet:
        ct.so_luong_tot = max(0, int(ct.so_luong_thuc_te or 0) - int(ct.so_luong_loi or 0))
    return render(request, 'kho/kiem_ke_print.html', {
        'kk': kk,
        'chi_tiet': chi_tiet,
        'page_title': f'In phiếu kiểm kê {kk.ma_phieu}',
        'active_menu': 'kiem_ke',
    })


# ─── BÁO CÁO KHO ────────────────────────────────────────────
@login_required
def bao_cao_xuat_nhap_ton(request):
    thang_hien_tai = (request.GET.get('thang') or '').strip()
    kho_id = (request.GET.get('kho') or '').strip()
    q = (request.GET.get('q') or '').strip()

    today = date.today()
    if not thang_hien_tai:
        thang_hien_tai = f'{today.year:04d}-{today.month:02d}'

    try:
        year = int(thang_hien_tai[:4])
        month = int(thang_hien_tai[5:7])
        tu_ngay = date(year, month, 1)
    except Exception:
        year = today.year
        month = today.month
        thang_hien_tai = f'{year:04d}-{month:02d}'
        tu_ngay = date(year, month, 1)

    if month == 12:
        den_ngay = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        den_ngay = date(year, month + 1, 1) - timedelta(days=1)

    nhap_base = PhieuNhap_CT.objects.select_related('phieu_nhap', 'hang_hoa', 'hang_hoa__don_vi_tinh').filter(
        phieu_nhap__trang_thai__in=('2', '3')
    )
    xuat_base = PhieuXuat_CT.objects.select_related('phieu_xuat', 'hang_hoa', 'hang_hoa__don_vi_tinh').filter(
        phieu_xuat__trang_thai__in=('2', '3')
    )

    if kho_id:
        nhap_base = nhap_base.filter(phieu_nhap__kho_id=kho_id)
        xuat_base = xuat_base.filter(phieu_xuat__kho_id=kho_id)
    if q:
        nhap_base = nhap_base.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))
        xuat_base = xuat_base.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))

    dau_ky_nhap = nhap_base.filter(phieu_nhap__ngay_chung_tu__lt=tu_ngay)
    dau_ky_xuat = xuat_base.filter(phieu_xuat__ngay_chung_tu__lt=tu_ngay)

    trong_ky_nhap = nhap_base.filter(phieu_nhap__ngay_chung_tu__range=[tu_ngay, den_ngay])
    trong_ky_xuat = xuat_base.filter(phieu_xuat__ngay_chung_tu__range=[tu_ngay, den_ngay])

    data_map = {}

    def ensure_row(hang_hoa):
        row = data_map.get(hang_hoa.id)
        if row:
            return row
        dvt_name = ''
        if getattr(hang_hoa, 'don_vi_tinh', None):
            dvt_name = getattr(hang_hoa.don_vi_tinh, 'ten', '') or getattr(hang_hoa.don_vi_tinh, 'ten_dvt', '')
        row = {
            'ma_hang': hang_hoa.ma_hang,
            'ten_hang': hang_hoa.ten_hang,
            'dvt': dvt_name or '-',
            'sl_dau_ky': 0,
            'gt_dau_ky': 0,
            'sl_nhap': 0,
            'gt_nhap': 0,
            'sl_xuat': 0,
            'gt_xuat': 0,
            'sl_cuoi_ky': 0,
            'gt_cuoi_ky': 0,
        }
        data_map[hang_hoa.id] = row
        return row

    for ct in dau_ky_nhap:
        row = ensure_row(ct.hang_hoa)
        row['sl_dau_ky'] += int(ct.so_luong_nhan or 0)
        row['gt_dau_ky'] += int(ct.thanh_tien or 0)

    for ct in dau_ky_xuat:
        row = ensure_row(ct.hang_hoa)
        row['sl_dau_ky'] -= int(ct.so_luong or 0)
        row['gt_dau_ky'] -= int(ct.tong_gia_von or 0)

    for ct in trong_ky_nhap:
        row = ensure_row(ct.hang_hoa)
        row['sl_nhap'] += int(ct.so_luong_nhan or 0)
        row['gt_nhap'] += int(ct.thanh_tien or 0)

    for ct in trong_ky_xuat:
        row = ensure_row(ct.hang_hoa)
        row['sl_xuat'] += int(ct.so_luong or 0)
        row['gt_xuat'] += int(ct.tong_gia_von or 0)

    data = []
    total_sl_dau_ky = 0
    total_sl_nhap = 0
    total_sl_xuat = 0
    total_sl_cuoi_ky = 0
    total_gt_cuoi_ky = 0
    
    for row in data_map.values():
        row['sl_cuoi_ky'] = row['sl_dau_ky'] + row['sl_nhap'] - row['sl_xuat']
        row['gt_cuoi_ky'] = row['gt_dau_ky'] + row['gt_nhap'] - row['gt_xuat']
        data.append(row)
        
        total_sl_dau_ky += int(row['sl_dau_ky'] or 0)
        total_sl_nhap += int(row['sl_nhap'] or 0)
        total_sl_xuat += int(row['sl_xuat'] or 0)
        total_sl_cuoi_ky += int(row['sl_cuoi_ky'] or 0)
        total_gt_cuoi_ky += int(row['gt_cuoi_ky'] or 0)

    data.sort(key=lambda x: x['ma_hang'])

    return render(request, 'kho/bao_cao_ton_kho.html', {
        'data': data,
        'total_sl_dau_ky': total_sl_dau_ky,
        'total_sl_nhap': total_sl_nhap,
        'total_sl_xuat': total_sl_xuat,
        'total_sl_cuoi_ky': total_sl_cuoi_ky,
        'total_gt_cuoi_ky': total_gt_cuoi_ky,
        'thang_hien_tai': thang_hien_tai,
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'kho_filter': kho_id,
        'kho_da_chon': Kho.objects.filter(pk=kho_id).values_list('ma_kho', flat=True).first() if kho_id else '',
        'q': q,
        'page_title': 'Báo cáo xuất nhập tồn',
        'active_menu': 'bao_cao_xnt',
    })


@login_required
def bao_cao_ton_hien_tai(request):
    kho_id = (request.GET.get('kho') or '').strip()
    q = (request.GET.get('q') or '').strip()

    _resync_tonkho_from_vitri_scope(kho_id or None)

    items = TonKho.objects.select_related('hang_hoa', 'hang_hoa__don_vi_tinh', 'kho').filter(so_luong__gt=0)
    if kho_id:
        items = items.filter(kho_id=kho_id)
    if q:
        items = items.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))

    rows = []
    tong_so_luong = 0
    tong_gia_tri = Decimal('0')
    tong_gia_tri_ban = Decimal('0')
    gia_ban_cache = {}
    for tk in items.order_by('hang_hoa__ma_hang'):
        dvt_name = ''
        if getattr(tk.hang_hoa, 'don_vi_tinh', None):
            dvt_name = getattr(tk.hang_hoa.don_vi_tinh, 'ten', '') or getattr(tk.hang_hoa.don_vi_tinh, 'ten_dvt', '')
        gia_tri = Decimal(int(tk.so_luong or 0)) * Decimal(tk.gia_von_tb or 0)
        if tk.hang_hoa_id not in gia_ban_cache:
            gia_ban_cache[tk.hang_hoa_id] = Decimal(getattr(tk.hang_hoa, 'gia_ban', 0) or 0)
        gia_ban_hien_tai = gia_ban_cache[tk.hang_hoa_id]
        gia_tri_ban = Decimal(int(tk.so_luong or 0)) * gia_ban_hien_tai
        tong_so_luong += int(tk.so_luong or 0)
        tong_gia_tri += gia_tri
        tong_gia_tri_ban += gia_tri_ban
        rows.append({
            'ma_hang': tk.hang_hoa.ma_hang,
            'ten_hang': tk.hang_hoa.ten_hang,
            'dvt': dvt_name or '-',
            'ma_kho': tk.kho.ma_kho,
            'so_luong': int(tk.so_luong or 0),
            'gia_von_tb': Decimal(tk.gia_von_tb or 0),
            'gia_ban_hien_tai': gia_ban_hien_tai,
            'gia_tri_ton': gia_tri,
            'gia_tri_ban': gia_tri_ban,
        })

    return render(request, 'kho/bao_cao_ton_hien_tai.html', {
        'rows': rows,
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'kho_filter': kho_id,
        'kho_da_chon': Kho.objects.filter(pk=kho_id).values_list('ma_kho', flat=True).first() if kho_id else '',
        'q': q,
        'tong_so_luong': tong_so_luong,
        'tong_gia_tri': tong_gia_tri,
        'tong_gia_tri_ban': tong_gia_tri_ban,
        'page_title': 'Báo cáo tồn kho hiện tại',
        'active_menu': 'bao_cao_ton_hien_tai',
    })


@login_required
def bao_cao_ton_kho(request):
    kho_id = request.GET.get('kho', '').strip()
    q = request.GET.get('q', '').strip()
    tu_ngay = request.GET.get('tu_ngay', '').strip()
    den_ngay = request.GET.get('den_ngay', '').strip()

    nhap_qs = PhieuNhap_CT.objects.select_related('phieu_nhap', 'hang_hoa', 'phieu_nhap__kho').filter(
        phieu_nhap__trang_thai__in=('2', '3')
    )
    xuat_qs = PhieuXuat_CT.objects.select_related('phieu_xuat', 'hang_hoa', 'phieu_xuat__kho').filter(
        phieu_xuat__trang_thai__in=('2', '3')
    )

    if kho_id:
        nhap_qs = nhap_qs.filter(phieu_nhap__kho_id=kho_id)
        xuat_qs = xuat_qs.filter(phieu_xuat__kho_id=kho_id)
    if q:
        nhap_qs = nhap_qs.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))
        xuat_qs = xuat_qs.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))

    if tu_ngay:
        dau_ky_nhap = nhap_qs.filter(phieu_nhap__ngay_chung_tu__lt=tu_ngay)
        dau_ky_xuat = xuat_qs.filter(phieu_xuat__ngay_chung_tu__lt=tu_ngay)
    else:
        dau_ky_nhap = nhap_qs.none()
        dau_ky_xuat = xuat_qs.none()

    sl_dau_ky_nhap = sum((i.so_luong_nhan or 0) for i in dau_ky_nhap)
    sl_dau_ky_xuat = sum((i.so_luong or 0) for i in dau_ky_xuat)
    gt_dau_ky_nhap = sum(int(i.thanh_tien or 0) for i in dau_ky_nhap)
    gt_dau_ky_xuat = sum(int(i.tong_gia_von or 0) for i in dau_ky_xuat)

    sl_dau_ky = sl_dau_ky_nhap - sl_dau_ky_xuat
    gt_dau_ky = gt_dau_ky_nhap - gt_dau_ky_xuat

    trong_ky_nhap = nhap_qs
    trong_ky_xuat = xuat_qs
    if tu_ngay and den_ngay:
        trong_ky_nhap = trong_ky_nhap.filter(phieu_nhap__ngay_chung_tu__range=[tu_ngay, den_ngay])
        trong_ky_xuat = trong_ky_xuat.filter(phieu_xuat__ngay_chung_tu__range=[tu_ngay, den_ngay])
    elif tu_ngay:
        trong_ky_nhap = trong_ky_nhap.filter(phieu_nhap__ngay_chung_tu__gte=tu_ngay)
        trong_ky_xuat = trong_ky_xuat.filter(phieu_xuat__ngay_chung_tu__gte=tu_ngay)
    elif den_ngay:
        trong_ky_nhap = trong_ky_nhap.filter(phieu_nhap__ngay_chung_tu__lte=den_ngay)
        trong_ky_xuat = trong_ky_xuat.filter(phieu_xuat__ngay_chung_tu__lte=den_ngay)

    trong_ky_nhap = trong_ky_nhap.order_by('phieu_nhap__ngay_chung_tu', 'phieu_nhap__so_phieu', 'id')
    trong_ky_xuat = trong_ky_xuat.order_by('phieu_xuat__ngay_chung_tu', 'phieu_xuat__so_phieu', 'id')

    sl_nhap_ky = sum((i.so_luong_nhan or 0) for i in trong_ky_nhap)
    gt_nhap_ky = sum(int(i.thanh_tien or 0) for i in trong_ky_nhap)
    sl_xuat_ky = sum((i.so_luong or 0) for i in trong_ky_xuat)
    gt_xuat_ky = sum(int(i.tong_gia_von or 0) for i in trong_ky_xuat)

    sl_cuoi_ky = sl_dau_ky + sl_nhap_ky - sl_xuat_ky
    gt_cuoi_ky = gt_dau_ky + gt_nhap_ky - gt_xuat_ky

    rows = []
    for i in trong_ky_nhap:
        rows.append({
            'sort_key': (i.phieu_nhap.ngay_chung_tu, i.phieu_nhap.so_phieu, 0, i.id),
            'loai_dong': 'nhap',
            'doi_tuong_id': i.phieu_nhap_id,
            'action_url': f'/kho/nhap/{i.phieu_nhap_id}/',
            'ngay_ghi_so': i.phieu_nhap.ngay_chung_tu,
            'ngay_ct': i.phieu_nhap.ngay_chung_tu,
            'ma_ct': 'PN',
            'so_ct': i.phieu_nhap.so_phieu,
            'dien_giai': i.phieu_nhap.ghi_chu or '',
            'ma_nhap_xuat': i.tk_no or '',
            'tk_doi_ung': i.tk_co or '',
            'gia': int(i.don_gia or 0),
            'sl_nhap': i.so_luong_nhan or 0,
            'tien_nhap': int(i.thanh_tien or 0),
            'sl_xuat': 0,
            'tien_xuat': 0,
        })
    for i in trong_ky_xuat:
        rows.append({
            'sort_key': (i.phieu_xuat.ngay_chung_tu, i.phieu_xuat.so_phieu, 1, i.id),
            'loai_dong': 'xuat',
            'doi_tuong_id': i.phieu_xuat_id,
            'action_url': f'/kho/xuat/{i.phieu_xuat_id}/',
            'ngay_ghi_so': i.phieu_xuat.ngay_chung_tu,
            'ngay_ct': i.phieu_xuat.ngay_chung_tu,
            'ma_ct': 'HD',
            'so_ct': i.phieu_xuat.so_phieu,
            'dien_giai': i.phieu_xuat.ghi_chu or '',
            'ma_nhap_xuat': i.tk_co or '156',
            'tk_doi_ung': i.tk_no or '632',
            'gia': int(i.gia_von or 0),
            'sl_nhap': 0,
            'tien_nhap': 0,
            'sl_xuat': i.so_luong or 0,
            'tien_xuat': int(i.tong_gia_von or 0),
        })

    rows.sort(key=lambda x: x['sort_key'])

    sl_ton_chay = sl_dau_ky
    gt_ton_chay = gt_dau_ky
    for r in rows:
        sl_ton_chay = sl_ton_chay + r['sl_nhap'] - r['sl_xuat']
        gt_ton_chay = gt_ton_chay + r['tien_nhap'] - r['tien_xuat']
        r['sl_ton'] = sl_ton_chay
        r['gt_ton'] = gt_ton_chay

    kho_da_chon = ''
    if kho_id:
        kho_obj = Kho.objects.filter(pk=kho_id).first()
        if kho_obj:
            kho_da_chon = f'{kho_obj.ma_kho} - {kho_obj.ten_kho}'

    return render(request, 'kho/bao_cao_ton.html', {
        'rows': rows,
        'kho_list': Kho.objects.filter(trang_thai=True),
        'hang_hoa_list': HangHoa.objects.filter(trang_thai='dang_ban').order_by('ma_hang').values('ma_hang', 'ten_hang')[:1000],
        'kho_filter': kho_id,
        'kho_da_chon': kho_da_chon,
        'q': q,
        'tu_ngay': tu_ngay,
        'den_ngay': den_ngay,
        'sl_dau_ky': int(sl_dau_ky),
        'gt_dau_ky': int(gt_dau_ky),
        'sl_nhap_ky': int(sl_nhap_ky),
        'gt_nhap_ky': int(gt_nhap_ky),
        'sl_xuat_ky': int(sl_xuat_ky),
        'gt_xuat_ky': int(gt_xuat_ky),
        'sl_cuoi_ky': int(sl_cuoi_ky),
        'gt_cuoi_ky': int(gt_cuoi_ky),
        'page_title': 'Sổ chi tiết hàng hóa',
        'active_menu': 'so_chi_tiet_hang_hoa',
    })


@login_required
def bao_cao_ton_kho_loi(request):
    """Báo cáo tồn kho hàng lỗi"""
    kho_id = (request.GET.get('kho') or '').strip()
    q = (request.GET.get('q') or '').strip()

    items = TonKho.objects.select_related('hang_hoa', 'hang_hoa__don_vi_tinh', 'kho').filter(so_luong_loi__gt=0)
    if kho_id:
        items = items.filter(kho_id=kho_id)
    if q:
        items = items.filter(Q(hang_hoa__ma_hang__icontains=q) | Q(hang_hoa__ten_hang__icontains=q))

    rows = []
    tong_so_luong_loi = 0
    for tk in items.order_by('hang_hoa__ma_hang'):
        dvt_name = ''
        if getattr(tk.hang_hoa, 'don_vi_tinh', None):
            dvt_name = getattr(tk.hang_hoa.don_vi_tinh, 'ten', '') or getattr(tk.hang_hoa.don_vi_tinh, 'ten_dvt', '')
        tong_so_luong_loi += int(tk.so_luong_loi or 0)
        rows.append({
            'ma_hang': tk.hang_hoa.ma_hang,
            'ten_hang': tk.hang_hoa.ten_hang,
            'dvt': dvt_name or '-',
            'ma_kho': tk.kho.ma_kho,
            'so_luong_loi': int(tk.so_luong_loi or 0),
        })

    return render(request, 'kho/bao_cao_ton_kho_loi.html', {
        'rows': rows,
        'kho_list': Kho.objects.filter(trang_thai=True).order_by('ma_kho'),
        'kho_filter': kho_id,
        'kho_da_chon': Kho.objects.filter(pk=kho_id).values_list('ma_kho', flat=True).first() if kho_id else '',
        'q': q,
        'tong_so_luong_loi': tong_so_luong_loi,
        'page_title': 'Báo cáo tồn kho hàng lỗi',
    })


_period_guard_fallbacks = {
    'phieu_nhap_them': 'phieu_nhap_list',
    'phieu_nhap_sua': 'phieu_nhap_detail',
    'phieu_xuat_them': 'phieu_xuat_list',
    'phieu_xuat_sua': 'phieu_xuat_detail',
    'kiem_ke_them': 'kiem_ke_list',
}

for _view_name, _fallback in _period_guard_fallbacks.items():
    if _view_name in globals():
        globals()[_view_name] = guard_accounting_period_error(_fallback)(globals()[_view_name])
